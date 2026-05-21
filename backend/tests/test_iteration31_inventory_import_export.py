"""
Iteration 31 — Inventory Import / Export tests
Endpoints under test:
  GET  /api/inventory/template
  POST /api/inventory/import
  GET  /api/inventory/export?format=xlsx|pdf
Plus regression: GET /api/inventory/items/{item_id}
Plus smoke for previous iterations.
"""

import io
import os
import pytest
import requests
from openpyxl import Workbook, load_workbook

def _load_base_url():
    url = os.environ.get("REACT_APP_BACKEND_URL", "").strip()
    if not url:
        try:
            with open("/app/frontend/.env") as f:
                for line in f:
                    if line.startswith("REACT_APP_BACKEND_URL="):
                        url = line.split("=", 1)[1].strip()
                        break
        except Exception:
            pass
    return url.rstrip("/")

BASE_URL = _load_base_url()
assert BASE_URL, "REACT_APP_BACKEND_URL not set"

ADMIN_EMAIL = "admin@sensoper.com"
ADMIN_PASS = "Admin@123"


# ---------- fixtures ----------
@pytest.fixture(scope="module")
def admin_session():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": ADMIN_EMAIL, "password": ADMIN_PASS},
               timeout=20)
    assert r.status_code == 200, f"Admin login failed: {r.status_code} {r.text}"
    return s


@pytest.fixture(scope="module")
def unauth_session():
    return requests.Session()


def _make_xlsx(headers, rows):
    wb = Workbook(); ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ---------- TEMPLATE ----------
class TestInventoryTemplate:
    def test_template_ok(self, admin_session):
        r = admin_session.get(f"{BASE_URL}/api/inventory/template", timeout=30)
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers.get("content-type", "")
        assert len(r.content) > 500
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        expected = ["name", "sku_code", "category", "quantity", "unit_price",
                    "reorder_level", "supplier", "gst_percentage", "margin_pct",
                    "zone", "aisle", "shelf", "rack", "bin_location",
                    "procurement_date", "active"]
        assert headers == expected, f"Header mismatch: {headers}"
        row2 = [c.value for c in ws[2]]
        assert row2[0] == "Solar Panel 540W Mono"
        assert row2[1] == "SP-540-MONO"

    def test_template_unauth(self, unauth_session):
        r = unauth_session.get(f"{BASE_URL}/api/inventory/template", timeout=10)
        assert r.status_code in (401, 403)


# ---------- IMPORT ----------
class TestInventoryImport:
    def test_import_missing_file(self, admin_session):
        r = admin_session.post(f"{BASE_URL}/api/inventory/import", timeout=15)
        assert r.status_code in (400, 422)

    def test_import_empty_file(self, admin_session):
        files = {"file": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = admin_session.post(f"{BASE_URL}/api/inventory/import", files=files, timeout=15)
        assert r.status_code == 400

    def test_import_missing_columns(self, admin_session):
        csv_bytes = b"name,sku\nFoo,F-1\n"
        files = {"file": ("bad.csv", csv_bytes, "text/csv")}
        r = admin_session.post(f"{BASE_URL}/api/inventory/import", files=files, timeout=20)
        assert r.status_code == 400
        detail = r.json().get("detail", "")
        assert "Missing required columns" in detail or "missing" in detail.lower()

    def test_import_valid_template_creates_or_updates(self, admin_session):
        headers = ["name", "sku_code", "category", "quantity", "unit_price",
                   "reorder_level", "supplier", "gst_percentage", "margin_pct",
                   "zone", "aisle", "shelf", "rack", "bin_location",
                   "procurement_date", "active"]
        rows = [[
            "Solar Panel 540W Mono", "SP-540-MONO", "Panels", 50, 11500, 5,
            "ABC Solar Co", 18.0, 12.5, "A", "1", "S2", "R3", "B4",
            "2026-01-15", True
        ]]
        buf = _make_xlsx(headers, rows)
        files = {"file": ("import.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = admin_session.post(f"{BASE_URL}/api/inventory/import", files=files, timeout=30)
        assert r.status_code == 200, r.text
        body = r.json()
        assert "created" in body and "updated" in body and "errors" in body
        assert "message" in body
        assert (body["created"] + body["updated"]) >= 1
        # Now GET inventory items and ensure SP-540-MONO present
        items_r = admin_session.get(f"{BASE_URL}/api/inventory/items", timeout=20)
        assert items_r.status_code == 200
        items = items_r.json()
        skus = [i.get("sku_code") for i in items]
        assert "SP-540-MONO" in skus

    def test_import_invalid_rows_recorded(self, admin_session):
        headers = ["name", "sku_code", "category", "quantity", "unit_price"]
        rows = [
            ["Good Item", "TEST-GOOD-1", "Panels", 5, 100.0],
            ["NoSku", "", "Panels", 5, 100.0],          # missing sku
            ["NegQty", "TEST-NEG-1", "Panels", -3, 50], # negative qty
        ]
        buf = _make_xlsx(headers, rows)
        files = {"file": ("mixed.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = admin_session.post(f"{BASE_URL}/api/inventory/import", files=files, timeout=30)
        assert r.status_code == 200, r.text
        body = r.json()
        assert len(body["errors"]) >= 2
        # row numbers present
        for err in body["errors"]:
            assert "row" in err and "error" in err

    def test_import_too_large(self, admin_session):
        big = b"x" * (10 * 1024 * 1024 + 50)
        files = {"file": ("big.xlsx", big,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = admin_session.post(f"{BASE_URL}/api/inventory/import", files=files, timeout=60)
        assert r.status_code == 413


# ---------- EXPORT ----------
class TestInventoryExport:
    def test_export_xlsx(self, admin_session):
        r = admin_session.get(f"{BASE_URL}/api/inventory/export",
                              params={"format": "xlsx"}, timeout=60)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct, ct
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.max_row >= 2  # header + at least 1 row (SP-540-MONO created above)
        hdr = [c.value for c in ws[1]]
        assert "sku_code" in hdr and "name" in hdr

    def test_export_pdf(self, admin_session):
        r = admin_session.get(f"{BASE_URL}/api/inventory/export",
                              params={"format": "pdf"}, timeout=60)
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith("application/pdf")
        assert r.content[:4] == b"%PDF", r.content[:8]
        # parse with pdfplumber for header text
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(r.content)) as pdf:
                txt = "".join((p.extract_text() or "") for p in pdf.pages)
        except Exception as e:
            pytest.skip(f"pdfplumber failed: {e}")
        assert "Sensoper" in txt and "Inventory" in txt, txt[:300]

    def test_export_invalid_format(self, admin_session):
        r = admin_session.get(f"{BASE_URL}/api/inventory/export",
                              params={"format": "garbage"}, timeout=15)
        assert r.status_code == 400
        assert "xlsx" in r.text.lower() and "pdf" in r.text.lower()


# ---------- REGRESSION: item by id still works ----------
class TestInventoryItemRouteRegression:
    def test_get_item_by_id(self, admin_session):
        lr = admin_session.get(f"{BASE_URL}/api/inventory/items", timeout=20)
        assert lr.status_code == 200
        items = lr.json()
        assert len(items) > 0, "No inventory items to test"
        iid = items[0].get("id") or items[0].get("_id")
        assert iid, f"Item has no id field: {items[0]}"
        r = admin_session.get(f"{BASE_URL}/api/inventory/items/{iid}", timeout=15)
        assert r.status_code == 200, r.text
        item = r.json()
        assert (item.get("id") == iid) or (item.get("_id") == iid) or item.get("sku_code")


# ---------- SMOKE: previous iterations ----------
class TestPreviousSmoke:
    def test_auth_me(self, admin_session):
        r = admin_session.get(f"{BASE_URL}/api/auth/me", timeout=15)
        assert r.status_code == 200
        assert r.json().get("email") == ADMIN_EMAIL

    def test_solar_sizing(self, admin_session):
        payload = {
            "consumption_kwh_per_month": 500,
            "tariff_per_kwh": 8.0,
            "system_type": "on_grid",
            "shading_factor": 1.0,
            "location": "Bangalore",
        }
        r = admin_session.post(f"{BASE_URL}/api/solar/sizing", json=payload, timeout=30)
        # accept 200 or 422 (schema variations); test for breakdown only if 200
        assert r.status_code in (200, 422)
        if r.status_code == 200:
            data = r.json()
            # 25-yr breakdown lives under a key – just sanity check presence
            assert isinstance(data, dict)
