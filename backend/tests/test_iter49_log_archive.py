"""Iteration 49 §5 — audit-log quarterly archive & retention: archive-before-purge ordering.
Seeds synthetic logs into an old quarter (2024-Q1, well outside the retention window) via Mongo directly,
then drives the HTTP API: purge without archive → 409; failed archive → purge still 409; successful
archive → purge allowed; deletion-snapshot entries counted in the archive; archive downloadable after purge."""
import os, uuid, pytest, requests
from datetime import datetime, timezone
from pymongo import MongoClient

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
API = f"{BASE_URL}/api"
Q = "2024-Q1"
TAG = f"TEST49ARC_{uuid.uuid4().hex[:5]}"


def _env(key):
    for line in open("/app/backend/.env"):
        if line.startswith(key + "="):
            return line.split("=", 1)[1].strip().strip('"')
    return None


@pytest.fixture(scope="module")
def dbc():
    return MongoClient(_env("MONGO_URL"))[_env("DB_NAME")]


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": "admin@sensoper.com", "password": "Admin@123"}, timeout=60)
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="module")
def seeded(dbc, client):
    # clean any prior state for this quarter
    dbc.audit_logs.delete_many({"timestamp": {"$gte": "2024-01-01", "$lt": "2024-04-01"}})
    dbc.log_archives.delete_many({"quarter": Q})
    rows = []
    for i in range(5):
        rows.append({"user_id": "t", "user_name": TAG, "action_type": "update", "entity_type": "project", "entity_id": f"p{i}",
                     "details": {"i": i}, "timestamp": f"2024-02-1{i}T10:00:00+00:00"})
    rows.append({"user_id": "t", "user_name": TAG, "action_type": "hard_delete", "entity_type": "sale", "entity_id": "s1",
                 "details": {"snapshot": {"amount": 999}, "reason": "test"}, "timestamp": "2024-03-01T10:00:00+00:00"})
    dbc.audit_logs.insert_many(rows)
    yield len(rows)
    dbc.audit_logs.delete_many({"user_name": TAG})
    dbc.log_archives.delete_many({"quarter": Q})


def test_purge_refused_without_archive(client, seeded):
    r = client.post(f"{API}/audit-logs/archives/{Q}/purge", timeout=60)
    assert r.status_code == 409 and "archive first" in r.text.lower()


def test_failed_archive_blocks_purge(client, seeded, dbc):
    r = client.post(f"{API}/audit-logs/archives/run", params={"quarter": Q, "fail_for_test": "true"}, timeout=60)
    assert r.status_code == 200 and r.json()["status"] == "failed" and r.json()["error"]
    r = client.post(f"{API}/audit-logs/archives/{Q}/purge", timeout=60)
    assert r.status_code == 409
    assert dbc.audit_logs.count_documents({"user_name": TAG}) == seeded, "nothing may be purged after a failed archive"


def test_successful_archive_then_purge(client, seeded, dbc):
    r = client.post(f"{API}/audit-logs/archives/run", params={"quarter": Q}, timeout=120)
    assert r.status_code == 200, r.text
    a = r.json()
    assert a["status"] == "archived" and a["row_count"] == seeded and a["deletion_snapshot_count"] == 1
    assert a["xlsx_path"] and a["pdf_path"]
    # archive is retrievable BEFORE purge
    x = client.get(f"{API}/audit-logs/archives/{a['id']}/download", params={"format": "xlsx"}, timeout=60)
    assert x.status_code == 200 and len(x.content) > 1000
    p = client.get(f"{API}/audit-logs/archives/{a['id']}/download", params={"format": "pdf"}, timeout=60)
    assert p.status_code == 200 and p.content[:4] == b"%PDF"
    # deletion snapshot is inside the archive
    from openpyxl import load_workbook
    import io
    ws = load_workbook(io.BytesIO(x.content)).active
    cells = [str(c.value) for row in ws.iter_rows() for c in row]
    assert any("hard_delete" in c for c in cells) and any("999" in c for c in cells)
    # now purge
    r = client.post(f"{API}/audit-logs/archives/{Q}/purge", timeout=60)
    assert r.status_code == 200, r.text
    assert r.json()["purged_rows"] == seeded
    assert dbc.audit_logs.count_documents({"user_name": TAG}) == 0
    # archive still downloadable after purge (permanent record)
    x2 = client.get(f"{API}/audit-logs/archives/{a['id']}/download", params={"format": "xlsx"}, timeout=60)
    assert x2.status_code == 200 and x2.content == x.content
    lst = client.get(f"{API}/audit-logs/archives", timeout=60).json()
    mine = [d for d in lst if d["quarter"] == Q and d["status"] == "archived"][0]
    assert mine["purged_at"] and mine["purged_rows"] == seeded


def test_recent_quarter_never_purgeable(client):
    now = datetime.now(timezone.utc)
    cur = f"{now.year}-Q{(now.month - 1) // 3 + 1}"
    r = client.post(f"{API}/audit-logs/archives/{cur}/purge", timeout=60)
    assert r.status_code == 409 and "retention window" in r.text


def test_config_roundtrip_and_guard(client):
    r = client.put(f"{API}/audit-logs/archive-config", json={"keep_quarters_live": 0, "auto_archive": True, "auto_purge": False}, timeout=60)
    assert r.status_code == 400
    r = client.put(f"{API}/audit-logs/archive-config", json={"keep_quarters_live": 3, "auto_archive": True, "auto_purge": False}, timeout=60)
    assert r.status_code == 200 and r.json()["keep_quarters_live"] == 3
    client.put(f"{API}/audit-logs/archive-config", json={"keep_quarters_live": 2, "auto_archive": True, "auto_purge": False}, timeout=60)


def test_run_due_is_idempotent(client):
    r1 = client.post(f"{API}/audit-logs/archives/run-due", timeout=120)
    r2 = client.post(f"{API}/audit-logs/archives/run-due", timeout=120)
    assert r1.status_code == 200 and r2.status_code == 200
    q = r1.json()["checked_quarter"]
    lst = [d for d in client.get(f"{API}/audit-logs/archives", timeout=60).json() if d["quarter"] == q and d["status"] == "archived"]
    assert len(lst) == 1, "scheduler must not create duplicate archives"
