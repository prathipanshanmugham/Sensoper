"""Iter-36 backend tests:
 - Inventory items hsn_code (create / list / update)
 - Inventory template includes hsn_code column after gst_percentage
 - Project installation_date / commissioning_date persist
 - till_date metrics in reference-candidates and reference-summary
 - notes-only PUT on completed projects still works
"""
import os
import io
import pytest
import requests

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().rstrip("/")

ADMIN_EMAIL = "admin@sensoper.com"
ADMIN_PASS = "Admin@123"


@pytest.fixture(scope="module")
def auth_headers():
    """Returns a requests.Session() with cookie-based auth set after login."""
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": ADMIN_EMAIL, "password": ADMIN_PASS})
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    token = r.json().get("access_token") or r.json().get("token")
    if token:
        s.headers.update({"Authorization": f"Bearer {token}"})
    return s


# ===================== Inventory HSN =====================
class TestInventoryHSN:
    def test_create_item_with_hsn(self, auth_headers):
        payload = {
            "sku_code": "TEST_HSN_NEW_36X",
            "name": "Test HSN Item",
            "category": "consumable",
            "quantity": 5,
            "unit_price": 100.0,
            "gst_percentage": 18.0,
            "hsn_code": "85414011",
        }
        r = auth_headers.post(f"{BASE_URL}/api/inventory/items",
                          json=payload)
        assert r.status_code in (200, 201), r.text
        data = r.json()
        item_id = data.get("id")
        assert item_id, f"missing id: {data}"

        # List endpoint should return it
        r2 = auth_headers.get(f"{BASE_URL}/api/inventory/items",
                          params={"q": "TEST_HSN_NEW_36X"})
        assert r2.status_code == 200
        items = r2.json() if isinstance(r2.json(), list) else r2.json().get("items", [])
        match = [i for i in items if (i.get("sku_code") or i.get("sku")) == "TEST_HSN_NEW_36X"]
        assert match, "created item not in list"
        assert match[0].get("hsn_code") == "85414011"

        # Update HSN
        r3 = auth_headers.put(f"{BASE_URL}/api/inventory/items/{item_id}",
                          json={"hsn_code": "85044090"})
        assert r3.status_code in (200, 201), r3.text

        # Verify update via GET (list)
        r4 = auth_headers.get(f"{BASE_URL}/api/inventory/items",
                              params={"q": "TEST_HSN_NEW_36X"})
        items = r4.json() if isinstance(r4.json(), list) else r4.json().get("items", [])
        m2 = [i for i in items if (i.get("sku_code") or i.get("sku")) == "TEST_HSN_NEW_36X"]
        assert m2 and m2[0].get("hsn_code") == "85044090", \
            f"updated hsn not persisted: {m2}"

        # Cleanup
        auth_headers.delete(f"{BASE_URL}/api/inventory/items/{item_id}",
                        )

    def test_seeded_hsn_item(self, auth_headers):
        # Seeded SKU HSN-TEST-001 per problem statement
        r = auth_headers.get(f"{BASE_URL}/api/inventory/items",
                         params={"q": "HSN-TEST-001"})
        assert r.status_code == 200
        items = r.json() if isinstance(r.json(), list) else r.json().get("items", [])
        match = [i for i in items if (i.get("sku_code") or i.get("sku")) == "HSN-TEST-001"]
        if match:
            assert match[0].get("hsn_code") == "85414011"


# ===================== Inventory Template =====================
class TestInventoryTemplate:
    def test_template_has_hsn_column(self, auth_headers):
        r = auth_headers.get(f"{BASE_URL}/api/inventory/template",
                         )
        assert r.status_code == 200, r.text
        # Try to read xlsx
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(r.content))
            ws = wb.active
            headers = [c.value for c in ws[1]]
            assert "hsn_code" in headers, f"hsn_code missing from template headers: {headers}"
            # Should be after gst_percentage (9th col = index 8)
            gst_idx = headers.index("gst_percentage")
            hsn_idx = headers.index("hsn_code")
            assert hsn_idx == gst_idx + 1, \
                f"hsn_code at {hsn_idx} not directly after gst_percentage at {gst_idx}"
        except ImportError:
            pytest.skip("openpyxl not available")


# ===================== Project install/commission dates =====================
class TestProjectDates:
    @pytest.fixture
    def created_project_id(self, auth_headers):
        payload = {
            "customer": {"name": "TEST_Iter36_DateProject", "phone": "9999900000", "email": "t@x.com", "address": "Test Addr"},
            "location": {"address": "Chennai", "city": "Chennai", "state": "TN", "pincode": "600001"},
            "electrical": {"connection_type": "single_phase", "sanction_load_kw": 5, "connected_load_kw": 4, "eb_tariff": 7, "monthly_consumption_units": 500, "monthly_bill": 3500, "tariff_per_unit": 7},
            "solar_system": {"system_type": "on_grid", "panel_type": "mono_perc", "panel_wattage": 540, "inverter_type": "string", "battery_required": False},
            "mounting": {"type": "rooftop", "roof_type": "rcc", "tilt_angle": 15, "structure_type": "standard"},
            "additional": {"cable_length_meters": 50, "inverter_to_panel_distance": 10},
            "selected_items": [], "manual_costs": [], "site_images": [], "custom_fields": {},
            "installation_date": "2024-01-15",
            "commissioning_date": "2024-02-01",
        }
        r = auth_headers.post(f"{BASE_URL}/api/projects",
                          json=payload)
        assert r.status_code in (200, 201), r.text
        pid = r.json().get("id")
        yield pid
        # Cleanup
        try:
            auth_headers.delete(f"{BASE_URL}/api/projects/{pid}")
        except Exception:
            pass

    def test_dates_persist_on_create(self, auth_headers, created_project_id):
        r = auth_headers.get(f"{BASE_URL}/api/projects/{created_project_id}",
                         )
        assert r.status_code == 200
        data = r.json()
        assert data.get("installation_date") == "2024-01-15"
        assert data.get("commissioning_date") == "2024-02-01"

    def test_dates_update_via_put(self, auth_headers, created_project_id):
        r = auth_headers.put(f"{BASE_URL}/api/projects/{created_project_id}",
                         json={"installation_date": "2024-03-10",
                               "commissioning_date": "2024-04-01"},
                         )
        assert r.status_code in (200, 201), r.text

        r2 = auth_headers.get(f"{BASE_URL}/api/projects/{created_project_id}",
                          )
        assert r2.status_code == 200
        d = r2.json()
        assert d.get("installation_date") == "2024-03-10"
        assert d.get("commissioning_date") == "2024-04-01"


# ===================== Till-date metrics =====================
class TestTillDateMetrics:
    SEEDED_PROJECT_ID = "6a0f4d9c857a4325028296d2"

    def test_reference_candidates_includes_till_date(self, auth_headers):
        r = auth_headers.get(f"{BASE_URL}/api/projects/reference-candidates",
                         )
        assert r.status_code == 200
        data = r.json()
        items = data if isinstance(data, list) else data.get("items", [])
        assert len(items) >= 1, "no reference candidates"
        # At least one item should have till_date attached
        with_till = [i for i in items if i.get("till_date")]
        assert with_till, "no candidate has till_date object"
        td = with_till[0]["till_date"]
        for k in ("installation_date", "months_elapsed", "years_elapsed",
                  "savings_inr", "units_generated", "co2_kg", "fuel_litres"):
            assert k in td, f"till_date missing {k}: {td}"

    def test_seeded_project_till_date(self, auth_headers):
        # Try direct summary
        r = auth_headers.get(
            f"{BASE_URL}/api/projects/{self.SEEDED_PROJECT_ID}/reference-summary")
        if r.status_code == 404:
            pytest.skip("seeded project not present in this env")
        assert r.status_code == 200, r.text
        data = r.json()
        td = data.get("till_date")
        assert td, f"till_date missing in summary: {data}"
        assert td.get("installation_date")
        # years_elapsed × 120000 ≈ savings_inr (allow wide tolerance)
        years = td.get("years_elapsed") or 0
        savings = td.get("savings_inr") or 0
        if years > 0:
            implied_annual = savings / years if years else 0
            # Should be close to 120000 per spec
            assert 100000 <= implied_annual <= 140000, \
                f"implied annual savings {implied_annual} not ~120000 (years={years}, savings={savings})"

    def test_reference_summary_has_till_date(self, auth_headers):
        # Use first reference candidate
        r = auth_headers.get(f"{BASE_URL}/api/projects/reference-candidates",
                         )
        assert r.status_code == 200
        items = r.json() if isinstance(r.json(), list) else r.json().get("items", [])
        if not items:
            pytest.skip("no candidates")
        pid = items[0].get("id")
        r2 = auth_headers.get(f"{BASE_URL}/api/projects/{pid}/reference-summary",
                          )
        assert r2.status_code == 200, r2.text
        data = r2.json()
        assert "till_date" in data
        # till_date can be None for projects with no install/commission/updated date


# ===================== Notes-only PUT regression =====================
class TestNotesOnlyOnCompleted:
    def test_notes_only_update_on_completed(self, auth_headers):
        # Find a completed project
        r = auth_headers.get(f"{BASE_URL}/api/projects")
        assert r.status_code == 200
        projects = r.json() if isinstance(r.json(), list) else r.json().get("items", [])
        completed = [p for p in projects if p.get("status") == "completed"]
        if not completed:
            pytest.skip("no completed projects")
        pid = completed[0].get("id")
        # PUT notes only — must not 403/422
        r2 = auth_headers.put(f"{BASE_URL}/api/projects/{pid}",
                          json={"notes": "TEST_iter36_notes_update"},
                          )
        assert r2.status_code in (200, 201), f"notes-only update failed: {r2.status_code} {r2.text}"
