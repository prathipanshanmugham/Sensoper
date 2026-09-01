from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Form
from fastapi.responses import JSONResponse, HTMLResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from pymongo import ReturnDocument
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Literal, Any, Dict
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import secrets
import json
import base64
import requests as http_requests

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration — fail fast if the secret isn't provisioned so we never
# fall back to an ephemeral secret that would invalidate every existing session
# on service restart.
JWT_SECRET = os.environ.get("JWT_SECRET")
if not JWT_SECRET or not JWT_SECRET.strip():
    raise RuntimeError(
        "JWT_SECRET environment variable is not set. Add JWT_SECRET to backend/.env "
        "(a long random string, e.g. output of `python -c 'import secrets; print(secrets.token_hex(32))'`). "
        "Refusing to start with an auto-generated secret — that would log out every user on every restart."
    )
JWT_ALGORITHM = "HS256"

# Object Storage Configuration
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "sensoper-solar"
storage_key = None

def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = http_requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = http_requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str):
    key = init_storage()
    resp = http_requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# Create the main app
app = FastAPI(title="Sensoper Solar Estimator API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ================== MODELS ==================

class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: Literal["admin", "manager", "staff"] = "staff"
    phone: Optional[str] = None

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Literal["admin", "manager", "staff"] = "staff"
    phone: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    phone: Optional[str] = None
    created_at: str

class CustomerDetails(BaseModel):
    name: str
    phone: str
    address: str
    email: Optional[str] = None

class LocationDetails(BaseModel):
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    address: Optional[str] = None
    site_location_words: Optional[str] = None
    pincode: Optional[str] = None
    district: Optional[str] = None
    state: Optional[str] = None
    discom_id: Optional[str] = None

class ElectricalDetails(BaseModel):
    sanction_load_kw: float
    connected_load_kw: float
    monthly_consumption_units: float
    eb_tariff: float
    service_type: Optional[str] = None
    connection_phase: Optional[str] = None       # Iter 41 Change 1 — Single/Three Phase
    _prefilled: Optional[dict] = None             # {eb_tariff: bool, service_type: bool}

    model_config = {"extra": "allow"}             # forward-compat for future fields

class SolarSystemInputs(BaseModel):
    system_type: str = "on-grid"
    inverter_model: Optional[str] = None
    panel_wattage: Optional[int] = 540
    battery_required: bool = False
    battery_capacity_ah: Optional[int] = None

class MountingStructure(BaseModel):
    roof_type: str  # Free text field (e.g., "RCC Flat Roof", "Metal Sheet", "Terrace with slope")
    tilt_angle: int
    structure_type: str

class AdditionalInputs(BaseModel):
    cable_length_meters: float
    inverter_to_panel_distance: float
    installation_complexity: str = "simple"
    shadow_analysis_notes: Optional[str] = None
    cable_length_unit: Optional[str] = "m"          # Iter 41 Change 1 — user-facing unit
    inverter_to_panel_unit: Optional[str] = "m"

    model_config = {"extra": "allow"}

class SelectedItem(BaseModel):
    inventory_item_id: Optional[str] = None
    name: str
    category: str
    unit_price: float
    gst_percentage: float = 18.0
    quantity: int = 1
    margin_percentage: float = 0

class ManualCost(BaseModel):
    description: str
    amount: float

class ProjectCreate(BaseModel):
    customer: CustomerDetails
    location: LocationDetails
    electrical: ElectricalDetails
    solar_system: SolarSystemInputs
    mounting: MountingStructure
    additional: AdditionalInputs
    selected_items: List[SelectedItem] = []
    manual_costs: List[ManualCost] = []
    site_images: List[str] = []
    drive_folder_name: Optional[str] = None
    drive_folder_link: Optional[str] = None
    drive_folder_id: Optional[str] = None
    site_measurements: Optional[dict] = None
    custom_fields: Optional[dict] = None
    solar_report: Optional[dict] = None
    calculation_snapshot: Optional[dict] = None
    terms_id: Optional[str] = None
    notes: Optional[str] = None
    reference_project_id: Optional[str] = None
    installation_date: Optional[str] = None
    commissioning_date: Optional[str] = None

class ProjectUpdate(BaseModel):
    customer: Optional[CustomerDetails] = None
    location: Optional[LocationDetails] = None
    electrical: Optional[ElectricalDetails] = None
    solar_system: Optional[SolarSystemInputs] = None
    mounting: Optional[MountingStructure] = None
    additional: Optional[AdditionalInputs] = None
    selected_items: Optional[List[SelectedItem]] = None
    manual_costs: Optional[List[ManualCost]] = None
    site_images: Optional[List[str]] = None
    drive_folder_name: Optional[str] = None
    drive_folder_link: Optional[str] = None
    drive_folder_id: Optional[str] = None
    site_measurements: Optional[dict] = None
    custom_fields: Optional[dict] = None
    solar_report: Optional[dict] = None
    calculation_snapshot: Optional[dict] = None
    terms_id: Optional[str] = None
    notes: Optional[str] = None
    reference_project_id: Optional[str] = None
    installation_date: Optional[str] = None
    commissioning_date: Optional[str] = None
    status: Optional[Literal["draft", "submitted", "approved", "rejected", "completed", "deletion_requested"]] = None

class ProjectNoteAppend(BaseModel):
    text: str

class AIRecommendationRequest(BaseModel):
    monthly_consumption_units: float
    sanction_load_kw: float
    roof_type: str
    budget_range: Optional[str] = None

# ================== NEW MODELS FOR ENTERPRISE FEATURES ==================

class TermsConditionsCreate(BaseModel):
    title: str
    content: str  # HTML content
    language: str = "en"

class TermsConditionsUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    is_active: Optional[bool] = None
    language: Optional[str] = None

class InventoryItemCreate(BaseModel):
    name: str
    sku_code: str
    category: str
    zone: Optional[str] = None
    aisle: Optional[str] = None
    shelf: Optional[str] = None
    rack: Optional[str] = None
    bin_location: Optional[str] = None
    quantity: int
    unit_price: float
    supplier: Optional[str] = None
    gst_percentage: float = 18.0
    hsn_code: Optional[str] = None
    reorder_level: int = 10
    image_url: Optional[str] = None
    margin_pct: float = 0
    active: bool = True
    qc_checklist: list = []
    procurement_date: Optional[str] = None
    addon_group: Optional[str] = None                # Iter 44 Change 3 — links to addon_groups collection
    location_id: Optional[str] = None

class InventoryItemUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    zone: Optional[str] = None
    aisle: Optional[str] = None
    shelf: Optional[str] = None
    rack: Optional[str] = None
    bin_location: Optional[str] = None
    quantity: Optional[int] = None
    unit_price: Optional[float] = None
    supplier: Optional[str] = None
    gst_percentage: Optional[float] = None
    hsn_code: Optional[str] = None
    reorder_level: Optional[int] = None
    image_url: Optional[str] = None
    margin_pct: Optional[float] = None
    active: Optional[bool] = None
    qc_checklist: Optional[list] = None
    procurement_date: Optional[str] = None
    addon_group: Optional[str] = None                # Iter 44 Change 3
    location_id: Optional[str] = None

class InventoryCategoryCreate(BaseModel):
    name: str
    slug: str
    description: Optional[str] = None

class DeletionRequestCreate(BaseModel):
    reason: str

# ================== MATERIAL KIT MODELS ==================

class MaterialKitLine(BaseModel):
    inventory_item_id: Optional[str] = None
    name: str
    category: Optional[str] = None
    quantity: float = 1
    qty_formula: Optional[str] = None   # e.g. "1 per kW", "1 fixed"
    notes: Optional[str] = None

class MaterialKitCreate(BaseModel):
    name: str
    system_type: Literal["on-grid", "off-grid", "hybrid", "solar-pump"]
    capacity_kw: float = 0
    capacity_min_kw: Optional[float] = None
    capacity_max_kw: Optional[float] = None
    description: Optional[str] = None
    lines: List[MaterialKitLine] = []
    active: bool = True

class MaterialKitUpdate(BaseModel):
    name: Optional[str] = None
    system_type: Optional[Literal["on-grid", "off-grid", "hybrid", "solar-pump"]] = None
    capacity_kw: Optional[float] = None
    capacity_min_kw: Optional[float] = None
    capacity_max_kw: Optional[float] = None
    description: Optional[str] = None
    lines: Optional[List[MaterialKitLine]] = None
    active: Optional[bool] = None

# ================== COMPANY PROFILE MODELS ==================

class BankDetails(BaseModel):
    account_name: Optional[str] = None
    account_number: Optional[str] = None
    ifsc_code: Optional[str] = None
    bank_name: Optional[str] = None
    branch: Optional[str] = None
    upi_id: Optional[str] = None

class CompanyProfileCreate(BaseModel):
    company_name: str
    tagline: Optional[str] = None
    logo_url: Optional[str] = None
    primary_color: str = "#4ADE40"
    secondary_color: str = "#2D9BF0"
    address: str
    phone: str
    email: EmailStr
    website: Optional[str] = None
    gst_number: Optional[str] = None
    pan_number: Optional[str] = None
    state: Optional[str] = None
    location_id: Optional[str] = None
    bank_details: Optional[BankDetails] = None
    authorized_signatory: Optional[str] = None
    designation: Optional[str] = None

class CompanyProfileUpdate(BaseModel):
    company_name: Optional[str] = None
    tagline: Optional[str] = None
    logo_url: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[EmailStr] = None
    website: Optional[str] = None
    gst_number: Optional[str] = None
    pan_number: Optional[str] = None
    state: Optional[str] = None
    location_id: Optional[str] = None
    bank_details: Optional[BankDetails] = None
    authorized_signatory: Optional[str] = None
    designation: Optional[str] = None
    is_active: Optional[bool] = None

# ================== FORM TABS (Dynamic Form Engine) MODELS ==================

class FormFieldDefinition(BaseModel):
    name: str
    label: str
    type: str  # text, number, select, textarea, checkbox, date
    required: bool = False
    placeholder: str = ""
    options: List[str] = []

class FormTabCreate(BaseModel):
    name: str
    fields: List[FormFieldDefinition]
    roles_visible: List[str] = ["admin", "manager", "staff"]

class FormTabUpdate(BaseModel):
    name: Optional[str] = None
    fields: Optional[List[FormFieldDefinition]] = None
    roles_visible: Optional[List[str]] = None
    active: Optional[bool] = None

# ================== DAILY UPDATE & PAYMENT MODELS ==================

class DailyUpdateCreate(BaseModel):
    project_id: str
    update_type: str  # progress, material, payment, installation, om
    data: dict

class DailyUpdateUpdate(BaseModel):
    data: Optional[dict] = None
    update_type: Optional[str] = None

class PaymentCreate(BaseModel):
    project_id: str
    amount: float
    payment_method: str  # cash, cheque, upi, bank_transfer, emi
    notes: str = ""

class MaterialUsageCreate(BaseModel):
    project_id: str
    item_name: str
    estimated_qty: float
    actual_qty: float
    wastage: float = 0
    notes: str = ""

# ================== ALERT & THRESHOLD MODELS ==================

class ThresholdUpdate(BaseModel):
    min_margin_pct: Optional[float] = None
    max_material_variance_pct: Optional[float] = None
    payment_delay_days: Optional[int] = None
    max_project_duration_days: Optional[int] = None
    underpriced_margin_pct: Optional[float] = None

# ================== OPERATIONAL MODULE MODELS ==================

class CustomerCreditCreate(BaseModel):
    customer_name: str
    customer_phone: str = ""
    invoice_ref: str = ""
    total_amount: float
    due_date: str = ""
    notes: str = ""

class CreditPaymentCreate(BaseModel):
    credit_id: str
    amount: float
    payment_method: str = "cash"
    notes: str = ""

class PurchaseOrderCreate(BaseModel):
    supplier_name: str
    supplier_contact: str = ""
    items: list  # [{name, qty, unit_price, inventory_item_id?, sku_code?}]
    expected_delivery: str = ""
    notes: str = ""
    location_id: Optional[str] = None

class InboundEditLine(BaseModel):
    inventory_item_id: str
    qty_received: float
    notes: Optional[str] = None

class InboundEditRequest(BaseModel):
    lines: List[InboundEditLine] = Field(..., min_items=1)
    storage_location: Optional[str] = None

class DeliveryOutboundCreate(BaseModel):
    project_id: str = ""
    customer_name: str
    customer_address: str = ""
    customer_contact: str = ""
    items: list  # [{name, qty}]
    transporter_name: str = ""
    vehicle_number: str = ""
    driver_contact: str = ""
    dispatch_date: str = ""
    delivery_date: str = ""
    distance_km: float = 0
    notes: str = ""

class DeliveryOutboundEdit(BaseModel):
    items: Optional[list] = None
    transporter_name: Optional[str] = None
    vehicle_number: Optional[str] = None
    driver_contact: Optional[str] = None
    dispatch_date: Optional[str] = None
    delivery_date: Optional[str] = None
    notes: Optional[str] = None
    confirm_reconciliation_impact: bool = False
    admin_reason: Optional[str] = None

class BrandReturnCreate(BaseModel):
    project_id: str = ""
    supplier_name: str = ""
    item_name: str
    quantity: float
    reason: str  # damage, excess, defect
    notes: str = ""

class AuditCreate(BaseModel):
    title: str
    project_id: str = ""
    auditor_name: str
    deadline: str = ""
    checklist: list  # [{item, status, notes}]
    notes: str = ""

# ================== HELPER FUNCTIONS ==================

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=60),
        "type": "access"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return {
            "id": str(user["_id"]),
            "email": user["email"],
            "name": user["name"],
            "role": user["role"],
            "phone": user.get("phone"),
            "created_at": user["created_at"],
            "location_ids": user.get("location_ids", []),
            "default_location_id": user.get("default_location_id")
        }
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(*roles):
    async def role_checker(request: Request):
        user = await get_current_user(request)
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return role_checker

def serialize_for_json(obj):
    """Helper to serialize objects for JSON, handling ObjectId"""
    if obj is None:
        return None
    if isinstance(obj, ObjectId):
        return str(obj)
    if isinstance(obj, dict):
        return {k: serialize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [serialize_for_json(item) for item in obj]
    return obj

async def create_audit_log(user_id: str, user_name: str, action_type: str, entity_type: str, entity_id: str, old_data: Any = None, new_data: Any = None, details: str = None):
    """Create an audit log entry"""
    log_entry = {
        "user_id": user_id,
        "user_name": user_name,
        "action_type": action_type,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "old_data": json.dumps(serialize_for_json(old_data)) if old_data else None,
        "new_data": json.dumps(serialize_for_json(new_data)) if new_data else None,
        "details": details,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(log_entry)

# ================== DYNAMIC PERMISSIONS ==================

DEFAULT_PERMISSIONS = {
    "admin": {
        "can_create_project": True, "can_edit_project": True, "can_delete_project": True,
        "can_request_delete": True, "can_approve_deletion": True,
        "can_approve_quotation": True, "can_set_margin": True, "can_approve_margin": True,
        "can_edit_inventory": True, "can_approve_inventory": True,
        "can_manage_users": True, "can_change_user_access": True,
        "can_view_reports": True, "can_view_audit_logs": True,
        "can_manage_company": True, "can_manage_terms": True,
        # Module-level (added Feb 2026 for Accounts, Readings, refreshed UI)
        "module_dashboard": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_ceo_dashboard": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_expansion": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_direct_sales": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_accounts": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_readings": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_inventory": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_purchase_inbound": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_delivery_outbound": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_credits": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_returns": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_audits": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_reports": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_alerts": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_approvals": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_users": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_permissions": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_settings": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_assets": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_amc": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_locations": {"view": True, "create": True, "edit": True, "delete": True, "export": True}
    },
    "manager": {
        "can_create_project": True, "can_edit_project": True, "can_delete_project": False,
        "can_request_delete": True, "can_approve_deletion": True,
        "can_approve_quotation": True, "can_set_margin": True, "can_approve_margin": True,
        "can_edit_inventory": True, "can_approve_inventory": True,
        "can_manage_users": False, "can_change_user_access": False,
        "can_view_reports": True, "can_view_audit_logs": True,
        "can_manage_company": False, "can_manage_terms": True,
        "module_dashboard": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_ceo_dashboard": {"view": True, "create": False, "edit": False, "delete": False, "export": True},
        "module_expansion": {"view": True, "create": False, "edit": False, "delete": False, "export": True},
        "module_direct_sales": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_accounts": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_readings": {"view": True, "create": True, "edit": True, "delete": True, "export": True},
        "module_inventory": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_purchase_inbound": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_delivery_outbound": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_credits": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_returns": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_audits": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_reports": {"view": True, "create": False, "edit": False, "delete": False, "export": True},
        "module_alerts": {"view": True, "create": False, "edit": True, "delete": False, "export": True},
        "module_approvals": {"view": True, "create": False, "edit": True, "delete": False, "export": False},
        "module_users": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_permissions": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_settings": {"view": True, "create": False, "edit": False, "delete": False, "export": False},
        "module_assets": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_amc": {"view": True, "create": True, "edit": True, "delete": False, "export": True},
        "module_locations": {"view": True, "create": False, "edit": False, "delete": False, "export": False}
    },
    "staff": {
        "can_create_project": True, "can_edit_project": True, "can_delete_project": False,
        "can_request_delete": True, "can_approve_deletion": False,
        "can_approve_quotation": False, "can_set_margin": False, "can_approve_margin": False,
        "can_edit_inventory": False, "can_approve_inventory": False,
        "can_manage_users": False, "can_change_user_access": False,
        "can_view_reports": False, "can_view_audit_logs": False,
        "can_manage_company": False, "can_manage_terms": False,
        "module_dashboard": {"view": True, "create": True, "edit": True, "delete": False, "export": False},
        "module_ceo_dashboard": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_expansion": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_direct_sales": {"view": True, "create": True, "edit": False, "delete": False, "export": False},
        "module_accounts": {"view": True, "create": True, "edit": False, "delete": False, "export": False},
        "module_readings": {"view": True, "create": True, "edit": True, "delete": False, "export": False},
        "module_inventory": {"view": True, "create": False, "edit": False, "delete": False, "export": False},
        "module_purchase_inbound": {"view": True, "create": False, "edit": False, "delete": False, "export": False},
        "module_delivery_outbound": {"view": True, "create": False, "edit": False, "delete": False, "export": False},
        "module_credits": {"view": True, "create": False, "edit": False, "delete": False, "export": False},
        "module_returns": {"view": True, "create": True, "edit": False, "delete": False, "export": False},
        "module_audits": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_reports": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_alerts": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_approvals": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_users": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_permissions": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_settings": {"view": False, "create": False, "edit": False, "delete": False, "export": False},
        "module_assets": {"view": True, "create": False, "edit": False, "delete": False, "export": False},
        "module_amc": {"view": True, "create": False, "edit": False, "delete": False, "export": False},
        "module_locations": {"view": True, "create": False, "edit": False, "delete": False, "export": False}
    }
}

async def get_permissions(role: str) -> dict:
    """Get permissions for a role from DB, falling back to defaults"""
    perm = await db.role_permissions.find_one({"role_name": role})
    if perm:
        return perm.get("permissions", DEFAULT_PERMISSIONS.get(role, {}))
    return DEFAULT_PERMISSIONS.get(role, {})

async def check_permission(user: dict, permission: str) -> bool:
    """Check if a user has a specific permission"""
    if user["role"] == "admin":
        admin_perms = await get_permissions("admin")
        return admin_perms.get(permission, True)
    perms = await get_permissions(user["role"])
    return perms.get(permission, False)

async def require_permission(request: Request, permission: str) -> dict:
    """Middleware-like function to check permission"""
    user = await get_current_user(request)
    has_perm = await check_permission(user, permission)
    if not has_perm:
        raise HTTPException(status_code=403, detail=f"Permission denied: {permission}")
    return user

async def check_module_permission(user: dict, module: str, action: str) -> bool:
    """Read permission for a module.action from the dynamic role_permissions matrix."""
    perms = await get_permissions(user["role"])
    mod = perms.get(module, {})
    if isinstance(mod, dict):
        return bool(mod.get(action, False))
    return False


def calculate_cost_estimation(selected_items: list, manual_costs: list) -> dict:
    """Calculate cost estimation from selected inventory items and manual costs with per-item margins"""
    items_breakdown = []
    total_items_cost = 0
    total_gst = 0
    total_margin = 0

    for item in selected_items:
        item_cost = item["unit_price"] * item["quantity"]
        item_margin_pct = item.get("margin_percentage", 0)
        item_margin = item_cost * (item_margin_pct / 100)
        item_gst = item_cost * (item["gst_percentage"] / 100)
        total_items_cost += item_cost
        total_gst += item_gst
        total_margin += item_margin
        items_breakdown.append({
            "name": item["name"],
            "category": item["category"],
            "unit_price": item["unit_price"],
            "quantity": item["quantity"],
            "gst_percentage": item["gst_percentage"],
            "margin_percentage": item_margin_pct,
            "amount": round(item_cost, 2),
            "gst_amount": round(item_gst, 2),
            "margin_amount": round(item_margin, 2)
        })

    manual_total = sum(c["amount"] for c in manual_costs)
    subtotal = total_items_cost + manual_total
    total_cost = subtotal + total_margin + total_gst

    return {
        "items_breakdown": items_breakdown,
        "manual_costs": [{"description": c["description"], "amount": round(c["amount"], 2)} for c in manual_costs],
        "items_subtotal": round(total_items_cost, 2),
        "manual_subtotal": round(manual_total, 2),
        "subtotal": round(subtotal, 2),
        "total_margin": round(total_margin, 2),
        "total_gst": round(total_gst, 2),
        "total_cost": round(total_cost, 2)
    }

# ================== AUTH ENDPOINTS ==================

@api_router.post("/auth/register")
async def register(user_data: UserCreate, response: Response):
    email = user_data.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed = hash_password(user_data.password)
    user_doc = {
        "email": email,
        "password_hash": hashed,
        "name": user_data.name,
        "role": user_data.role,
        "phone": user_data.phone,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    
    access_token = create_access_token(user_id, email, user_data.role)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    
    return {
        "id": user_id,
        "email": email,
        "name": user_data.name,
        "role": user_data.role,
        "phone": user_data.phone,
        "created_at": user_doc["created_at"]
    }

@api_router.post("/auth/login")
async def login(credentials: UserLogin, response: Response, request: Request):
    email = credentials.email.lower()
    
    # Brute force protection
    identifier = f"{request.client.host}:{email}"
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("count", 0) >= 5:
        lockout_time = attempt.get("locked_until")
        if lockout_time and datetime.fromisoformat(lockout_time) > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail="Too many failed attempts. Try again later.")
        else:
            await db.login_attempts.delete_one({"identifier": identifier})
    
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(credentials.password, user["password_hash"]):
        # Increment failed attempts
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {
                "$inc": {"count": 1},
                "$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}
            },
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Clear failed attempts on success
    await db.login_attempts.delete_one({"identifier": identifier})
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email, user["role"])
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    
    return {
        "id": user_id,
        "email": user["email"],
        "name": user["name"],
        "role": user["role"],
        "phone": user.get("phone"),
        "created_at": user["created_at"]
    }

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out successfully"}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        
        user_id = str(user["_id"])
        access_token = create_access_token(user_id, user["email"], user["role"])
        
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=3600, path="/")
        
        return {"message": "Token refreshed"}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ================== USER MANAGEMENT (Admin Only) ==================

@api_router.get("/users")
async def get_users(request: Request):
    user = await require_role("admin")(request)
    users = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    return [
        {
            "id": str(u["_id"]),
            "email": u["email"],
            "name": u["name"],
            "role": u["role"],
            "phone": u.get("phone"),
            "created_at": u["created_at"],
            "location_ids": u.get("location_ids", []),
            "default_location_id": u.get("default_location_id")
        }
        for u in users
    ]

@api_router.post("/users")
async def create_user(user_data: UserCreate, request: Request):
    current_user = await require_role("admin")(request)
    
    email = user_data.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed = hash_password(user_data.password)
    user_doc = {
        "email": email,
        "password_hash": hashed,
        "name": user_data.name,
        "role": user_data.role,
        "phone": user_data.phone,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.users.insert_one(user_doc)
    
    await create_audit_log(
        current_user["id"], current_user["name"], "create", "user", 
        str(result.inserted_id), None, {"email": email, "role": user_data.role}
    )
    
    return {
        "id": str(result.inserted_id),
        "email": email,
        "name": user_data.name,
        "role": user_data.role,
        "phone": user_data.phone,
        "created_at": user_doc["created_at"]
    }

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, request: Request):
    current_user = await require_role("admin")(request)
    
    body = await request.json()
    update_data = {}
    
    if "name" in body:
        update_data["name"] = body["name"]
    if "role" in body:
        update_data["role"] = body["role"]
    if "phone" in body:
        update_data["phone"] = body["phone"]
    if "password" in body and body["password"]:
        update_data["password_hash"] = hash_password(body["password"])
    if "location_ids" in body:
        update_data["location_ids"] = body["location_ids"]
    if "default_location_id" in body:
        update_data["default_location_id"] = body["default_location_id"]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    await create_audit_log(
        current_user["id"], current_user["name"], "update", "user", 
        user_id, None, update_data
    )
    
    return {"message": "User updated successfully"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request):
    current_user = await require_role("admin")(request)
    
    if current_user["id"] == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    await create_audit_log(
        current_user["id"], current_user["name"], "delete", "user", user_id
    )
    
    return {"message": "User deleted successfully"}

# ================== COMPANY PROFILE ==================

@api_router.get("/company")
async def get_company_profiles(request: Request):
    """Get all company profiles"""
    await get_current_user(request)
    
    profiles = await db.company_profiles.find().to_list(100)
    return [
        {
            "id": str(p["_id"]),
            "company_name": p["company_name"],
            "tagline": p.get("tagline"),
            "logo_url": p.get("logo_url"),
            "primary_color": p.get("primary_color", "#4ADE40"),
            "secondary_color": p.get("secondary_color", "#2D9BF0"),
            "address": p["address"],
            "phone": p["phone"],
            "email": p["email"],
            "website": p.get("website"),
            "gst_number": p.get("gst_number"),
            "pan_number": p.get("pan_number"),
            "state": p.get("state"),
            "location_id": p.get("location_id"),
            "bank_details": p.get("bank_details"),
            "authorized_signatory": p.get("authorized_signatory"),
            "designation": p.get("designation"),
            "is_active": p.get("is_active", False),
            "created_at": p["created_at"]
        }
        for p in profiles
    ]

@api_router.get("/company/active")
async def get_active_company():
    """Get the active company profile (for PDF generation, public access)"""
    profile = await db.company_profiles.find_one({"is_active": True})
    
    if not profile:
        # Return default profile if none exists
        return {
            "id": None,
            "company_name": "Sensoper Controls & Renewables",
            "tagline": "Solar Solutions Provider",
            "logo_url": "https://customer-assets.emergentagent.com/job_8c20414a-b147-464e-9c68-aaa2fa40fdbf/artifacts/q52gayft_snspr.png",
            "primary_color": "#4ADE40",
            "secondary_color": "#2D9BF0",
            "address": "Tamil Nadu, India",
            "phone": "+91 XXXXX XXXXX",
            "email": "info@sensoper.com",
            "website": "www.sensoper.com",
            "gst_number": None,
            "pan_number": None,
            "state": "Tamil Nadu",
            "location_id": None,
            "bank_details": None,
            "authorized_signatory": None,
            "designation": None
        }
    
    return {
        "id": str(profile["_id"]),
        "company_name": profile["company_name"],
        "tagline": profile.get("tagline"),
        "logo_url": profile.get("logo_url"),
        "primary_color": profile.get("primary_color", "#4ADE40"),
        "secondary_color": profile.get("secondary_color", "#2D9BF0"),
        "address": profile["address"],
        "phone": profile["phone"],
        "email": profile["email"],
        "website": profile.get("website"),
        "gst_number": profile.get("gst_number"),
        "pan_number": profile.get("pan_number"),
        "state": profile.get("state"),
        "location_id": profile.get("location_id"),
        "bank_details": profile.get("bank_details"),
        "authorized_signatory": profile.get("authorized_signatory"),
        "designation": profile.get("designation")
    }

@api_router.post("/company")
async def create_company_profile(profile: CompanyProfileCreate, request: Request):
    """Create a new company profile"""
    current_user = await require_role("admin")(request)
    
    profile_doc = {
        "company_name": profile.company_name,
        "tagline": profile.tagline,
        "logo_url": profile.logo_url,
        "primary_color": profile.primary_color,
        "secondary_color": profile.secondary_color,
        "address": profile.address,
        "phone": profile.phone,
        "email": profile.email,
        "website": profile.website,
        "gst_number": profile.gst_number,
        "pan_number": profile.pan_number,
        "state": profile.state,
        "location_id": profile.location_id,
        "bank_details": profile.bank_details.model_dump() if profile.bank_details else None,
        "authorized_signatory": profile.authorized_signatory,
        "designation": profile.designation,
        "is_active": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.company_profiles.insert_one(profile_doc)
    
    await create_audit_log(
        current_user["id"], current_user["name"], "create", "company_profile",
        str(result.inserted_id), None, {"company_name": profile.company_name}
    )
    
    return {"id": str(result.inserted_id), "message": "Company profile created successfully"}

@api_router.put("/company/{profile_id}")
async def update_company_profile(profile_id: str, updates: CompanyProfileUpdate, request: Request):
    """Update a company profile"""
    current_user = await require_role("admin")(request)
    
    profile = await db.company_profiles.find_one({"_id": ObjectId(profile_id)})
    if not profile:
        raise HTTPException(status_code=404, detail="Company profile not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if updates.company_name is not None:
        update_data["company_name"] = updates.company_name
    if updates.tagline is not None:
        update_data["tagline"] = updates.tagline
    if updates.logo_url is not None:
        update_data["logo_url"] = updates.logo_url
    if updates.primary_color is not None:
        update_data["primary_color"] = updates.primary_color
    if updates.secondary_color is not None:
        update_data["secondary_color"] = updates.secondary_color
    if updates.address is not None:
        update_data["address"] = updates.address
    if updates.phone is not None:
        update_data["phone"] = updates.phone
    if updates.email is not None:
        update_data["email"] = updates.email
    if updates.website is not None:
        update_data["website"] = updates.website
    if updates.gst_number is not None:
        update_data["gst_number"] = updates.gst_number
    if updates.pan_number is not None:
        update_data["pan_number"] = updates.pan_number
    if updates.state is not None:
        update_data["state"] = updates.state
    if updates.location_id is not None:
        update_data["location_id"] = updates.location_id
    if updates.bank_details is not None:
        update_data["bank_details"] = updates.bank_details.model_dump()
    if updates.authorized_signatory is not None:
        update_data["authorized_signatory"] = updates.authorized_signatory
    if updates.designation is not None:
        update_data["designation"] = updates.designation
    
    # Handle activation
    if updates.is_active is True:
        # Deactivate all other profiles
        await db.company_profiles.update_many(
            {"_id": {"$ne": ObjectId(profile_id)}},
            {"$set": {"is_active": False}}
        )
        update_data["is_active"] = True
    elif updates.is_active is False:
        update_data["is_active"] = False
    
    await db.company_profiles.update_one(
        {"_id": ObjectId(profile_id)},
        {"$set": update_data}
    )
    
    await create_audit_log(
        current_user["id"], current_user["name"], "update", "company_profile",
        profile_id, None, update_data
    )
    
    return {"message": "Company profile updated successfully"}

@api_router.delete("/company/{profile_id}")
async def delete_company_profile(profile_id: str, request: Request):
    """Delete a company profile"""
    current_user = await require_role("admin")(request)
    
    profile = await db.company_profiles.find_one({"_id": ObjectId(profile_id)})
    if not profile:
        raise HTTPException(status_code=404, detail="Company profile not found")
    
    if profile.get("is_active"):
        raise HTTPException(status_code=400, detail="Cannot delete active profile")
    
    await db.company_profiles.delete_one({"_id": ObjectId(profile_id)})
    
    await create_audit_log(
        current_user["id"], current_user["name"], "delete", "company_profile", profile_id
    )
    
    return {"message": "Company profile deleted successfully"}

@api_router.post("/company/upload-logo")
async def upload_company_logo(request: Request, file: UploadFile = File(...)):
    """Upload a logo image and return a base64 data URL"""
    current_user = await require_role("admin")(request)
    
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Only image files are allowed")
    
    contents = await file.read()
    if len(contents) > 2 * 1024 * 1024:  # 2MB limit
        raise HTTPException(status_code=400, detail="File size must be under 2MB")
    
    import base64
    b64 = base64.b64encode(contents).decode("utf-8")
    data_url = f"data:{file.content_type};base64,{b64}"
    
    return {"logo_url": data_url}

# ================== TERMS & CONDITIONS ==================

@api_router.get("/terms")
async def get_all_terms(request: Request):
    await get_current_user(request)
    
    terms = await db.terms_conditions.find().sort("version", -1).to_list(100)
    return [
        {
            "id": str(t["_id"]),
            "title": t["title"],
            "content": t["content"],
            "version": t["version"],
            "is_active": t.get("is_active", False),
            "language": t.get("language", "en"),
            "created_by": t.get("created_by"),
            "created_by_name": t.get("created_by_name"),
            "created_at": t["created_at"],
            "updated_at": t.get("updated_at")
        }
        for t in terms
    ]

@api_router.get("/terms/active")
async def get_active_terms(language: str = "en"):
    """Get the active terms & conditions for PDF generation"""
    terms = await db.terms_conditions.find_one({"is_active": True, "language": language})
    if not terms:
        # Return default terms if none set
        return {
            "id": None,
            "title": "Standard Terms & Conditions",
            "content": """<ol>
<li>This quotation is valid for 30 days from the date of issue.</li>
<li>50% advance payment required to confirm the order.</li>
<li>Balance payment due upon installation completion.</li>
<li>Installation timeline: 7-14 working days after material delivery.</li>
<li>5-year warranty on installation workmanship.</li>
<li>Panel warranty as per manufacturer terms (typically 25 years).</li>
<li>Inverter warranty as per manufacturer terms.</li>
<li>All prices are subject to change without prior notice.</li>
</ol>""",
            "version": 0
        }
    return {
        "id": str(terms["_id"]),
        "title": terms["title"],
        "content": terms["content"],
        "version": terms["version"]
    }

@api_router.get("/terms/{terms_id}")
async def get_terms_by_id(terms_id: str, request: Request):
    """Get a specific terms & conditions by ID — used when a project has a selected terms_id."""
    await get_current_user(request)
    try:
        oid = ObjectId(terms_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid terms id")
    terms = await db.terms_conditions.find_one({"_id": oid})
    if not terms:
        raise HTTPException(status_code=404, detail="Terms not found")
    return {
        "id": str(terms["_id"]),
        "title": terms["title"],
        "content": terms["content"],
        "version": terms["version"],
        "language": terms.get("language", "en")
    }

@api_router.post("/terms")
async def create_terms(terms_data: TermsConditionsCreate, request: Request):
    current_user = await require_role("admin", "manager")(request)
    
    # Get the next version number
    latest = await db.terms_conditions.find_one(
        {"language": terms_data.language}, 
        sort=[("version", -1)]
    )
    next_version = (latest["version"] + 1) if latest else 1
    
    terms_doc = {
        "title": terms_data.title,
        "content": terms_data.content,
        "version": next_version,
        "is_active": False,
        "language": terms_data.language,
        "created_by": current_user["id"],
        "created_by_name": current_user["name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.terms_conditions.insert_one(terms_doc)
    
    await create_audit_log(
        current_user["id"], current_user["name"], "create", "terms_conditions",
        str(result.inserted_id), None, {"title": terms_data.title, "version": next_version}
    )
    
    return {
        "id": str(result.inserted_id),
        "version": next_version,
        "message": "Terms created successfully"
    }

@api_router.put("/terms/{terms_id}")
async def update_terms(terms_id: str, updates: TermsConditionsUpdate, request: Request):
    current_user = await require_role("admin", "manager")(request)
    
    terms = await db.terms_conditions.find_one({"_id": ObjectId(terms_id)})
    if not terms:
        raise HTTPException(status_code=404, detail="Terms not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if updates.title is not None:
        update_data["title"] = updates.title
    if updates.content is not None:
        update_data["content"] = updates.content
    if updates.language is not None:
        update_data["language"] = updates.language
    
    # Handle activation
    if updates.is_active is True:
        # Deactivate all other terms for this language
        lang = updates.language or terms.get("language", "en")
        await db.terms_conditions.update_many(
            {"language": lang, "_id": {"$ne": ObjectId(terms_id)}},
            {"$set": {"is_active": False}}
        )
        update_data["is_active"] = True
    elif updates.is_active is False:
        update_data["is_active"] = False
    
    await db.terms_conditions.update_one(
        {"_id": ObjectId(terms_id)},
        {"$set": update_data}
    )
    
    await create_audit_log(
        current_user["id"], current_user["name"], "update", "terms_conditions",
        terms_id, {"is_active": terms.get("is_active")}, update_data
    )
    
    return {"message": "Terms updated successfully"}

@api_router.delete("/terms/{terms_id}")
async def delete_terms(terms_id: str, request: Request):
    current_user = await require_role("admin")(request)
    
    terms = await db.terms_conditions.find_one({"_id": ObjectId(terms_id)})
    if not terms:
        raise HTTPException(status_code=404, detail="Terms not found")
    
    await db.terms_conditions.delete_one({"_id": ObjectId(terms_id)})
    
    await create_audit_log(
        current_user["id"], current_user["name"], "delete", "terms_conditions", terms_id
    )
    
    return {"message": "Terms deleted successfully"}

# ================== GOOGLE DRIVE SETTINGS ==================

@api_router.get("/drive/settings")
async def get_drive_settings(request: Request):
    await get_current_user(request)
    settings = await db.drive_settings.find_one({}, {"_id": 0})
    if not settings:
        return {"folder_name": "", "folder_link": ""}
    return {"folder_name": settings.get("folder_name", ""), "folder_link": settings.get("folder_link", "")}

@api_router.put("/drive/settings")
async def update_drive_settings(request: Request):
    await require_role("admin", "manager")(request)
    body = await request.json()
    folder_name = body.get("folder_name", "").strip()
    folder_link = body.get("folder_link", "").strip()
    await db.drive_settings.update_one(
        {},
        {"$set": {"folder_name": folder_name, "folder_link": folder_link, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"folder_name": folder_name, "folder_link": folder_link}

@api_router.post("/upload/site-image")
async def upload_site_image(request: Request, file: UploadFile = File(...)):
    await get_current_user(request)
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Only image files are allowed")
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be under 10MB")
    ext = file.filename.split(".")[-1] if "." in file.filename else "png"
    path = f"{APP_NAME}/site-images/{uuid.uuid4()}.{ext}"
    try:
        result = put_object(path, contents, file.content_type or "image/png")
        return {"storage_path": result["path"], "filename": file.filename, "size": result.get("size", len(contents))}
    except Exception as e:
        logger.error(f"Site image upload failed: {e}")
        raise HTTPException(status_code=500, detail="Image upload failed")

# ================== INVENTORY CATEGORIES ==================

@api_router.get("/inventory/categories")
async def get_inventory_categories(request: Request):
    await get_current_user(request)
    categories = await db.inventory_categories.find().to_list(100)
    return [
        {"id": str(c["_id"]), "name": c["name"], "slug": c["slug"], "description": c.get("description", "")}
        for c in categories
    ]

@api_router.post("/inventory/categories")
async def create_inventory_category(cat: InventoryCategoryCreate, request: Request):
    current_user = await require_role("admin", "manager")(request)
    existing = await db.inventory_categories.find_one({"slug": cat.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Category slug already exists")
    result = await db.inventory_categories.insert_one({
        "name": cat.name, "slug": cat.slug, "description": cat.description or "",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    return {"id": str(result.inserted_id), "message": "Category created"}

@api_router.delete("/inventory/categories/{cat_id}")
async def delete_inventory_category(cat_id: str, request: Request):
    current_user = await require_role("admin")(request)
    cat = await db.inventory_categories.find_one({"_id": ObjectId(cat_id)})
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    items_count = await db.inventory_items.count_documents({"category": cat["slug"]})
    if items_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete category with {items_count} items")
    await db.inventory_categories.delete_one({"_id": ObjectId(cat_id)})
    return {"message": "Category deleted"}

# ================== FILE UPLOAD ==================

@api_router.post("/upload/image")
async def upload_image(request: Request, file: UploadFile = File(...)):
    current_user = await get_current_user(request)
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Only image files are allowed")
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be under 5MB")
    ext = file.filename.split(".")[-1] if "." in file.filename else "png"
    path = f"{APP_NAME}/images/{uuid.uuid4()}.{ext}"
    try:
        result = put_object(path, contents, file.content_type or "image/png")
        return {"storage_path": result["path"], "size": result.get("size", len(contents))}
    except Exception as e:
        logger.error(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail="Image upload failed")

@api_router.post("/upload/media")
async def upload_media(request: Request, file: UploadFile = File(...)):
    """Upload photos or videos for project completion"""
    current_user = await get_current_user(request)
    allowed_prefixes = ["image/", "video/"]
    if not file.content_type or not any(file.content_type.startswith(t) for t in allowed_prefixes):
        raise HTTPException(status_code=400, detail="Only image and video files are allowed")
    contents = await file.read()
    if len(contents) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be under 50MB")
    ext = file.filename.split(".")[-1] if "." in file.filename else "bin"
    media_type = "images" if file.content_type.startswith("image/") else "videos"
    path = f"{APP_NAME}/completion/{media_type}/{uuid.uuid4()}.{ext}"
    try:
        result = put_object(path, contents, file.content_type)
        return {
            "storage_path": result["path"],
            "media_type": media_type,
            "filename": file.filename,
            "content_type": file.content_type,
            "size": result.get("size", len(contents))
        }
    except Exception as e:
        logger.error(f"Media upload failed: {e}")
        raise HTTPException(status_code=500, detail="Media upload failed")

@api_router.get("/files/{path:path}")
async def serve_file(path: str, request: Request):
    try:
        data, content_type = get_object(path)
        return Response(content=data, media_type=content_type)
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")

# ================== MARGIN UPDATE ==================

@api_router.put("/projects/{project_id}/margin")
async def update_project_margin(project_id: str, request: Request):
    current_user = await require_role("admin", "manager")(request)
    body = await request.json()
    item_margins = body.get("item_margins", [])
    
    if not item_margins:
        raise HTTPException(status_code=400, detail="No margin updates provided")
    
    project = await db.projects.find_one({"_id": ObjectId(project_id)})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    selected_items = project.get("selected_items", [])
    
    for margin_update in item_margins:
        idx = margin_update.get("index")
        pct = margin_update.get("margin_percentage", 0)
        if idx is not None and 0 <= idx < len(selected_items):
            selected_items[idx]["margin_percentage"] = float(pct)
    
    manual_costs = project.get("manual_costs", [])
    new_estimation = calculate_cost_estimation(selected_items, manual_costs)
    
    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$set": {
            "selected_items": selected_items,
            "cost_estimation": new_estimation,
            "margin_added_by": current_user["name"],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_audit_log(
        current_user["id"], current_user["name"], "update", "project_margin",
        project_id, None, {"item_margins": item_margins}
    )
    
    return {"message": "Margins updated", "cost_estimation": new_estimation, "selected_items": selected_items}

# ================== INVENTORY MANAGEMENT ==================

@api_router.get("/inventory/items")
async def get_inventory_items(
    request: Request, 
    category: Optional[str] = None,
    low_stock: bool = False,
    location_id: Optional[str] = None
):
    user = await get_current_user(request)
    
    query = {}
    if category:
        query["category"] = category
    loc_filter = location_scope_filter(user, location_id)
    if loc_filter:
        query.update(loc_filter)
    
    items = await db.inventory_items.find(query).to_list(500)
    
    result = []
    for item in items:
        item_data = {
            "id": str(item["_id"]),
            "name": item["name"],
            "sku_code": item["sku_code"],
            "category": item["category"],
            "location_id": item.get("location_id"),
            "addon_group": item.get("addon_group"),
            "zone": item.get("zone", ""),
            "aisle": item.get("aisle", ""),
            "shelf": item.get("shelf", ""),
            "rack": item.get("rack", ""),
            "bin_location": item.get("bin_location", ""),
            "quantity": item["quantity"],
            "unit_price": item["unit_price"],
            "supplier": item.get("supplier"),
            "gst_percentage": item.get("gst_percentage", 18.0),
            "hsn_code": item.get("hsn_code"),
            "reorder_level": item.get("reorder_level", 10),
            "image_url": item.get("image_url"),
            "margin_pct": item.get("margin_pct", 0),
            "active": item.get("active", True),
            "qc_checklist": item.get("qc_checklist", []),
            "procurement_date": item.get("procurement_date"),
            "created_at": item["created_at"],
            "updated_at": item.get("updated_at")
        }
        
        if item["quantity"] <= item.get("reorder_level", 10):
            item_data["low_stock_alert"] = True
        
        if low_stock and not item_data.get("low_stock_alert"):
            continue
            
        result.append(item_data)
    
    return result

@api_router.get("/inventory/items/{item_id}")
async def get_inventory_item(item_id: str, request: Request):
    await get_current_user(request)
    
    item = await db.inventory_items.find_one({"_id": ObjectId(item_id)})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Get transaction history
    transactions = await db.inventory_transactions.find(
        {"item_id": item_id}
    ).sort("timestamp", -1).limit(50).to_list(50)
    
    return {
        "id": str(item["_id"]),
        "name": item["name"],
        "sku_code": item["sku_code"],
        "category": item["category"],
        "location_id": item.get("location_id"),
        "addon_group": item.get("addon_group"),
        "zone": item.get("zone", ""),
        "aisle": item.get("aisle", ""),
        "shelf": item.get("shelf", ""),
        "rack": item.get("rack", ""),
        "bin_location": item.get("bin_location", ""),
        "quantity": item["quantity"],
        "unit_price": item["unit_price"],
        "supplier": item.get("supplier"),
        "gst_percentage": item.get("gst_percentage", 18.0),
        "hsn_code": item.get("hsn_code"),
        "reorder_level": item.get("reorder_level", 10),
        "image_url": item.get("image_url"),
        "margin_pct": item.get("margin_pct", 0),
        "active": item.get("active", True),
        "qc_checklist": item.get("qc_checklist", []),
        "procurement_date": item.get("procurement_date"),
        "created_at": item["created_at"],
        "updated_at": item.get("updated_at"),
        "transactions": [
            {
                "id": str(t["_id"]),
                "type": t["type"],
                "quantity": t["quantity"],
                "project_id": t.get("project_id"),
                "notes": t.get("notes"),
                "user_name": t.get("user_name"),
                "timestamp": t["timestamp"]
            }
            for t in transactions
        ]
    }

@api_router.post("/inventory/items")
async def create_inventory_item(item: InventoryItemCreate, request: Request):
    current_user = await require_role("admin", "manager")(request)
    
    existing = await db.inventory_items.find_one({"sku_code": item.sku_code})
    if existing:
        raise HTTPException(status_code=400, detail="SKU code already exists")
    
    item_doc = {
        "name": item.name,
        "sku_code": item.sku_code,
        "category": item.category,
        "zone": item.zone or "",
        "aisle": item.aisle or "",
        "shelf": item.shelf or "",
        "rack": item.rack or "",
        "bin_location": item.bin_location or "",
        "quantity": item.quantity,
        "unit_price": item.unit_price,
        "supplier": item.supplier,
        "gst_percentage": item.gst_percentage,
        "hsn_code": item.hsn_code,
        "reorder_level": item.reorder_level,
        "image_url": item.image_url,
        "margin_pct": item.margin_pct,
        "active": item.active,
        "qc_checklist": item.qc_checklist,
        "procurement_date": item.procurement_date,
        "location_id": item.location_id or None,
        "addon_group": item.addon_group,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.inventory_items.insert_one(item_doc)
    
    # Create initial stock transaction
    await db.inventory_transactions.insert_one({
        "item_id": str(result.inserted_id),
        "type": "initial_stock",
        "quantity": item.quantity,
        "user_id": current_user["id"],
        "user_name": current_user["name"],
        "notes": "Initial stock entry",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    await create_audit_log(
        current_user["id"], current_user["name"], "create", "inventory_item",
        str(result.inserted_id), None, item_doc
    )
    
    return {"id": str(result.inserted_id), "message": "Item created successfully"}

@api_router.put("/inventory/items/{item_id}")
async def update_inventory_item(item_id: str, updates: InventoryItemUpdate, request: Request):
    current_user = await require_role("admin", "manager")(request)
    
    item = await db.inventory_items.find_one({"_id": ObjectId(item_id)})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    old_data = {k: item.get(k) for k in ["quantity", "unit_price"]}
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if updates.name is not None:
        update_data["name"] = updates.name
    if updates.zone is not None:
        update_data["zone"] = updates.zone
    if updates.aisle is not None:
        update_data["aisle"] = updates.aisle
    if updates.shelf is not None:
        update_data["shelf"] = updates.shelf
    if updates.rack is not None:
        update_data["rack"] = updates.rack
    if updates.bin_location is not None:
        update_data["bin_location"] = updates.bin_location
    if updates.image_url is not None:
        update_data["image_url"] = updates.image_url
    if updates.category is not None:
        update_data["category"] = updates.category
    if updates.location_id is not None:
        update_data["location_id"] = updates.location_id or None
    if updates.addon_group is not None:
        update_data["addon_group"] = updates.addon_group
    if updates.unit_price is not None:
        update_data["unit_price"] = updates.unit_price
    if updates.supplier is not None:
        update_data["supplier"] = updates.supplier
    if updates.gst_percentage is not None:
        update_data["gst_percentage"] = updates.gst_percentage
    if updates.hsn_code is not None:
        update_data["hsn_code"] = updates.hsn_code or None
    if updates.reorder_level is not None:
        update_data["reorder_level"] = updates.reorder_level
    if updates.margin_pct is not None:
        update_data["margin_pct"] = updates.margin_pct
    if updates.active is not None:
        update_data["active"] = updates.active
    if updates.qc_checklist is not None:
        update_data["qc_checklist"] = updates.qc_checklist
    if updates.procurement_date is not None:
        update_data["procurement_date"] = updates.procurement_date
    
    # Handle quantity adjustment
    if updates.quantity is not None and updates.quantity != item["quantity"]:
        quantity_diff = updates.quantity - item["quantity"]
        update_data["quantity"] = updates.quantity
        
        # Record transaction
        await db.inventory_transactions.insert_one({
            "item_id": item_id,
            "type": "adjustment",
            "quantity": quantity_diff,
            "user_id": current_user["id"],
            "user_name": current_user["name"],
            "notes": "Manual stock adjustment",
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
    
    await db.inventory_items.update_one(
        {"_id": ObjectId(item_id)},
        {"$set": update_data}
    )
    
    await create_audit_log(
        current_user["id"], current_user["name"], "update", "inventory_item",
        item_id, old_data, update_data
    )
    
    return {"message": "Item updated successfully"}

@api_router.delete("/inventory/items/{item_id}")
async def delete_inventory_item(item_id: str, request: Request):
    current_user = await require_role("admin")(request)
    
    item = await db.inventory_items.find_one({"_id": ObjectId(item_id)})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    await db.inventory_items.delete_one({"_id": ObjectId(item_id)})
    await db.inventory_transactions.delete_many({"item_id": item_id})
    
    await create_audit_log(
        current_user["id"], current_user["name"], "delete", "inventory_item", item_id
    )
    
    return {"message": "Item deleted successfully"}

# ===================== INVENTORY: IMPORT / EXPORT =====================

INVENTORY_EXPORT_COLUMNS = [
    "name", "sku_code", "category", "quantity", "unit_price", "reorder_level",
    "supplier", "gst_percentage", "hsn_code", "margin_pct", "zone", "aisle", "shelf", "rack",
    "bin_location", "procurement_date", "active",
]

@api_router.get("/inventory/template")
async def inventory_import_template(request: Request):
    """Download a blank XLSX template with the required columns."""
    await require_role("admin", "manager")(request)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = Workbook(); ws = wb.active; ws.title = "Inventory Template"
    headers = INVENTORY_EXPORT_COLUMNS
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="10B981")
    bold_white = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_white; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.append([
        "Solar Panel 540W Mono", "SP-540-MONO", "Panels", 50, 11500, 5,
        "ABC Solar Co", 18.0, "85414011", 12.5, "A", "1", "S2", "R3", "B4",
        "2026-01-15", True
    ])
    ws.append([])
    ws.append(["NOTE:", "sku_code must be unique. quantity/unit_price/reorder_level/gst_percentage/margin_pct are numeric. active is TRUE/FALSE. procurement_date is YYYY-MM-DD."])
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=12)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 35)
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="inventory_import_template.xlsx"'})


from inventory_import import (
    build_column_mapping, read_spreadsheet, validate_rows, REQUIRED_CANONICAL,
)


async def _load_import_file(file: UploadFile, column_map_json: Optional[str]):
    """Shared by preview + commit. Returns either a 'needs_mapping' dict or (df, mapping, clean_rows, error_rows)."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="This file could not be read. Save it as .xlsx or .csv and try again.")
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 10MB)")
    if not file.filename.lower().endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="This file could not be read. Save it as .xlsx or .csv and try again.")
    try:
        df, raw_columns = read_spreadsheet(raw, file.filename)
    except Exception:
        raise HTTPException(status_code=400, detail="This file could not be read. Save it as .xlsx or .csv and try again.")
    if df.empty:
        raise HTTPException(status_code=400, detail="File has no rows")

    overrides = json.loads(column_map_json) if column_map_json else None
    mapping = build_column_mapping(raw_columns, overrides)
    unmapped_required = [c for c in REQUIRED_CANONICAL if not mapping.get(c)]
    if unmapped_required:
        return {
            "status": "needs_mapping",
            "detected_columns": raw_columns,
            "column_mapping": mapping,
            "unmapped_required": unmapped_required,
            "sample_rows": df.head(5).fillna("").astype(str).to_dict(orient="records"),
            "total_rows": int(len(df)),
        }
    clean_rows, error_rows = validate_rows(df, mapping)
    return df, mapping, clean_rows, error_rows


@api_router.post("/inventory/import/preview")
async def inventory_import_preview(request: Request, file: UploadFile = File(...), column_map: Optional[str] = Form(None)):
    """Parse + validate only — never writes to the database. Shows the first 20 rows with a per-row status."""
    await require_role("admin", "manager")(request)
    result = await _load_import_file(file, column_map)
    if isinstance(result, dict):
        return result
    df, mapping, clean_rows, error_rows = result

    existing_skus = set()
    if clean_rows:
        docs = await db.inventory_items.find({"sku_code": {"$in": [r["sku_code"] for r in clean_rows]}}, {"sku_code": 1}).to_list(5000)
        existing_skus = {d["sku_code"] for d in docs}

    preview_rows = []
    for r in clean_rows[:20]:
        will_update = r["sku_code"] in existing_skus
        preview_rows.append({**r, "status": "will_update" if will_update else "will_create", "reason": None})
    for e in error_rows[:20 - len(preview_rows)] if len(preview_rows) < 20 else []:
        preview_rows.append({"row": e["row"], "status": "will_skip", "reason": e["error"]})

    will_create = sum(1 for r in clean_rows if r["sku_code"] not in existing_skus)
    will_update = sum(1 for r in clean_rows if r["sku_code"] in existing_skus)
    return {
        "status": "ready",
        "column_mapping": mapping,
        "total_rows": int(len(df)),
        "preview_rows": preview_rows,
        "summary": {"will_create": will_create, "will_update": will_update, "will_skip": len(error_rows)},
        "errors": error_rows,
    }


@api_router.post("/inventory/import")
async def inventory_import(request: Request, file: UploadFile = File(...),
                            dry_run: Optional[str] = Form(None), column_map: Optional[str] = Form(None)):
    """Bulk import inventory items from XLSX/CSV. Existing SKUs are updated; new SKUs are created."""
    current_user = await require_role("admin", "manager")(request)
    result = await _load_import_file(file, column_map)
    if isinstance(result, dict):
        return result
    df, mapping, clean_rows, error_rows = result
    is_dry_run = str(dry_run).strip().lower() in ("true", "1", "yes")

    created, updated = 0, 0
    now_iso = datetime.now(timezone.utc).isoformat()
    for r in clean_rows:
        sku = r["sku_code"]
        doc = {k: v for k, v in r.items() if k != "row"}
        doc["updated_at"] = now_iso
        if is_dry_run:
            existing = await db.inventory_items.find_one({"sku_code": sku}, {"_id": 1})
            updated += 1 if existing else 0
            created += 0 if existing else 1
            continue
        existing = await db.inventory_items.find_one({"sku_code": sku})
        if existing:
            await db.inventory_items.update_one({"_id": existing["_id"]}, {"$set": doc})
            updated += 1
        else:
            doc["created_at"] = now_iso
            doc["qc_checklist"] = []
            res = await db.inventory_items.insert_one(doc)
            await db.inventory_transactions.insert_one({
                "item_id": str(res.inserted_id),
                "transaction_type": "purchase",
                "quantity": doc["quantity"],
                "previous_quantity": 0,
                "new_quantity": doc["quantity"],
                "performed_by_id": current_user["id"],
                "performed_by_name": current_user["name"],
                "notes": "Bulk import",
                "timestamp": now_iso,
            })
            created += 1

    if not is_dry_run:
        await create_audit_log(current_user["id"], current_user["name"], "import", "inventory_item",
                               f"created={created} updated={updated} errors={len(error_rows)}")
    verb = "Would import" if is_dry_run else "Imported"
    return {"status": "done", "dry_run": is_dry_run, "created": created, "updated": updated, "errors": error_rows,
            "total_rows": int(len(df)),
            "message": f"{verb} {created} new and {'would update' if is_dry_run else 'updated'} {updated} existing item(s)"
                        + (f" — {len(error_rows)} row(s) skipped" if error_rows else "")}


@api_router.get("/inventory/export")
async def inventory_export(request: Request, format: str = "xlsx"):
    """Export full inventory as XLSX or PDF."""
    await require_role("admin", "manager", "staff")(request)
    items = await db.inventory_items.find({}).to_list(5000)

    fmt = (format or "xlsx").lower()
    from io import BytesIO

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook(); ws = wb.active; ws.title = "Inventory"
        headers = INVENTORY_EXPORT_COLUMNS
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="10B981")
        bold_white = Font(bold=True, color="FFFFFF")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold_white; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for item in items:
            row_vals = []
            for c in headers:
                v = item.get(c, "")
                if isinstance(v, bool):
                    row_vals.append("TRUE" if v else "FALSE")
                else:
                    row_vals.append(v)
            ws.append(row_vals)
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 35)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="inventory_{datetime.now(timezone.utc).date()}.xlsx"'})

    if fmt == "pdf":
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            raise HTTPException(status_code=500, detail="reportlab not installed on backend")

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12, rightMargin=12, topMargin=14, bottomMargin=14)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#0F172A'))
        sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#64748B'))
        story = [
            Paragraph("Sensoper - Inventory Export", title_style),
            Paragraph(f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} - {len(items)} items", sub_style),
            Spacer(1, 6),
        ]
        cols = ["SKU", "Name", "Category", "Qty", "Unit Price", "Reorder", "Supplier", "Zone", "Active"]
        data = [cols]
        for it in items:
            data.append([
                str(it.get("sku_code", ""))[:24],
                str(it.get("name", ""))[:38],
                str(it.get("category", ""))[:18],
                str(it.get("quantity", "")),
                f'Rs.{it.get("unit_price", 0):,.0f}',
                str(it.get("reorder_level", "")),
                str(it.get("supplier", ""))[:18],
                str(it.get("zone", ""))[:8],
                "Yes" if it.get("active") else "No",
            ])
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(tbl)
        doc.build(story); buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="inventory_{datetime.now(timezone.utc).date()}.pdf"'})

    raise HTTPException(status_code=400, detail="format must be 'xlsx' or 'pdf'")


@api_router.get("/inventory/alerts")
async def get_inventory_alerts(request: Request, location_id: Optional[str] = None):
    """Get all low stock alerts"""
    user = await get_current_user(request)
    
    # Use aggregation to find items below reorder level
    match_stage = {"$expr": {"$lte": ["$quantity", "$reorder_level"]}}
    loc_filter = location_scope_filter(user, location_id)
    if loc_filter:
        match_stage = {"$and": [match_stage, loc_filter]}
    pipeline = [{"$match": match_stage}]
    
    items = await db.inventory_items.aggregate(pipeline).to_list(100)
    
    return [
        {
            "id": str(item["_id"]),
            "name": item["name"],
            "sku_code": item["sku_code"],
            "category": item["category"],
            "zone": item.get("zone", ""),
            "location_id": item.get("location_id"),
            "quantity": item["quantity"],
            "reorder_level": item.get("reorder_level", 10)
        }
        for item in items
    ]

# ================== MATERIAL KITS (Solution Kits) ==================

def _serialize_kit(k):
    return {
        "id": str(k["_id"]),
        "name": k.get("name"),
        "system_type": k.get("system_type"),
        "capacity_kw": k.get("capacity_kw", 0),
        "capacity_min_kw": k.get("capacity_min_kw"),
        "capacity_max_kw": k.get("capacity_max_kw"),
        "description": k.get("description"),
        "lines": k.get("lines", []),
        "active": k.get("active", True),
        "created_at": k.get("created_at"),
        "updated_at": k.get("updated_at"),
    }

@api_router.get("/material-kits")
async def list_material_kits(request: Request, system_type: Optional[str] = None, capacity_kw: Optional[float] = None):
    await get_current_user(request)
    query = {"active": True}
    if system_type:
        query["system_type"] = system_type
    kits = await db.material_kits.find(query).sort("capacity_kw", 1).to_list(500)
    result = [_serialize_kit(k) for k in kits]
    # If a capacity_kw is specified, order by proximity so first item is best-match
    if capacity_kw is not None:
        result.sort(key=lambda r: abs(float(r.get("capacity_kw") or 0) - float(capacity_kw)))
    return result

@api_router.get("/material-kits/match")
async def match_material_kit(request: Request, system_type: str, capacity_kw: float):
    """Return the best-matching kit for a given system_type + capacity.
    Preference: capacity within [min,max], else closest capacity_kw."""
    await get_current_user(request)
    kits = await db.material_kits.find({"system_type": system_type, "active": True}).to_list(500)
    if not kits:
        return {"match": None, "candidates": []}
    # Range match first
    within = [k for k in kits
              if (k.get("capacity_min_kw") is not None and k.get("capacity_max_kw") is not None
                  and float(k["capacity_min_kw"]) <= capacity_kw <= float(k["capacity_max_kw"]))]
    if within:
        best = min(within, key=lambda k: abs(float(k.get("capacity_kw") or 0) - capacity_kw))
    else:
        best = min(kits, key=lambda k: abs(float(k.get("capacity_kw") or 0) - capacity_kw))
    return {"match": _serialize_kit(best), "candidates": [_serialize_kit(k) for k in kits]}

@api_router.get("/material-kits/{kit_id}")
async def get_material_kit(kit_id: str, request: Request):
    await get_current_user(request)
    kit = await db.material_kits.find_one({"_id": ObjectId(kit_id)})
    if not kit:
        raise HTTPException(status_code=404, detail="Kit not found")
    return _serialize_kit(kit)

@api_router.post("/material-kits")
async def create_material_kit(kit: MaterialKitCreate, request: Request):
    current_user = await require_role("admin", "manager")(request)
    doc = kit.model_dump()
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["updated_at"] = doc["created_at"]
    doc["created_by"] = current_user["id"]
    result = await db.material_kits.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Kit created"}

@api_router.put("/material-kits/{kit_id}")
async def update_material_kit(kit_id: str, updates: MaterialKitUpdate, request: Request):
    await require_role("admin", "manager")(request)
    kit = await db.material_kits.find_one({"_id": ObjectId(kit_id)})
    if not kit:
        raise HTTPException(status_code=404, detail="Kit not found")
    upd = {k: v for k, v in updates.model_dump(exclude_unset=True).items() if v is not None}
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.material_kits.update_one({"_id": ObjectId(kit_id)}, {"$set": upd})
    return {"message": "Kit updated"}

@api_router.delete("/material-kits/{kit_id}")
async def delete_material_kit(kit_id: str, request: Request):
    await require_role("admin", "manager")(request)
    res = await db.material_kits.delete_one({"_id": ObjectId(kit_id)})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kit not found")
    return {"message": "Kit deleted"}

@api_router.post("/material-kits/seed-starter")
async def seed_material_kits(request: Request):
    """Idempotent seed of 8 starter kits — 2 per system type across common capacities."""
    await require_role("admin", "manager")(request)
    starter = [
        # On-Grid
        {"name": "On-Grid Starter · 3 kW", "system_type": "on-grid", "capacity_kw": 3,
         "capacity_min_kw": 2, "capacity_max_kw": 4,
         "description": "Residential on-grid rooftop with net-metering (2-4 kW).",
         "lines": [
             {"name": "540W Mono PERC panels", "category": "panels", "quantity": 6, "qty_formula": "1 per 0.5 kW"},
             {"name": "3 kW String Inverter", "category": "inverter", "quantity": 1},
             {"name": "DC/AC combiner box", "category": "combiner", "quantity": 1},
             {"name": "4 sqmm DC cable", "category": "cables", "quantity": 60, "qty_formula": "10m per panel"},
             {"name": "MC4 connectors (pair)", "category": "connectors", "quantity": 12},
             {"name": "GI mounting structure", "category": "mounting", "quantity": 6},
             {"name": "Earthing kit + LA", "category": "earthing", "quantity": 1},
         ]},
        {"name": "On-Grid Family · 5 kW", "system_type": "on-grid", "capacity_kw": 5,
         "capacity_min_kw": 4, "capacity_max_kw": 7,
         "description": "Family home on-grid (4-7 kW).",
         "lines": [
             {"name": "540W Mono PERC panels", "category": "panels", "quantity": 10},
             {"name": "5 kW String Inverter", "category": "inverter", "quantity": 1},
             {"name": "DC/AC combiner box", "category": "combiner", "quantity": 1},
             {"name": "6 sqmm DC cable", "category": "cables", "quantity": 100},
             {"name": "MC4 connectors (pair)", "category": "connectors", "quantity": 20},
             {"name": "GI mounting structure", "category": "mounting", "quantity": 10},
             {"name": "Earthing kit + LA", "category": "earthing", "quantity": 1},
         ]},
        # Off-Grid
        {"name": "Off-Grid Cabin · 3 kW", "system_type": "off-grid", "capacity_kw": 3,
         "capacity_min_kw": 2, "capacity_max_kw": 4,
         "description": "Standalone off-grid with C10 batteries (2-4 kW).",
         "lines": [
             {"name": "540W Mono PERC panels", "category": "panels", "quantity": 6},
             {"name": "3 kW Off-Grid Inverter", "category": "inverter", "quantity": 1},
             {"name": "150Ah C10 Tubular battery", "category": "battery", "quantity": 4},
             {"name": "MPPT charge controller 60A", "category": "charge_controller", "quantity": 1},
             {"name": "4 sqmm DC cable", "category": "cables", "quantity": 60},
             {"name": "MC4 connectors (pair)", "category": "connectors", "quantity": 12},
             {"name": "GI mounting structure", "category": "mounting", "quantity": 6},
             {"name": "Battery cables + lugs", "category": "cables", "quantity": 1},
         ]},
        {"name": "Off-Grid Farmhouse · 5 kW", "system_type": "off-grid", "capacity_kw": 5,
         "capacity_min_kw": 4, "capacity_max_kw": 7,
         "description": "Extended off-grid autonomy (4-7 kW).",
         "lines": [
             {"name": "540W Mono PERC panels", "category": "panels", "quantity": 10},
             {"name": "5 kW Off-Grid Inverter", "category": "inverter", "quantity": 1},
             {"name": "200Ah C10 Tubular battery", "category": "battery", "quantity": 6},
             {"name": "MPPT charge controller 100A", "category": "charge_controller", "quantity": 1},
             {"name": "6 sqmm DC cable", "category": "cables", "quantity": 100},
             {"name": "GI mounting structure", "category": "mounting", "quantity": 10},
         ]},
        # Hybrid
        {"name": "Hybrid Home · 5 kW", "system_type": "hybrid", "capacity_kw": 5,
         "capacity_min_kw": 4, "capacity_max_kw": 7,
         "description": "Grid-tied with battery backup (4-7 kW).",
         "lines": [
             {"name": "540W Mono PERC panels", "category": "panels", "quantity": 10},
             {"name": "5 kW Hybrid Inverter", "category": "inverter", "quantity": 1},
             {"name": "5 kWh LiFePO4 battery", "category": "battery", "quantity": 1},
             {"name": "DC/AC combiner box", "category": "combiner", "quantity": 1},
             {"name": "6 sqmm DC cable", "category": "cables", "quantity": 100},
             {"name": "GI mounting structure", "category": "mounting", "quantity": 10},
             {"name": "Earthing kit + LA", "category": "earthing", "quantity": 1},
         ]},
        {"name": "Hybrid Villa · 10 kW", "system_type": "hybrid", "capacity_kw": 10,
         "capacity_min_kw": 7, "capacity_max_kw": 15,
         "description": "Large hybrid with battery + net-metering (7-15 kW).",
         "lines": [
             {"name": "540W Mono PERC panels", "category": "panels", "quantity": 19},
             {"name": "10 kW Hybrid Inverter", "category": "inverter", "quantity": 1},
             {"name": "10 kWh LiFePO4 battery", "category": "battery", "quantity": 1},
             {"name": "10 sqmm DC cable", "category": "cables", "quantity": 200},
             {"name": "GI mounting structure", "category": "mounting", "quantity": 19},
             {"name": "Earthing kit + LA", "category": "earthing", "quantity": 1},
         ]},
        # Solar Pump
        {"name": "Solar Pump · 3 HP Submersible", "system_type": "solar-pump", "capacity_kw": 2.2,
         "capacity_min_kw": 1.5, "capacity_max_kw": 3,
         "description": "3 HP submersible pump kit (up to 50m head, ~10 kLPH).",
         "lines": [
             {"name": "330W Poly panels", "category": "panels", "quantity": 8},
             {"name": "3 HP DC Solar Pump Controller", "category": "controller", "quantity": 1},
             {"name": "3 HP Submersible Pump", "category": "pump", "quantity": 1},
             {"name": "4 sqmm DC cable", "category": "cables", "quantity": 80},
             {"name": "Delivery pipe (32mm)", "category": "pipe", "quantity": 50, "qty_formula": "per m of head"},
             {"name": "GI mounting structure", "category": "mounting", "quantity": 8},
         ]},
        {"name": "Solar Pump · 5 HP Surface", "system_type": "solar-pump", "capacity_kw": 3.7,
         "capacity_min_kw": 3, "capacity_max_kw": 5.5,
         "description": "5 HP surface pump for open wells / canals.",
         "lines": [
             {"name": "540W Mono PERC panels", "category": "panels", "quantity": 8},
             {"name": "5 HP AC Solar Pump Controller (VFD)", "category": "controller", "quantity": 1},
             {"name": "5 HP Surface AC Pump", "category": "pump", "quantity": 1},
             {"name": "6 sqmm DC cable", "category": "cables", "quantity": 80},
             {"name": "Delivery pipe (50mm)", "category": "pipe", "quantity": 50},
             {"name": "GI mounting structure", "category": "mounting", "quantity": 8},
         ]},
    ]
    created = 0
    for k in starter:
        existing = await db.material_kits.find_one({"name": k["name"], "system_type": k["system_type"]})
        if existing:
            continue
        k["active"] = True
        k["created_at"] = datetime.now(timezone.utc).isoformat()
        k["updated_at"] = k["created_at"]
        await db.material_kits.insert_one(k)
        created += 1
    total = await db.material_kits.count_documents({})
    return {"created": created, "total": total}

# ================== SOLAR CALCULATION ENGINE ==================

from calculators import (
    calculate_solution as _calc_solution,
    lookup_pincode as _lookup_pincode,
    compute_bill_from_slabs as _compute_bill,
    back_solve_units as _back_solve,
)
from calculators.seed_data import get_default_discoms, get_default_pincodes

# ═══════════ DIRECT SALES ROUTER (Iter 39 Change 1) ═══════════
from sales import create_router as _create_sales_router


async def _get_active_company_profile(location_id: Optional[str] = None):
    """Iter 43 Change 3b — a location-scoped profile (own GST/state) wins over the global active one."""
    if location_id:
        loc_profile = await db.company_profiles.find_one({"location_id": location_id, "is_active": True})
        if loc_profile:
            return loc_profile
    return await db.company_profiles.find_one({"is_active": True}) or {}


_sales_router = _create_sales_router(
    db=db, get_current_user=get_current_user, require_role=require_role,
    company_profile_fn=_get_active_company_profile, check_module_permission=check_module_permission,
)
api_router.include_router(_sales_router)

# ═══════════ EXCESS MATERIAL RECONCILIATION (Iter 42 Change 4) ═══════════
from reconciliation import create_router as _create_reconciliation_router
_reconciliation_router = _create_reconciliation_router(
    db=db, get_current_user=get_current_user, require_role=require_role,
    create_audit_log=create_audit_log,
)
api_router.include_router(_reconciliation_router)

# ═══════════ ASSETS & TOOLS (Iter 42 Change 6) ═══════════
from assets import create_router as _create_assets_router
_assets_router = _create_assets_router(
    db=db, get_current_user=get_current_user, require_role=require_role,
    create_audit_log=create_audit_log, check_module_permission=check_module_permission,
)
api_router.include_router(_assets_router)

# ═══════════ AMC — RECURRING REVENUE (Iter 42 Change 5) ═══════════
from amc import create_router as _create_amc_router
_amc_router = _create_amc_router(
    db=db, get_current_user=get_current_user, require_role=require_role,
    create_audit_log=create_audit_log,
)
api_router.include_router(_amc_router)

# ═══════════ MULTI-LOCATION ACCESS CONTROL (Iter 42 Change 8) ═══════════
from locations import create_router as _create_locations_router, location_scope_filter
_locations_router = _create_locations_router(
    db=db, get_current_user=get_current_user, require_role=require_role,
    create_audit_log=create_audit_log,
)
api_router.include_router(_locations_router)

# ═══════════ CATALOGUE + FUEL MODEL (Iter 44 Phase 1 — Changes 6 & 7) ═══════════
from catalogue import attach as _attach_catalogue  # noqa: E402
_attach_catalogue(api_router, db, get_current_user)


# ═══════════ SUBSIDY TRACKING (Iter 39 Change 2c) ═══════════

class SubsidyTracking(BaseModel):
    project_id: str
    scheme: Optional[str] = "pm_surya_ghar"    # pm_surya_ghar|pm_kusum_b|pm_kusum_c|state_scheme|none
    eligible_amount: Optional[float] = 0
    claimed_amount: Optional[float] = 0
    approved_amount: Optional[float] = 0
    disbursed_amount: Optional[float] = 0
    status: Optional[str] = "eligible"          # eligible|application_pending|applied|under_review|approved|disbursed|rejected|not_applicable
    application_number: Optional[str] = None
    application_date: Optional[str] = None
    approval_date: Optional[str] = None
    disbursement_date: Optional[str] = None
    discom_inspection_date: Optional[str] = None
    inspection_status: Optional[str] = None
    net_meter_installation_date: Optional[str] = None
    rejection_reason: Optional[str] = None
    documents: Optional[List[str]] = []
    notes: Optional[str] = None


@api_router.get("/subsidy/tracking/{project_id}")
async def get_subsidy_tracking(project_id: str, request: Request):
    await get_current_user(request)
    doc = await db.subsidy_tracking.find_one({"project_id": project_id})
    if not doc: return {"project_id": project_id, "status": "not_started"}
    return {k: v for k, v in doc.items() if k != "_id"}


@api_router.post("/subsidy/tracking")
async def upsert_subsidy_tracking(payload: SubsidyTracking, request: Request):
    user = await get_current_user(request)
    body = payload.model_dump(exclude_unset=True)
    body["updated_at"] = datetime.now(timezone.utc).isoformat()
    body["updated_by"] = user.get("id")
    # Merge with existing so days_to_disburse can be computed on incremental updates
    existing = await db.subsidy_tracking.find_one({"project_id": payload.project_id}) or {}
    merged_app = body.get("application_date") or existing.get("application_date")
    merged_disb = body.get("disbursement_date") or existing.get("disbursement_date")
    if merged_app and merged_disb:
        try:
            d1 = datetime.fromisoformat(merged_app.replace("Z", "+00:00"))
            d2 = datetime.fromisoformat(merged_disb.replace("Z", "+00:00"))
            body["days_to_disburse"] = (d2 - d1).days
        except Exception: pass
    await db.subsidy_tracking.update_one(
        {"project_id": payload.project_id},
        {"$set": body, "$setOnInsert": {"created_at": body["updated_at"]}},
        upsert=True,
    )
    return {"message": "saved"}


@api_router.get("/subsidy/analytics")
async def subsidy_analytics(request: Request):
    user = await get_current_user(request)
    if user["role"] not in ("admin", "manager"):
        raise HTTPException(403, "Forbidden")
    docs = await db.subsidy_tracking.find({}).to_list(5000)
    total_eligible = sum(d.get("eligible_amount", 0) for d in docs)
    total_claimed = sum(d.get("claimed_amount", 0) for d in docs)
    total_approved = sum(d.get("approved_amount", 0) for d in docs)
    total_disbursed = sum(d.get("disbursed_amount", 0) for d in docs)
    by_scheme = {}
    by_status = {}
    stuck = []
    now = datetime.now(timezone.utc)
    threshold_days = 60
    for d in docs:
        s = d.get("scheme") or "unknown"
        by_scheme[s] = by_scheme.get(s, 0) + d.get("disbursed_amount", 0)
        st = d.get("status") or "eligible"
        by_status[st] = by_status.get(st, 0) + 1
        # Stuck detection
        if st in ("applied", "under_review") and d.get("application_date"):
            try:
                appd = datetime.fromisoformat(d["application_date"].replace("Z", "+00:00"))
                if (now - appd).days > threshold_days:
                    stuck.append({"project_id": d.get("project_id"), "status": st,
                                  "days": (now - appd).days,
                                  "amount": d.get("claimed_amount", 0)})
            except Exception: pass
    rejections = [d for d in docs if d.get("status") == "rejected"]
    days = [d["days_to_disburse"] for d in docs if d.get("days_to_disburse")]
    return {
        "total_eligible": _round_v(total_eligible), "total_claimed": _round_v(total_claimed),
        "total_approved": _round_v(total_approved), "total_disbursed": _round_v(total_disbursed),
        "by_scheme": by_scheme, "by_status": by_status,
        "stuck_applications": stuck[:50],
        "rejection_count": len(rejections),
        "rejection_reasons": [{"reason": r.get("rejection_reason", "?"), "count": 1} for r in rejections[:20]],
        "avg_days_to_disburse": round(sum(days) / len(days), 1) if days else None,
        "conversion_pct": round((total_disbursed / total_eligible * 100), 1) if total_eligible else 0,
        "count": len(docs),
    }


def _round_v(x):
    try: return round(float(x), 2)
    except Exception: return 0


# ═══════════ MARKETING SUMMARY + CAC REPORT (Iter 39 Change 3) ═══════════

@api_router.get("/accounts/marketing-summary")
async def marketing_summary(request: Request, start: Optional[str] = None, end: Optional[str] = None):
    await get_current_user(request)
    now = datetime.now(timezone.utc)
    if not start: start = (now - timedelta(days=90)).date().isoformat()
    if not end: end = now.date().isoformat()
    entries = await db.account_entries.find({
        "entry_type": "marketing_expense",
        "entry_date": {"$gte": start, "$lte": end}
    }).to_list(5000)
    total = sum(e.get("amount", 0) for e in entries)
    by_channel, by_campaign, by_district, by_month = {}, {}, {}, {}
    for e in entries:
        ch = e.get("marketing_channel", "other")
        by_channel[ch] = by_channel.get(ch, 0) + e.get("amount", 0)
        cp = e.get("campaign_name", "-")
        by_campaign[cp] = by_campaign.get(cp, 0) + e.get("amount", 0)
        dist = e.get("target_district", "-")
        by_district[dist] = by_district.get(dist, 0) + e.get("amount", 0)
        m = (e.get("entry_date") or "")[:7]
        by_month[m] = by_month.get(m, 0) + e.get("amount", 0)
    return {
        "period_start": start, "period_end": end,
        "total_spend": _round_v(total),
        "entry_count": len(entries),
        "by_channel": by_channel, "by_campaign": by_campaign,
        "by_district": by_district, "by_month": by_month,
    }


@api_router.get("/reports/cac")
async def cac_report(request: Request, start: Optional[str] = None, end: Optional[str] = None,
                     attribution_window_days: int = 90):
    user = await get_current_user(request)
    if user["role"] not in ("admin", "manager"):
        raise HTTPException(403, "Forbidden")
    now = datetime.now(timezone.utc)
    if not start: start = (now - timedelta(days=365)).date().isoformat()
    if not end: end = now.date().isoformat()

    # Total marketing spend
    m_entries = await db.account_entries.find({
        "entry_type": "marketing_expense",
        "entry_date": {"$gte": start, "$lte": end}
    }).to_list(5000)
    total_spend = sum(e.get("amount", 0) for e in m_entries)
    spend_by_channel = {}
    for e in m_entries:
        ch = e.get("marketing_channel", "other")
        spend_by_channel[ch] = spend_by_channel.get(ch, 0) + e.get("amount", 0)

    # New customers acquired (projects + direct sales)
    proj_query = {"created_at": {"$gte": start + "T00:00:00", "$lte": end + "T23:59:59"},
                  "status": {"$in": ["approved", "completed"]},
                  "deleted_at": {"$exists": False}}
    projects = await db.projects.find(proj_query).to_list(5000)
    sales = await db.sales.find({"sale_date": {"$gte": start, "$lte": end},
                                  "status": {"$ne": "cancelled"}}).to_list(5000)

    # Deduplicate by phone (unified customer)
    customers = {}
    unattributed = 0
    by_channel_customers = {}
    total_revenue = 0
    for p in projects:
        phone = (p.get("customer") or {}).get("phone", "?")
        src = (p.get("custom_fields", {}) or {}).get("lead_source") or p.get("lead_source") or "unattributed"
        rev = (p.get("cost_estimation") or {}).get("total_cost", 0)
        total_revenue += rev
        if phone not in customers:
            customers[phone] = {"channel": src, "revenue": rev, "type": "project"}
        else:
            customers[phone]["revenue"] += rev
    for s in sales:
        phone = (s.get("customer") or {}).get("phone", "?")
        src = s.get("lead_source") or "unattributed"
        rev = s.get("grand_total", 0)
        total_revenue += rev
        if phone not in customers:
            customers[phone] = {"channel": src, "revenue": rev, "type": "sale"}
        else:
            customers[phone]["revenue"] += rev

    for phone, c in customers.items():
        ch = c["channel"] or "unattributed"
        if ch == "unattributed":
            unattributed += 1
        by_channel_customers[ch] = by_channel_customers.get(ch, 0) + 1

    total_customers = len(customers)
    paid_customers = sum(v for k, v in by_channel_customers.items() if k != "unattributed" and k not in ("referral", "organic"))

    def _safe(a, b): return _round_v(a / b) if b else None

    blended_cac = _safe(total_spend, total_customers)
    paid_cac = _safe(total_spend, paid_customers)

    channel_perf = []
    for ch, spend in spend_by_channel.items():
        n = by_channel_customers.get(ch, 0)
        ch_rev = sum(c["revenue"] for c in customers.values() if c["channel"] == ch)
        channel_perf.append({
            "channel": ch,
            "spend": _round_v(spend),
            "customers": n,
            "revenue": _round_v(ch_rev),
            "cac": _safe(spend, n),
            "roi": _safe(ch_rev - spend, spend) if spend else None,
        })
    channel_perf.sort(key=lambda r: -(r["revenue"] or 0))

    # LTV = average revenue per customer (proxy)
    ltv = _round_v(total_revenue / total_customers) if total_customers else 0
    ltv_cac_ratio = _safe(ltv, blended_cac) if blended_cac else None

    return {
        "period_start": start, "period_end": end,
        "attribution_window_days": attribution_window_days,
        "total_spend": _round_v(total_spend),
        "total_customers": total_customers,
        "unattributed_customers": unattributed,
        "unattributed_pct": _round_v(unattributed / total_customers * 100) if total_customers else 0,
        "blended_cac": blended_cac,
        "paid_cac": paid_cac,
        "ltv": ltv,
        "ltv_cac_ratio": ltv_cac_ratio,
        "marketing_pct_of_revenue": _round_v(total_spend / total_revenue * 100) if total_revenue else 0,
        "channels": channel_perf,
        "by_channel_customers": by_channel_customers,
        "spend_by_channel": {k: _round_v(v) for k, v in spend_by_channel.items()},
        "total_revenue": _round_v(total_revenue),
    }



class CalculateSolutionRequest(BaseModel):
    system_type: str
    pincode: Optional[str] = None
    discom_id: Optional[str] = None
    inputs: Dict[str, Any] = {}
    overrides: Dict[str, Any] = {}


async def _get_calc_config() -> Dict[str, Any]:
    doc = await db.calc_config.find_one({"_id": "singleton"})
    if not doc:
        default = {
            "_id": "singleton",
            "default_specific_yield": 4.4,
            "cost_per_kwp": {"on-grid": 55000, "hybrid": 75000, "off-grid": 95000, "solar-pump": 65000},
            "battery_unit_kwh": 5,
            "system_life_years": 25,
            "panel_degradation_pct_per_year": 0.7,
            "pm_surya_ghar": {
                "cap": 78000,
                "slabs": [
                    {"upto_kw": 1, "amount": 30000},
                    {"upto_kw": 2, "amount": 48000},
                    {"upto_kw": 3, "amount": 78000},
                ],
            },
            "pm_kusum": {"benchmark_per_kw": 40000},
            "diesel_price_per_liter": 95,
            "diesel_lph_per_kw": 0.3,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.calc_config.insert_one(default)
        return default
    return doc


async def _get_pincode_and_discom(pincode: Optional[str], discom_id: Optional[str]):
    """Return (pin_dict_or_None, discom_dict_or_None)."""
    pin_doc = None
    discom_doc = None
    if pincode:
        pin_doc = await db.pincodes.find_one({"pincode": pincode})
    # DISCOM priority: explicit discom_id > pin's discom > FALLBACK
    if discom_id:
        discom_doc = await db.discoms.find_one({"$or": [{"id": discom_id}, {"short_code": discom_id}]})
    if not discom_doc and pin_doc and pin_doc.get("discom"):
        discom_doc = await db.discoms.find_one({"$or": [{"id": pin_doc["discom"]}, {"short_code": pin_doc["discom"]}]})
    if not discom_doc:
        discom_doc = await db.discoms.find_one({"id": "FALLBACK"})
    return pin_doc, discom_doc


@api_router.get("/calculate/lookup/{pincode}")
async def calc_lookup_pincode(pincode: str, request: Request):
    """Debounced endpoint — returns district/state/DISCOM/yield for a PIN."""
    await get_current_user(request)
    # Preload all discoms into a code-keyed map
    all_discoms = await db.discoms.find({}).to_list(500)
    discoms_by_code = {d["id"]: d for d in all_discoms}
    pin_doc = await db.pincodes.find_one({"pincode": pincode})
    pincodes_map = {pin_doc["pincode"]: pin_doc} if pin_doc else {}
    resolved = _lookup_pincode(pincode, pincodes_map, discoms_by_code)
    # Present only category names + fixed_charge in the light response
    if resolved.get("categories"):
        resolved["categories"] = [
            {"name": c.get("name"), "fixed_charge": c.get("fixed_charge", 0),
             "export_rate": c.get("export_rate", 0),
             "slab_count": len(c.get("slabs", []))}
            for c in resolved["categories"]
        ]
    return resolved


@api_router.post("/calculate/solution")
async def calc_solution(payload: CalculateSolutionRequest, request: Request):
    await get_current_user(request)
    config = await _get_calc_config()
    pin_doc, discom_doc = await _get_pincode_and_discom(payload.pincode, payload.discom_id)
    if pin_doc:
        pin_doc = {k: v for k, v in pin_doc.items() if k != "_id"}
    if discom_doc:
        discom_doc = {k: v for k, v in discom_doc.items() if k != "_id"}
    computed = _calc_solution(
        system_type=payload.system_type,
        pincode=payload.pincode,
        inputs=payload.inputs,
        overrides=payload.overrides,
        config=config,
        discom=discom_doc,
        pin=pin_doc,
    )
    # Snapshot the versioned constants used, so a project saving this can replay it later
    computed["snapshot"] = {
        "computed_at": datetime.now(timezone.utc).isoformat(),
        "config_version": config.get("updated_at"),
        "pincode": payload.pincode,
        "discom_id": (discom_doc or {}).get("id") or (discom_doc or {}).get("short_code"),
        "discom_effective_from": (discom_doc or {}).get("effective_from"),
        "specific_yield_used": computed["result"].get("specific_yield_used"),
        "cost_per_kwp_used": computed["result"].get("cost_per_kwp_used"),
    }
    # 4-stage guided-flow trace (Iter 44 Phase 2 — Change 1) — for Show Working UI
    try:
        from calculators.working import compile_stages_ongrid, compile_stages_pump, pump_roi
        st = (payload.system_type or "on-grid").lower()
        if st == "on-grid":
            computed["stages"] = compile_stages_ongrid(
                computed["result"], computed.get("breakdown", []), payload.inputs, config
            )
        elif st in ("solar-pump", "pump-dc", "pump-ac"):
            # Resolve ROI mode + fuel for pump
            mode = (payload.inputs.get("roi_mode") or "diesel").lower()
            fuel_doc = None
            fuel_id = payload.inputs.get("fuel_id")
            if fuel_id:
                try:
                    fuel_doc = await db.fuel_types.find_one({"_id": ObjectId(fuel_id)})
                except Exception:
                    fuel_doc = None
            if not fuel_doc:
                fuel_doc = await db.fuel_types.find_one({"name": "Diesel"})
            if fuel_doc:
                fuel_doc = {k: v for k, v in fuel_doc.items() if k != "_id"}
            roi_params = {
                "fuel_price": payload.inputs.get("fuel_price_per_unit"),
                "tariff_per_unit": payload.inputs.get("existing_ag_tariff"),
                "hours_per_year": payload.inputs.get("hire_hours_per_year"),
                "hire_rate": payload.inputs.get("hire_rate"),
                "crop_value_per_year": payload.inputs.get("crop_value_per_year"),
                "hours_gained_per_year": payload.inputs.get("hours_gained_per_year"),
            }
            roi = pump_roi(
                input_kw=computed["result"].get("input_power_kw", 0),
                operating_hours=payload.inputs.get("daily_operating_hours", 6),
                mode=mode, mode_params=roi_params, fuel=fuel_doc,
            )
            # Overwrite the legacy annual_saving/payback if ROI helper produced them
            if roi.get("annual_saving") and roi["annual_saving"] > 0:
                computed["result"]["annual_saving"] = roi["annual_saving"]
                if computed["result"].get("net_cost"):
                    computed["result"]["payback_years"] = round(computed["result"]["net_cost"] / roi["annual_saving"], 2)
            if roi.get("annual_co2_offset_kg"):
                computed["result"]["annual_co2_offset_kg"] = roi["annual_co2_offset_kg"]
            if roi.get("annual_fuel_units"):
                computed["result"]["fuel_offset_units_yearly"] = roi["annual_fuel_units"]
                computed["result"]["fuel_saved_units_yearly"] = roi["annual_fuel_units"]  # alias
                if fuel_doc:
                    computed["result"]["fuel_name"] = fuel_doc.get("name")
                    computed["result"]["fuel_unit"] = fuel_doc.get("unit")
            computed["roi_details"] = roi
            computed["stages"] = compile_stages_pump(
                computed["result"], computed.get("breakdown", []), payload.inputs, config, roi
            )
    except Exception as _stages_err:
        logger.warning(f"stages compilation failed: {_stages_err}")
        computed["stages"] = None
    return computed


@api_router.post("/calculate/pump/string-voltage")
async def validate_pump_string_voltage(payload: Dict[str, Any], request: Request):
    """Iter 44 Phase 2 — validate a pump string design against controller MPPT + absolute-max limits
    using per-DISCOM/pincode admin-configurable low-temp reference.
    """
    await get_current_user(request)
    from calculators.working import validate_string_voltage

    pump_id = payload.get("pump_product_id")
    panel_id = payload.get("panel_product_id")
    modules_in_series = int(payload.get("modules_in_series") or 0)
    strings_in_parallel = int(payload.get("strings_in_parallel") or 1)
    pincode = payload.get("pincode")

    if not (pump_id and panel_id):
        raise HTTPException(400, "pump_product_id and panel_product_id are required")

    try:
        pump = await db.pump_products.find_one({"_id": ObjectId(pump_id)})
        panel = await db.panel_products.find_one({"_id": ObjectId(panel_id)})
    except Exception:
        raise HTTPException(400, "invalid product id(s)")
    if not (pump and panel):
        raise HTTPException(404, "pump or panel not found in catalogue")

    # Resolve low-temp reference: per-pincode override > global default (from pricing_config)
    site_tmin = None
    if pincode:
        pin_doc = await db.pincodes.find_one({"pincode": str(pincode)})
        if pin_doc and pin_doc.get("min_temp_c") is not None:
            site_tmin = float(pin_doc["min_temp_c"])
    if site_tmin is None:
        pc = await db.pricing_config.find_one({"key": "defaults"}) or {}
        site_tmin = float(pc.get("string_low_temp_default_c", -10))

    return validate_string_voltage(
        pump_product=pump, panel_product=panel,
        modules_in_series=modules_in_series, strings_in_parallel=strings_in_parallel,
        site_min_temp_c=site_tmin,
    )


@api_router.get("/calculate/config")
async def get_calc_config(request: Request):
    await get_current_user(request)
    cfg = await _get_calc_config()
    cfg.pop("_id", None)
    return cfg


@api_router.put("/calculate/config")
async def update_calc_config(payload: Dict[str, Any], request: Request):
    await require_role("admin")(request)
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    payload.pop("_id", None)
    await db.calc_config.update_one({"_id": "singleton"}, {"$set": payload}, upsert=True)
    doc = await db.calc_config.find_one({"_id": "singleton"})
    doc.pop("_id", None)
    return doc


# ── DISCOMs ─────────────────────────────────────────────────────────────
@api_router.get("/calculate/discoms")
async def list_discoms(request: Request, state: Optional[str] = None, active_only: bool = True):
    await get_current_user(request)
    q = {}
    if state: q["state"] = state
    if active_only: q["active"] = True
    docs = await db.discoms.find(q).sort("name", 1).to_list(500)
    return [{k: v for k, v in d.items() if k != "_id"} for d in docs]


@api_router.get("/calculate/discoms/{discom_id}")
async def get_discom(discom_id: str, request: Request):
    await get_current_user(request)
    d = await db.discoms.find_one({"$or": [{"id": discom_id}, {"short_code": discom_id}]})
    if not d:
        raise HTTPException(status_code=404, detail="DISCOM not found")
    d.pop("_id", None)
    return d


@api_router.post("/calculate/discoms")
async def create_discom(payload: Dict[str, Any], request: Request):
    await require_role("admin")(request)
    if not payload.get("id"):
        raise HTTPException(status_code=400, detail="'id' is required")
    existing = await db.discoms.find_one({"id": payload["id"]})
    if existing:
        raise HTTPException(status_code=400, detail="DISCOM id already exists")
    payload["created_at"] = datetime.now(timezone.utc).isoformat()
    payload["active"] = payload.get("active", True)
    await db.discoms.insert_one(payload)
    return {"message": "DISCOM created", "id": payload["id"]}


@api_router.put("/calculate/discoms/{discom_id}")
async def update_discom(discom_id: str, payload: Dict[str, Any], request: Request):
    await require_role("admin")(request)
    payload.pop("_id", None)
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.discoms.update_one({"$or": [{"id": discom_id}, {"short_code": discom_id}]}, {"$set": payload})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="DISCOM not found")
    return {"message": "DISCOM updated"}


@api_router.delete("/calculate/discoms/{discom_id}")
async def delete_discom(discom_id: str, request: Request):
    await require_role("admin")(request)
    if discom_id == "FALLBACK":
        raise HTTPException(status_code=400, detail="Cannot delete FALLBACK DISCOM")
    res = await db.discoms.delete_one({"id": discom_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="DISCOM not found")
    return {"message": "DISCOM deleted"}


# ── Pincodes ────────────────────────────────────────────────────────────
@api_router.get("/calculate/pincodes")
async def list_pincodes(request: Request, state: Optional[str] = None, q: Optional[str] = None, limit: int = 100):
    await get_current_user(request)
    query = {}
    if state: query["state"] = state
    if q:
        query["$or"] = [{"pincode": {"$regex": f"^{q}"}}, {"district": {"$regex": q, "$options": "i"}}]
    docs = await db.pincodes.find(query).limit(limit).to_list(limit)
    return [{k: v for k, v in d.items() if k != "_id"} for d in docs]


@api_router.post("/calculate/pincodes")
async def create_pincode(payload: Dict[str, Any], request: Request):
    await require_role("admin", "manager")(request)
    if not payload.get("pincode") or len(payload["pincode"]) != 6:
        raise HTTPException(status_code=400, detail="pincode must be 6 digits")
    existing = await db.pincodes.find_one({"pincode": payload["pincode"]})
    if existing:
        raise HTTPException(status_code=400, detail="PIN already exists")
    payload["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.pincodes.insert_one(payload)
    return {"message": "PIN created"}


@api_router.post("/calculate/pincodes/import")
async def import_pincodes_csv(request: Request, file: UploadFile = File(...)):
    """Bulk-import PINs from a CSV (up to 20 MB). Accepts the India-Post schema:
        Pincode, StateName, District, Region, Country, ... — or a compact
        schema: pincode, district, state, latitude, longitude, discom
    De-duplicates by `pincode`. Skips rows with invalid PINs. Returns
    { inserted, skipped_existing, skipped_invalid, total_after }.
    """
    import csv, io
    await require_role("admin")(request)
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File exceeds 20 MB limit")
    try:
        text = content.decode("utf-8-sig", errors="replace")
    except Exception:
        raise HTTPException(status_code=400, detail="Could not decode CSV as UTF-8")

    reader = csv.DictReader(io.StringIO(text))
    # Normalise header aliases
    def _get(row, *names):
        for n in names:
            for k in row:
                if k and k.strip().lower() == n.lower():
                    return (row[k] or "").strip()
        return ""

    # Preload existing PINs to skip
    existing = await db.pincodes.find({}, {"pincode": 1}).to_list(200000)
    existing_set = {d["pincode"] for d in existing}

    # State → DISCOM fallback
    from calculators.geo import STATE_FALLBACK
    state_to_discom = {s: v["discom"] for s, v in STATE_FALLBACK.items()}
    state_to_yield = {s: v["yield"] for s, v in STATE_FALLBACK.items()}

    to_insert = []
    inserted = 0
    skipped_existing = 0
    skipped_invalid = 0
    now_iso = datetime.now(timezone.utc).isoformat()

    for row in reader:
        pin = _get(row, "Pincode", "pincode", "PIN Code", "PIN")
        if not pin or len(pin) != 6 or not pin.isdigit():
            skipped_invalid += 1
            continue
        if pin in existing_set:
            skipped_existing += 1
            continue
        existing_set.add(pin)

        district = _get(row, "District", "district", "Districtname", "DistrictName")
        state = _get(row, "StateName", "State", "state")
        # normalise state casing
        state_norm = state.strip().title() if state else ""
        # map common state name spellings
        state_lookup = {"Tamilnadu": "Tamil Nadu", "Andhrapradesh": "Andhra Pradesh"}
        state = state_lookup.get(state_norm.replace(" ", ""), state_norm)

        lat = _get(row, "Latitude", "latitude", "lat")
        lon = _get(row, "Longitude", "longitude", "lon", "lng")
        try: lat_f = float(lat) if lat else None
        except (ValueError, TypeError): lat_f = None
        try: lon_f = float(lon) if lon else None
        except (ValueError, TypeError): lon_f = None

        discom = _get(row, "discom", "DISCOM") or state_to_discom.get(state) or None
        sy = state_to_yield.get(state, 4.5)

        to_insert.append({
            "pincode": pin,
            "district": district or "Unknown",
            "state": state or "Unknown",
            "discom": discom,
            "latitude": lat_f, "longitude": lon_f,
            "specific_yield_kwh_per_kwp_day": sy,
            "peak_sun_hours": round(sy / 0.75, 2),
            "region_cost_factor": 1.0,
            "created_at": now_iso,
        })

        # Insert in chunks of 5000 for memory safety
        if len(to_insert) >= 5000:
            await db.pincodes.insert_many(to_insert, ordered=False)
            inserted += len(to_insert)
            to_insert = []

    if to_insert:
        await db.pincodes.insert_many(to_insert, ordered=False)
        inserted += len(to_insert)

    total = await db.pincodes.count_documents({})
    return {
        "inserted": inserted,
        "skipped_existing": skipped_existing,
        "skipped_invalid": skipped_invalid,
        "total_after": total,
        "file_size_kb": round(len(content) / 1024, 1),
    }


# ── Seed defaults (idempotent) ──────────────────────────────────────────
@api_router.post("/calculate/seed-defaults")
async def seed_calc_defaults(request: Request):
    await require_role("admin")(request)
    created_discoms = 0
    created_pins = 0
    for d in get_default_discoms():
        existing = await db.discoms.find_one({"id": d["id"]})
        if existing: continue
        payload = dict(d)
        payload["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.discoms.insert_one(payload)
        created_discoms += 1
    for p in get_default_pincodes():
        existing = await db.pincodes.find_one({"pincode": p["pincode"]})
        if existing: continue
        payload = dict(p)
        payload["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.pincodes.insert_one(payload)
        created_pins += 1
    return {
        "discoms_created": created_discoms,
        "pincodes_created": created_pins,
        "total_discoms": await db.discoms.count_documents({}),
        "total_pincodes": await db.pincodes.count_documents({}),
    }


# ── PIN backfill migration ─────────────────────────────────────────────
# Regex-based PIN detection from existing project addresses. When a PIN is
# not present, we fall back to a keyword scan for state names — safer than
# calling an external geocoder.
STATE_HINTS = [
    ("Tamil Nadu",       "TANGEDCO"),
    ("TamilNadu",        "TANGEDCO"),
    ("Kerala",           "KSEB"),
    ("Karnataka",        "BESCOM"),
    ("Bengaluru",        "BESCOM"), ("Bangalore", "BESCOM"),
    ("Chennai",          "TANGEDCO"),
    ("Coimbatore",       "TANGEDCO"),
    ("Kochi",            "KSEB"), ("Ernakulam", "KSEB"),
    ("Andhra Pradesh",   "APEPDCL"),
    ("Telangana",        "TSSPDCL"), ("Hyderabad", "TSSPDCL"),
    ("Maharashtra",      "MSEDCL"), ("Mumbai", "MSEDCL"), ("Pune", "MSEDCL"),
    ("Gujarat",          "MGVCL"),
    ("Rajasthan",        "JVVNL"),
    ("Delhi",            "BSES"),
]


class BackfillLocationsRequest(BaseModel):
    dry_run: bool = False
    only_missing: bool = True   # only touch projects without pincode/district


@api_router.post("/projects/backfill-locations")
async def backfill_project_locations(payload: BackfillLocationsRequest, request: Request):
    """One-shot admin migration: extract PIN codes from every project's stored
    addresses (customer.address / location.address / location.site_location_words)
    and autofill location.{pincode,district,state,discom_id} from db.pincodes
    where possible. Projects that only contain a state hint get partial fill.
    """
    import re
    user = await require_role("admin")(request)

    query = {"deleted_at": {"$exists": False}}
    if payload.only_missing:
        query["$or"] = [
            {"location.pincode": {"$in": [None, ""]}},
            {"location.pincode": {"$exists": False}},
        ]

    projects = await db.projects.find(query).to_list(10000)

    # Preload PIN + DISCOM lookups once
    pin_docs = await db.pincodes.find({}).to_list(10000)
    pincodes_map = {p["pincode"]: p for p in pin_docs}
    all_discoms = await db.discoms.find({}).to_list(500)
    discoms_by_code = {d["id"]: d for d in all_discoms}

    pin_regex = re.compile(r"\b[1-9]\d{5}\b")

    resolved = 0
    partial = 0
    unresolved = 0
    updated_ids = []
    unresolved_ids = []

    for p in projects:
        loc = p.get("location") or {}
        text_bits = " ".join([
            (p.get("customer") or {}).get("address") or "",
            loc.get("address") or "",
            loc.get("site_location_words") or "",
        ])
        update = {}

        # PIN extraction
        m = pin_regex.search(text_bits)
        if m:
            pin = m.group(0)
            update["location.pincode"] = pin
            rec = pincodes_map.get(pin)
            if rec:
                update["location.district"] = rec.get("district")
                update["location.state"] = rec.get("state")
                update["location.discom_id"] = rec.get("discom") or rec.get("discom_id")
                resolved += 1
            else:
                partial += 1

        if "location.state" not in update:
            # Fallback: state keyword hint
            lowered = text_bits.lower()
            for name, discom in STATE_HINTS:
                if name.lower() in lowered:
                    update["location.state"] = name if name not in ("TamilNadu", "Bengaluru", "Bangalore",
                                                                    "Chennai", "Coimbatore", "Kochi", "Ernakulam",
                                                                    "Hyderabad", "Mumbai", "Pune") else \
                        {"TamilNadu": "Tamil Nadu", "Bengaluru": "Karnataka", "Bangalore": "Karnataka",
                         "Chennai": "Tamil Nadu", "Coimbatore": "Tamil Nadu",
                         "Kochi": "Kerala", "Ernakulam": "Kerala",
                         "Hyderabad": "Telangana", "Mumbai": "Maharashtra", "Pune": "Maharashtra"}.get(name, name)
                    update["location.discom_id"] = discom
                    if "location.pincode" not in update:
                        partial += 1
                    break
            else:
                if "location.pincode" not in update:
                    unresolved += 1
                    unresolved_ids.append(str(p.get("_id")))

        if update and not payload.dry_run:
            update["updated_at"] = datetime.now(timezone.utc).isoformat()
            await db.projects.update_one({"_id": p["_id"]}, {"$set": update})
        if update:
            updated_ids.append(str(p.get("_id")))

    return {
        "scanned": len(projects),
        "resolved_full": resolved,
        "resolved_partial": partial,
        "unresolved": unresolved,
        "dry_run": payload.dry_run,
        "updated_project_ids": updated_ids[:200],
        "unresolved_sample": unresolved_ids[:20],
        "run_by": user["email"],
        "completed_at": datetime.now(timezone.utc).isoformat(),
    }


class BillSavingsRequest(BaseModel):
    monthly_units_pre: float
    monthly_generation: float
    discom_id: Optional[str] = None
    category_name: Optional[str] = "Domestic"
    net_metering: bool = True


@api_router.post("/calculate/bill-savings")
async def calc_bill_savings(payload: BillSavingsRequest, request: Request):
    """Slab-aware pre-vs-post bill comparison. Standalone endpoint for the UI to
    show a clear 3-number savings breakdown without recomputing the full solution."""
    await get_current_user(request)
    from calculators.tariffs import compute_bill_savings as _cbs, pick_category as _pk
    _, discom_doc = await _get_pincode_and_discom(None, payload.discom_id)
    category = _pk(discom_doc, payload.category_name)
    return _cbs(payload.monthly_units_pre, payload.monthly_generation, category,
                net_metering=payload.net_metering)


# ================== PROJECTS ==================

@api_router.post("/projects")
async def create_project(project: ProjectCreate, request: Request):
    user = await get_current_user(request)
    
    # Build selected items list for cost calculation
    selected_items_data = [si.model_dump() for si in project.selected_items]
    manual_costs_data = [mc.model_dump() for mc in project.manual_costs]
    
    project_doc = {
        "customer": project.customer.model_dump(),
        "location": project.location.model_dump(),
        "electrical": project.electrical.model_dump(),
        "solar_system": project.solar_system.model_dump(),
        "mounting": project.mounting.model_dump(),
        "additional": project.additional.model_dump(),
        "selected_items": selected_items_data,
        "manual_costs": manual_costs_data,
        "site_images": project.site_images,
        "drive_folder_name": project.drive_folder_name or "",
        "drive_folder_link": project.drive_folder_link or "",
        "drive_folder_id": project.drive_folder_id or "",
        "site_measurements": project.site_measurements or {},
        "custom_fields": project.custom_fields or {},
        "solar_report": project.solar_report or None,
        "calculation_snapshot": project.calculation_snapshot or None,
        "terms_id": project.terms_id or None,
        "reference_project_id": project.reference_project_id or None,
        "installation_date": project.installation_date or None,
        "commissioning_date": project.commissioning_date or None,
        "notes": project.notes or "",
        "notes_history": [],
        "status": "draft",
        "location_id": user.get("default_location_id"),
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Calculate cost estimation from selected items
    project_doc["cost_estimation"] = calculate_cost_estimation(selected_items_data, manual_costs_data)
    
    result = await db.projects.insert_one(project_doc)
    
    # Auto-generate reference number
    ref_number = f"SCR-{str(result.inserted_id)[-6:].upper()}"
    await db.projects.update_one(
        {"_id": result.inserted_id},
        {"$set": {"reference_number": ref_number}}
    )
    
    await create_audit_log(
        user["id"], user["name"], "create", "project",
        str(result.inserted_id), None, {"customer": project.customer.name, "status": "draft"}
    )
    
    return {
        "id": str(result.inserted_id),
        "message": "Project created successfully"
    }

def compute_till_date_metrics(project: dict, derived: dict) -> dict:
    """Calculate live 'savings till date' metrics for a project based on its
    installation/commissioning date. Returns None if neither explicit date is set —
    we never use `updated_at` as a stand-in because that would attribute savings
    to projects that were never actually installed."""
    from datetime import datetime as _dt, timezone as _tz
    derived = derived or {}
    install_str = project.get("installation_date") or project.get("commissioning_date")
    if not install_str:
        return None
    try:
        # Accept ISO date or full timestamp
        s = str(install_str).replace("Z", "+00:00")
        if "T" in s:
            install_dt = _dt.fromisoformat(s)
        else:
            install_dt = _dt.fromisoformat(s + "T00:00:00+00:00")
        if install_dt.tzinfo is None:
            install_dt = install_dt.replace(tzinfo=_tz.utc)
    except Exception:
        return None
    now = _dt.now(_tz.utc)
    if install_dt > now:
        return {
            "installation_date": install_str,
            "months_elapsed": 0, "years_elapsed": 0.0,
            "savings_inr": 0, "units_generated": 0, "co2_kg": 0, "fuel_litres": 0,
        }
    days = (now - install_dt).total_seconds() / 86400.0
    months = days / 30.4375
    years = days / 365.25
    annual_savings = float(derived.get("annual_savings") or 0)
    annual_gen = float(derived.get("annual_generation_units") or 0)
    co2_year = float(derived.get("co2_kg_year") or 0)
    fuel_year = float(derived.get("diesel_petrol_saved_liters_yearly") or 0)
    return {
        "installation_date": install_str,
        "months_elapsed": round(months, 1),
        "years_elapsed": round(years, 2),
        "savings_inr": round(annual_savings * years),
        "units_generated": round(annual_gen * years),
        "co2_kg": round(co2_year * years),
        "fuel_litres": round(fuel_year * years),
    }

@api_router.get("/projects/reference-candidates")
async def get_reference_candidates(request: Request, q: Optional[str] = None):
    """Return completed projects suitable to attach as a reference to a new quotation.
    Lightweight summary — customer name, system size, location, key metrics.
    Used by the New Project wizard's 'Reference Site' dropdown."""
    await get_current_user(request)
    query = {"deleted_at": {"$exists": False}, "status": "completed"}
    projects = await db.projects.find(query).sort("updated_at", -1).limit(200).to_list(200)
    out = []
    for p in projects:
        cust = p.get("customer", {}) or {}
        loc = p.get("location", {}) or {}
        cf = p.get("custom_fields", {}) or {}
        ps = cf.get("proposed_solution", {}) or {}
        derived = ps.get("_derived", {}) or {}
        # Pull system size from proposed_solution or legacy solar_report
        sr = p.get("solar_report", {}) or {}
        system_size = ps.get("system_size_kw") or (sr.get("sizing") or {}).get("kwp_recommended") or None
        # Image preference: completion_media first, then site_images, then drive folder
        image_url = None
        for m in (p.get("completion_media") or []):
            if m.get("type") == "image" and m.get("url"):
                image_url = m["url"]; break
        if not image_url:
            for s in (p.get("site_images") or []):
                if s:
                    image_url = s; break
        till = compute_till_date_metrics(p, derived)
        item = {
            "id": str(p["_id"]),
            "reference_number": p.get("reference_number", f"SCR-{str(p['_id'])[-6:].upper()}"),
            "name": p.get("project_name") or cust.get("name", "Unnamed"),
            "customer_name": cust.get("name", ""),
            "phone": cust.get("phone", ""),
            "location": loc.get("address", ""),
            "system_size_kw": system_size,
            "image_url": image_url,
            "completed_at": p.get("updated_at"),
            "installation_date": p.get("installation_date"),
            "commissioning_date": p.get("commissioning_date"),
            "till_date": till,
            "metrics": {
                "monthly_savings": derived.get("monthly_savings"),
                "annual_savings": derived.get("annual_savings"),
                "lifetime_savings": derived.get("lifetime_savings"),
                "roi_pct": derived.get("roi_pct"),
                "payback_years": derived.get("payback_years"),
                "annual_generation_units": derived.get("annual_generation_units"),
                "co2_kg_year": derived.get("co2_kg_year"),
                "diesel_petrol_saved_liters_yearly": derived.get("diesel_petrol_saved_liters_yearly"),
            }
        }
        # Filter by q (search by customer/location/reference number)
        if q:
            qlow = q.lower()
            hay = " ".join([
                str(item.get("name") or ""),
                str(item.get("customer_name") or ""),
                str(item.get("location") or ""),
                str(item.get("reference_number") or ""),
            ]).lower()
            if qlow not in hay:
                continue
        out.append(item)
    return out

@api_router.get("/projects/{project_id}/reference-summary")
async def get_reference_summary(project_id: str, request: Request):
    """Full reference summary for a project — used to render the
    'Reference Project Performance' section in PDFs / preview modal."""
    await get_current_user(request)
    p = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    if not p:
        raise HTTPException(status_code=404, detail="Reference project not found")
    cust = p.get("customer", {}) or {}
    loc = p.get("location", {}) or {}
    cf = p.get("custom_fields", {}) or {}
    ps = cf.get("proposed_solution", {}) or {}
    derived = ps.get("_derived", {}) or {}
    image_url = None
    for m in (p.get("completion_media") or []):
        if m.get("type") == "image" and m.get("url"):
            image_url = m["url"]; break
    if not image_url:
        for s in (p.get("site_images") or []):
            if s:
                image_url = s; break
    return {
        "id": str(p["_id"]),
        "reference_number": p.get("reference_number", f"SCR-{str(p['_id'])[-6:].upper()}"),
        "customer_name": cust.get("name", ""),
        "phone": cust.get("phone", ""),
        "location": loc.get("address", ""),
        "system_size_kw": ps.get("system_size_kw") or (p.get("solar_report", {}) or {}).get("sizing", {}).get("kwp_recommended"),
        "panel_count": ps.get("panel_count"),
        "inverter_kw": ps.get("inverter_kw"),
        "total_cost": ps.get("total_cost"),
        "subsidy": ps.get("subsidy"),
        "completed_at": p.get("updated_at"),
        "installation_date": p.get("installation_date"),
        "commissioning_date": p.get("commissioning_date"),
        "till_date": compute_till_date_metrics(p, derived),
        "image_url": image_url,
        "metrics": derived,
        "notes": p.get("notes", ""),
    }

@api_router.get("/projects")
async def get_projects(request: Request, status: Optional[str] = None, location_id: Optional[str] = None):
    user = await get_current_user(request)
    
    query = {"deleted_at": {"$exists": False}}  # Exclude soft-deleted projects
    
    # Staff can only see their own projects
    if user["role"] == "staff":
        query["created_by"] = user["id"]
    
    if status:
        query["status"] = status

    loc_filter = location_scope_filter(user, location_id)
    if loc_filter:
        query.update(loc_filter)
    
    projects = await db.projects.find(query).sort("created_at", -1).to_list(1000)
    
    return [
        {
            "id": str(p["_id"]),
            "reference_number": p.get("reference_number", f"SCR-{str(p['_id'])[-6:].upper()}"),
            "customer": p["customer"],
            "location": p["location"],
            "status": p["status"],
            "cost_estimation": p.get("cost_estimation", {}),
            "created_by_name": p.get("created_by_name", "Unknown"),
            "created_at": p["created_at"],
            "updated_at": p["updated_at"],
            "location_id": p.get("location_id")
        }
        for p in projects
    ]

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str, request: Request):
    user = await get_current_user(request)
    
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Staff can only see their own projects
    if user["role"] == "staff" and project["created_by"] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get deletion request if exists
    deletion_request = None
    if project["status"] == "deletion_requested":
        dr = await db.deletion_requests.find_one({"project_id": project_id, "status": "pending"})
        if dr:
            deletion_request = {
                "id": str(dr["_id"]),
                "requested_by": dr["requested_by_name"],
                "reason": dr["reason"],
                "requested_at": dr["requested_at"]
            }
    
    return {
        "id": str(project["_id"]),
        "reference_number": project.get("reference_number", f"SCR-{str(project['_id'])[-6:].upper()}"),
        "customer": project["customer"],
        "location": project["location"],
        "electrical": project["electrical"],
        "solar_system": project["solar_system"],
        "mounting": project["mounting"],
        "additional": project["additional"],
        "selected_items": project.get("selected_items", []),
        "manual_costs": project.get("manual_costs", []),
        "site_images": project.get("site_images", []),
        "drive_folder_name": project.get("drive_folder_name", ""),
        "drive_folder_link": project.get("drive_folder_link", ""),
        "drive_folder_id": project.get("drive_folder_id", ""),
        "site_measurements": project.get("site_measurements", {}),
        "custom_fields": project.get("custom_fields", {}),
        "solar_report": project.get("solar_report"),
        "calculation_snapshot": project.get("calculation_snapshot"),
        "terms_id": project.get("terms_id"),
        "reference_project_id": project.get("reference_project_id"),
        "installation_date": project.get("installation_date"),
        "commissioning_date": project.get("commissioning_date"),
        "notes": (project.get("notes") if project.get("notes") is not None else (project.get("additional", {}) or {}).get("shadow_analysis_notes") or ""),
        "notes_history": project.get("notes_history", []),
        "completion_media": project.get("completion_media", []),
        "completion_drive_link": project.get("completion_drive_link", ""),
        "inverter_login": project.get("inverter_login", {}),
        "customer_feedback": project.get("customer_feedback"),
        "status": project["status"],
        "cost_estimation": project.get("cost_estimation", {}),
        "created_by": project["created_by"],
        "created_by_name": project.get("created_by_name", "Unknown"),
        "created_at": project["created_at"],
        "updated_at": project["updated_at"],
        "approved_by": project.get("approved_by"),
        "approved_at": project.get("approved_at"),
        "rejection_reason": project.get("rejection_reason"),
        "margin_added_by": project.get("margin_added_by"),
        "deletion_request": deletion_request
    }

@api_router.put("/projects/{project_id}")
async def update_project(project_id: str, updates: ProjectUpdate, request: Request):
    user = await get_current_user(request)
    
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    editable_statuses = ["draft", "approved"]
    # Detect a notes-only update — those are allowed for any status (incl. completed/rejected)
    notes_only = (
        updates.notes is not None
        and all(getattr(updates, f) is None for f in [
            "customer", "location", "electrical", "solar_system", "mounting", "additional",
            "selected_items", "manual_costs", "site_images", "drive_folder_name", "drive_folder_link",
            "drive_folder_id", "site_measurements", "custom_fields", "solar_report", "terms_id",
            "reference_project_id", "installation_date", "commissioning_date", "status"
        ])
    )
    
    # Staff can only edit their own draft projects (notes are exception — see below)
    if user["role"] == "staff":
        if project["created_by"] != user["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        if project["status"] != "draft" and not notes_only:
            raise HTTPException(status_code=400, detail="Can only edit draft projects")
    elif user["role"] in ["admin", "manager"]:
        if project["status"] not in editable_statuses and not notes_only:
            raise HTTPException(status_code=400, detail=f"Can only edit projects in: {', '.join(editable_statuses)}")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if updates.customer:
        update_data["customer"] = updates.customer.model_dump()
    if updates.location:
        update_data["location"] = updates.location.model_dump()
    if updates.electrical:
        update_data["electrical"] = updates.electrical.model_dump()
    if updates.solar_system:
        update_data["solar_system"] = updates.solar_system.model_dump()
    if updates.mounting:
        update_data["mounting"] = updates.mounting.model_dump()
    if updates.additional:
        update_data["additional"] = updates.additional.model_dump()
    if updates.site_images is not None:
        update_data["site_images"] = updates.site_images
    if updates.drive_folder_name is not None:
        update_data["drive_folder_name"] = updates.drive_folder_name
    if updates.drive_folder_link is not None:
        update_data["drive_folder_link"] = updates.drive_folder_link
    if updates.drive_folder_id is not None:
        update_data["drive_folder_id"] = updates.drive_folder_id
    if updates.site_measurements is not None:
        update_data["site_measurements"] = updates.site_measurements
    if updates.custom_fields is not None:
        update_data["custom_fields"] = updates.custom_fields
    if updates.solar_report is not None:
        update_data["solar_report"] = updates.solar_report
    if updates.terms_id is not None:
        update_data["terms_id"] = updates.terms_id or None
    if updates.reference_project_id is not None:
        update_data["reference_project_id"] = updates.reference_project_id or None
    if updates.installation_date is not None:
        update_data["installation_date"] = updates.installation_date or None
    if updates.commissioning_date is not None:
        update_data["commissioning_date"] = updates.commissioning_date or None
    if updates.notes is not None:
        update_data["notes"] = updates.notes
    if updates.selected_items is not None:
        update_data["selected_items"] = [si.model_dump() for si in updates.selected_items]
    if updates.manual_costs is not None:
        update_data["manual_costs"] = [mc.model_dump() for mc in updates.manual_costs]
    if updates.status and user["role"] in ["admin", "manager"]:
        update_data["status"] = updates.status
    
    # If editing an approved project, revert to submitted for re-approval
    if project["status"] == "approved" and user["role"] in ["admin", "manager"]:
        has_content_changes = any(k in update_data for k in ["customer", "location", "electrical", "solar_system", "mounting", "additional", "site_images", "selected_items", "manual_costs"])
        if has_content_changes and "status" not in update_data:
            update_data["status"] = "submitted"
    
    # Recalculate cost if items changed
    if "selected_items" in update_data or "manual_costs" in update_data:
        sel_items = update_data.get("selected_items", project.get("selected_items", []))
        man_costs = update_data.get("manual_costs", project.get("manual_costs", []))
        update_data["cost_estimation"] = calculate_cost_estimation(sel_items, man_costs)
    
    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$set": update_data}
    )
    
    await create_audit_log(
        user["id"], user["name"], "update", "project", project_id
    )
    
    return {"message": "Project updated successfully"}

@api_router.post("/projects/{project_id}/notes")
async def append_project_note(project_id: str, payload: ProjectNoteAppend, request: Request):
    """Append a timestamped note entry to the project's notes_history.
    Allowed for any project status — including completed — so the team can keep
    adding service / handover / follow-up updates after delivery."""
    user = await get_current_user(request)
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    # Staff may only append to their own projects; admin/manager can touch any
    if user["role"] == "staff" and project["created_by"] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    text = (payload.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Note text is required")
    entry = {
        "id": str(ObjectId()),
        "text": text,
        "author_id": user["id"],
        "author_name": user["name"],
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$push": {"notes_history": entry},
         "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await create_audit_log(user["id"], user["name"], "append_note", "project", project_id, None, {"len": len(text)})
    return {"message": "Note appended", "entry": entry}

@api_router.post("/projects/{project_id}/submit")
async def submit_project(project_id: str, request: Request):
    user = await get_current_user(request)
    
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    if user["role"] == "staff" and project["created_by"] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if project["status"] != "draft":
        raise HTTPException(status_code=400, detail="Only draft projects can be submitted")
    
    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$set": {
            "status": "submitted",
            "submitted_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_audit_log(
        user["id"], user["name"], "submit", "project", project_id
    )
    
    return {"message": "Project submitted for review"}

@api_router.post("/projects/{project_id}/approve")
async def approve_project(project_id: str, request: Request):
    user = await require_role("admin", "manager")(request)
    
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    if project["status"] != "submitted":
        raise HTTPException(status_code=400, detail="Only submitted projects can be approved")
    
    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$set": {
            "status": "approved",
            "approved_by": user["id"],
            "approved_by_name": user["name"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_audit_log(
        user["id"], user["name"], "approve", "project", project_id
    )
    
    return {"message": "Project approved"}

@api_router.post("/projects/{project_id}/reject")
async def reject_project(project_id: str, request: Request):
    user = await require_role("admin", "manager")(request)
    
    body = await request.json()
    reason = body.get("reason", "No reason provided")
    
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    if project["status"] != "submitted":
        raise HTTPException(status_code=400, detail="Only submitted projects can be rejected")
    
    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$set": {
            "status": "rejected",
            "rejected_by": user["id"],
            "rejection_reason": reason,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_audit_log(
        user["id"], user["name"], "reject", "project", project_id, None, {"reason": reason}
    )
    
    return {"message": "Project rejected"}

@api_router.post("/projects/{project_id}/complete")
async def complete_project(project_id: str, request: Request):
    user = await require_role("admin", "manager")(request)

    try:
        body = await request.json()
    except Exception:
        body = {}

    # New schema: completion_drive_link + inverter_login (manual entry).
    # completion_media kept for backward-compat.
    completion_media = body.get("completion_media", [])
    completion_drive_link = (body.get("completion_drive_link") or "").strip()
    inverter_login = body.get("inverter_login") or {}
    customer_feedback = body.get("customer_feedback", "")

    # At least one proof of completion: drive link OR legacy media
    if not completion_drive_link and (not completion_media or len(completion_media) == 0):
        raise HTTPException(status_code=400, detail="Completion Drive link is required")
    # Validate drive link format if provided
    if completion_drive_link and not (completion_drive_link.startswith("http://") or completion_drive_link.startswith("https://")):
        raise HTTPException(status_code=400, detail="Completion Drive link must be a valid URL")
    # inverter_login must be a dict if provided
    if inverter_login and not isinstance(inverter_login, dict):
        raise HTTPException(status_code=400, detail="inverter_login must be an object")

    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if project["status"] != "approved":
        raise HTTPException(status_code=400, detail="Only approved projects can be completed")

    update_doc = {
        "status": "completed",
        "completion_media": completion_media,
        "completion_drive_link": completion_drive_link,
        "inverter_login": {
            "url": (inverter_login.get("url") or "").strip(),
            "username": (inverter_login.get("username") or "").strip(),
            "password": (inverter_login.get("password") or "").strip(),
            "notes": (inverter_login.get("notes") or "").strip(),
        } if inverter_login else {},
        "customer_feedback": customer_feedback,
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    await db.projects.update_one({"_id": ObjectId(project_id)}, {"$set": update_doc})
    await create_audit_log(user["id"], user["name"], "complete", "project", project_id)
    return {"message": "Project marked as completed"}

@api_router.put("/projects/{project_id}/reference")
async def update_reference_number(project_id: str, request: Request):
    """Update project reference number (Admin/Manager only)"""
    current_user = await require_role("admin", "manager")(request)
    body = await request.json()
    ref = body.get("reference_number", "").strip()
    if not ref:
        raise HTTPException(status_code=400, detail="Reference number is required")
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$set": {"reference_number": ref, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await create_audit_log(current_user["id"], current_user["name"], "update", "project_reference", project_id, None, {"reference_number": ref})
    return {"message": "Reference number updated", "reference_number": ref}

@api_router.put("/projects/{project_id}/status")
async def update_project_status(project_id: str, request: Request):
    """Manually change project status (Admin/Manager only)"""
    current_user = await require_role("admin", "manager")(request)
    body = await request.json()
    new_status = body.get("status", "").strip()
    valid_statuses = ["draft", "submitted", "approved", "rejected", "completed"]
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    old_status = project["status"]
    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await create_audit_log(current_user["id"], current_user["name"], "update", "project_status", project_id, {"status": old_status}, {"status": new_status})
    return {"message": f"Status changed to {new_status}", "status": new_status}

@api_router.get("/projects/{project_id}/gallery")
async def project_gallery(project_id: str):
    """Public gallery page for site images - used for QR code in PDF"""
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    site_images = project.get("site_images", [])
    completion_media = project.get("completion_media", [])
    customer_name = project.get("customer", {}).get("name", "Project")
    ref = project.get("reference_number", f"SCR-{str(project['_id'])[-6:].upper()}")
    
    img_tags = ""
    for i, url in enumerate(site_images):
        img_tags += f'<div style="break-inside:avoid;margin-bottom:12px;"><img src="{url}" alt="Site Photo {i+1}" style="width:100%;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.1);" loading="lazy"/><p style="text-align:center;color:#666;font-size:13px;margin:6px 0;">Site Photo {i+1}</p></div>'
    
    media_tags = ""
    for m in completion_media:
        sp = m.get("storage_path", "") if isinstance(m, dict) else ""
        ct = m.get("content_type", "") if isinstance(m, dict) else ""
        fn = m.get("filename", "File") if isinstance(m, dict) else ""
        file_url = f"/api/files/{sp}"
        if ct.startswith("video/"):
            media_tags += f'<div style="break-inside:avoid;margin-bottom:12px;"><video controls style="width:100%;border-radius:8px;" preload="metadata"><source src="{file_url}" type="{ct}"/></video><p style="text-align:center;color:#666;font-size:13px;">{fn}</p></div>'
        elif ct.startswith("image/"):
            media_tags += f'<div style="break-inside:avoid;margin-bottom:12px;"><img src="{file_url}" alt="{fn}" style="width:100%;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.1);" loading="lazy"/><p style="text-align:center;color:#666;font-size:13px;">{fn}</p></div>'
    
    html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{customer_name} - Site Gallery</title>
    <style>*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:system-ui,sans-serif;background:#f8fafc;color:#1e293b;padding:24px;max-width:900px;margin:0 auto}}
    h1{{font-size:1.5rem;margin-bottom:4px}}h2{{font-size:1.1rem;color:#475569;margin:24px 0 12px;border-bottom:2px solid #e2e8f0;padding-bottom:6px}}
    .ref{{color:#64748b;font-size:.9rem;margin-bottom:20px}}.grid{{columns:2;column-gap:16px}}@media(max-width:600px){{.grid{{columns:1}}}}</style></head>
    <body><h1>{customer_name}</h1><p class="ref">{ref} &bull; Sensoper Controls &amp; Renewables</p>"""
    
    if img_tags:
        html += f'<h2>Site Images ({len(site_images)})</h2><div class="grid">{img_tags}</div>'
    if media_tags:
        html += f'<h2>Completion Media</h2><div class="grid">{media_tags}</div>'
    if not img_tags and not media_tags:
        html += '<p style="text-align:center;padding:40px;color:#94a3b8;">No images available for this project.</p>'
    
    html += '</body></html>'
    return HTMLResponse(content=html)

# ================== PROJECT DELETION WORKFLOW ==================

@api_router.post("/projects/{project_id}/request-deletion")
async def request_project_deletion(project_id: str, deletion: DeletionRequestCreate, request: Request):
    """Staff requests deletion, requires manager approval"""
    user = await get_current_user(request)
    
    project = await db.projects.find_one({"_id": ObjectId(project_id), "deleted_at": {"$exists": False}})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Staff can only request deletion of their own projects
    if user["role"] == "staff" and project["created_by"] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Check for existing pending request
    existing = await db.deletion_requests.find_one({
        "project_id": project_id,
        "status": "pending"
    })
    if existing:
        raise HTTPException(status_code=400, detail="Deletion request already pending")
    
    # Create deletion request
    request_doc = {
        "project_id": project_id,
        "project_name": project["customer"]["name"],
        "requested_by": user["id"],
        "requested_by_name": user["name"],
        "reason": deletion.reason,
        "status": "pending",
        "requested_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.deletion_requests.insert_one(request_doc)
    
    # Update project status
    await db.projects.update_one(
        {"_id": ObjectId(project_id)},
        {"$set": {
            "status": "deletion_requested",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_audit_log(
        user["id"], user["name"], "deletion_request", "project", project_id,
        None, {"reason": deletion.reason}
    )
    
    return {"id": str(result.inserted_id), "message": "Deletion request submitted"}

@api_router.get("/deletion-requests")
async def get_deletion_requests(request: Request, status: Optional[str] = None):
    """Get all deletion requests (Manager/Admin only)"""
    user = await require_role("admin", "manager")(request)
    
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.deletion_requests.find(query).sort("requested_at", -1).to_list(100)
    
    return [
        {
            "id": str(r["_id"]),
            "project_id": r["project_id"],
            "project_name": r["project_name"],
            "requested_by": r["requested_by_name"],
            "reason": r["reason"],
            "status": r["status"],
            "requested_at": r["requested_at"],
            "resolved_by": r.get("resolved_by_name"),
            "resolved_at": r.get("resolved_at")
        }
        for r in requests
    ]

@api_router.post("/deletion-requests/{request_id}/approve")
async def approve_deletion(request_id: str, request: Request):
    """Approve deletion request - soft delete the project"""
    user = await require_role("admin", "manager")(request)
    
    del_request = await db.deletion_requests.find_one({"_id": ObjectId(request_id)})
    
    if not del_request:
        raise HTTPException(status_code=404, detail="Deletion request not found")
    
    if del_request["status"] != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")
    
    # Soft delete the project
    await db.projects.update_one(
        {"_id": ObjectId(del_request["project_id"])},
        {"$set": {
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": user["id"],
            "deleted_by_name": user["name"],
            "status": "deleted"
        }}
    )
    
    # Update deletion request
    await db.deletion_requests.update_one(
        {"_id": ObjectId(request_id)},
        {"$set": {
            "status": "approved",
            "resolved_by": user["id"],
            "resolved_by_name": user["name"],
            "resolved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_audit_log(
        user["id"], user["name"], "deletion_approved", "project", del_request["project_id"]
    )
    
    return {"message": "Deletion approved, project soft-deleted"}

@api_router.post("/deletion-requests/{request_id}/reject")
async def reject_deletion(request_id: str, request: Request):
    """Reject deletion request - restore project status"""
    user = await require_role("admin", "manager")(request)
    
    del_request = await db.deletion_requests.find_one({"_id": ObjectId(request_id)})
    
    if not del_request:
        raise HTTPException(status_code=404, detail="Deletion request not found")
    
    if del_request["status"] != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")
    
    # Restore project status to draft
    await db.projects.update_one(
        {"_id": ObjectId(del_request["project_id"])},
        {"$set": {
            "status": "draft",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update deletion request
    await db.deletion_requests.update_one(
        {"_id": ObjectId(request_id)},
        {"$set": {
            "status": "rejected",
            "resolved_by": user["id"],
            "resolved_by_name": user["name"],
            "resolved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await create_audit_log(
        user["id"], user["name"], "deletion_rejected", "project", del_request["project_id"]
    )
    
    return {"message": "Deletion rejected, project restored"}

@api_router.delete("/projects/{project_id}/force")
async def force_delete_project(project_id: str, request: Request):
    """Admin force delete (hard delete)"""
    user = await require_role("admin")(request)
    
    project = await db.projects.find_one({"_id": ObjectId(project_id)})
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Hard delete
    await db.projects.delete_one({"_id": ObjectId(project_id)})
    await db.deletion_requests.delete_many({"project_id": project_id})
    
    await create_audit_log(
        user["id"], user["name"], "force_delete", "project", project_id,
        {"customer": project["customer"]["name"]}
    )
    
    return {"message": "Project permanently deleted"}

# ================== AUDIT LOGS ==================

@api_router.get("/audit-logs")
async def get_audit_logs(
    request: Request,
    entity_type: Optional[str] = None,
    action_type: Optional[str] = None,
    user_id: Optional[str] = None,
    limit: int = 100
):
    """Get audit logs (Admin only, Manager limited)"""
    user = await get_current_user(request)
    
    if user["role"] not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    
    # Manager can only see project-related logs
    if user["role"] == "manager":
        query["entity_type"] = "project"
    elif entity_type:
        query["entity_type"] = entity_type
    
    if action_type:
        query["action_type"] = action_type
    if user_id:
        query["user_id"] = user_id
    
    logs = await db.audit_logs.find(query).sort("timestamp", -1).limit(limit).to_list(limit)
    
    return [
        {
            "id": str(log["_id"]),
            "user_id": log["user_id"],
            "user_name": log["user_name"],
            "action_type": log["action_type"],
            "entity_type": log["entity_type"],
            "entity_id": log["entity_id"],
            "details": log.get("details"),
            "timestamp": log["timestamp"]
        }
        for log in logs
    ]

# ================== DASHBOARD STATS ==================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request):
    user = await get_current_user(request)
    
    query = {"deleted_at": {"$exists": False}}
    if user["role"] == "staff":
        query["created_by"] = user["id"]
    
    # Get project counts by status
    pipeline = [
        {"$match": query},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]
    
    status_counts = await db.projects.aggregate(pipeline).to_list(100)
    
    stats = {
        "draft": 0,
        "submitted": 0,
        "approved": 0,
        "rejected": 0,
        "completed": 0,
        "deletion_requested": 0,
        "total": 0
    }
    
    for item in status_counts:
        if item["_id"] in stats:
            stats[item["_id"]] = item["count"]
            stats["total"] += item["count"]
    
    # Calculate total revenue from completed projects
    revenue_pipeline = [
        {"$match": {**query, "status": "completed"}},
        {"$group": {"_id": None, "total": {"$sum": "$cost_estimation.total_cost"}}}
    ]
    
    revenue_result = await db.projects.aggregate(revenue_pipeline).to_list(1)
    stats["total_revenue"] = revenue_result[0]["total"] if revenue_result else 0
    
    # Conversion rate — based on real leads (drafts excluded; they are work-in-progress, not leads)
    leads = stats["total"] - stats["draft"]
    if leads > 0:
        stats["conversion_rate"] = round((stats["completed"] / leads) * 100, 1)
    else:
        stats["conversion_rate"] = 0
    
    # Pending deletion requests count (for managers/admins)
    if user["role"] in ["admin", "manager"]:
        pending_deletions = await db.deletion_requests.count_documents({"status": "pending"})
        stats["pending_deletions"] = pending_deletions
        
        # Pending approvals count
        pending_approvals = await db.approvals.count_documents({"status": "pending"})
        stats["pending_approvals"] = pending_approvals
        
        # Low stock alerts count — scoped to the user's assigned location(s)
        loc_filter = location_scope_filter(user)
        low_stock_match = {"$expr": {"$lte": ["$quantity", "$reorder_level"]}}
        if loc_filter:
            low_stock_match.update(loc_filter)
        low_stock_pipeline = [
            {"$match": low_stock_match}
        ]
        low_stock_items = await db.inventory_items.aggregate(low_stock_pipeline).to_list(100)
        stats["low_stock_alerts"] = len(low_stock_items)
    
    return stats

# ================== CEO DASHBOARD ==================

@api_router.get("/dashboard/ceo")
async def get_ceo_dashboard(request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="CEO dashboard is admin/manager only")
    
    query = {"deleted_at": {"$exists": False}}
    all_projects = await db.projects.find(query).to_list(5000)
    
    # Status counts
    status_counts = {}
    for p in all_projects:
        s = p.get("status", "draft")
        status_counts[s] = status_counts.get(s, 0) + 1
    
    total = len(all_projects)
    completed = status_counts.get("completed", 0)
    approved = status_counts.get("approved", 0)
    
    # Revenue & Profit
    total_revenue = sum(p.get("cost_estimation", {}).get("total_cost", 0) for p in all_projects if p.get("status") in ["completed", "approved"])
    # Iter 39 Change 1c: include direct sales revenue
    sales_docs = await db.sales.find({"status": {"$ne": "cancelled"}}).to_list(10000)
    direct_sales_revenue = sum(s.get("grand_total", 0) for s in sales_docs)
    direct_sales_margin = sum(sum(l.get("margin_amount", 0) for l in (s.get("lines") or [])) for s in sales_docs)
    project_revenue = total_revenue
    total_revenue = project_revenue + direct_sales_revenue
    total_margin = sum(p.get("cost_estimation", {}).get("margin_total", 0) for p in all_projects if p.get("status") in ["completed", "approved"])
    
    # Conversion rate — proportion of real leads (drafts excluded) that became wins (approved + completed)
    drafts = status_counts.get("draft", 0)
    leads = total - drafts
    wins = approved + completed
    conversion_rate = round((wins / leads) * 100, 1) if leads > 0 else 0
    
    # Monthly revenue trend (last 12 months)
    from collections import defaultdict
    monthly_revenue = defaultdict(float)
    monthly_projects = defaultdict(int)
    for p in all_projects:
        created = p.get("created_at", "")
        if created:
            month_key = created[:7]  # YYYY-MM
            monthly_projects[month_key] += 1
            if p.get("status") in ["completed", "approved"]:
                monthly_revenue[month_key] += p.get("cost_estimation", {}).get("total_cost", 0)
    
    sorted_months = sorted(monthly_revenue.keys())[-12:]
    revenue_trend = [{"month": m, "revenue": round(monthly_revenue.get(m, 0)), "projects": monthly_projects.get(m, 0)} for m in sorted_months]
    
    # Sales funnel
    funnel = {
        "total_leads": total - drafts,
        "quotes_generated": sum(1 for p in all_projects if p.get("status") != "draft"),
        "approved": approved + completed,
        "completed": completed
    }
    
    # Top staff
    staff_perf = defaultdict(lambda: {"name": "", "count": 0, "revenue": 0})
    for p in all_projects:
        uid = p.get("created_by", "")
        staff_perf[uid]["name"] = p.get("created_by_name", "Unknown")
        staff_perf[uid]["count"] += 1
        if p.get("status") in ["completed", "approved"]:
            staff_perf[uid]["revenue"] += p.get("cost_estimation", {}).get("total_cost", 0)
    top_staff = sorted(staff_perf.values(), key=lambda x: x["revenue"], reverse=True)[:5]
    
    # Inventory value
    inv_items = await db.inventory_items.find().to_list(1000)
    inventory_value = sum(i.get("unit_price", 0) * i.get("quantity", 0) for i in inv_items)
    low_stock = sum(1 for i in inv_items if i.get("quantity", 0) <= i.get("reorder_level", 5))
    
    # Pending approvals
    pending_approvals = await db.approvals.count_documents({"status": "pending"})
    
    # Customer credit data
    credits = await db.customer_credits.find({"status": {"$ne": "closed"}}).to_list(500)
    total_outstanding = sum(c.get("balance", 0) for c in credits)
    overdue_credits = [c for c in credits if c.get("status") == "overdue"]
    overdue_amount = sum(c.get("balance", 0) for c in overdue_credits)
    # Top 5 debtors
    top_debtors = sorted(credits, key=lambda x: x.get("balance", 0), reverse=True)[:5]
    top_debtors_list = [{"name": c.get("customer_name", ""), "balance": round(c.get("balance", 0)), "status": c.get("status", "")} for c in top_debtors]
    # Aging
    credit_aging = {"0_30": 0, "30_60": 0, "60_plus": 0}
    for c in credits:
        created = c.get("created_at", "")
        if created:
            try:
                created_dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                days = (datetime.now(timezone.utc) - created_dt).days
                if days <= 30: credit_aging["0_30"] += c.get("balance", 0)
                elif days <= 60: credit_aging["30_60"] += c.get("balance", 0)
                else: credit_aging["60_plus"] += c.get("balance", 0)
            except (ValueError, TypeError):
                pass
    
    return {
        "kpis": {
            "total_revenue": round(total_revenue),
            "total_profit": round(total_margin),
            "conversion_rate": conversion_rate,
            "wins": wins,
            "approved_projects": approved,
            "active_projects": status_counts.get("submitted", 0) + status_counts.get("approved", 0),
            "completed_projects": completed,
            "pending_approvals": pending_approvals,
            "inventory_value": round(inventory_value),
            "low_stock_alerts": low_stock,
            "total_projects": total,
            "total_outstanding": round(total_outstanding),
            "overdue_amount": round(overdue_amount)
        },
        "status_distribution": [{"name": k, "value": v} for k, v in status_counts.items()],
        "revenue_trend": revenue_trend,
        "sales_funnel": funnel,
        "top_staff": top_staff,
        "credit_data": {
            "total_outstanding": round(total_outstanding),
            "overdue_amount": round(overdue_amount),
            "top_debtors": top_debtors_list,
            "aging": credit_aging
        },
        "accounts_summary": await _ceo_accounts_summary(),
        "readings_summary": await _ceo_readings_summary(),
        "health_score": await _ceo_health_score(all_projects, credits, inv_items, pending_approvals),
        "direct_sales": {
            "revenue": direct_sales_revenue,
            "margin": direct_sales_margin,
            "count": len(sales_docs),
        },
        "project_revenue": project_revenue,
    }


# ================== CEO HEALTH SCORE ==================
from health import compute_pillars as _compute_health_pillars, DEFAULT_HEALTH_CONFIG


async def _get_health_config():
    doc = await db.health_config.find_one({"_id": "singleton"})
    if not doc:
        await db.health_config.insert_one(dict(DEFAULT_HEALTH_CONFIG))
        return dict(DEFAULT_HEALTH_CONFIG)
    return doc


async def _ceo_health_score(all_projects, credits, inv_items, pending_approvals_count):
    cfg = await _get_health_config()
    # Fetch approvals list, daily_updates & weekly_audits + brand_returns for the pillars
    approvals_docs = await db.approvals.find({}).to_list(500)
    daily_updates_docs = await db.daily_updates.find({}).to_list(500)
    audit_docs = await db.weekly_audits.find({}).to_list(200)
    return_docs = await db.brand_returns.find({}).to_list(500)
    return _compute_health_pillars(
        projects=all_projects, credits=credits, inv_items=inv_items,
        approvals=approvals_docs, health_cfg=cfg,
        daily_updates=daily_updates_docs, weekly_audits=audit_docs, brand_returns=return_docs
    )


@api_router.get("/dashboard/health/config")
async def get_health_config_api(request: Request):
    await get_current_user(request)
    cfg = await _get_health_config()
    cfg.pop("_id", None)
    return cfg


@api_router.put("/dashboard/health/config")
async def update_health_config_api(payload: Dict[str, Any], request: Request):
    await require_role("admin")(request)
    payload.pop("_id", None)
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.health_config.update_one({"_id": "singleton"}, {"$set": payload}, upsert=True)
    doc = await db.health_config.find_one({"_id": "singleton"})
    doc.pop("_id", None)
    return doc


@api_router.post("/dashboard/health/snapshot")
async def snapshot_health(request: Request):
    """Persist the current health score as a monthly datapoint. Idempotent per month."""
    user = await require_role("admin", "manager")(request)
    query = {"deleted_at": {"$exists": False}}
    projects = await db.projects.find(query).to_list(5000)
    credits = await db.customer_credits.find({}).to_list(2000)
    inv_items = await db.inventory_items.find({}).to_list(2000)
    approvals = await db.approvals.count_documents({"status": "pending"})
    hs = await _ceo_health_score(projects, credits, inv_items, approvals)
    month_key = datetime.now(timezone.utc).strftime("%Y-%m")
    snap = {
        "month": month_key,
        "score": hs["score"], "band": hs["band"], "verdict": hs["verdict"],
        "pillars": hs["pillars"], "weakest_pillar": hs["weakest_pillar"],
        "created_by": user["id"], "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.health_snapshots.update_one({"month": month_key}, {"$set": snap}, upsert=True)
    return {"month": month_key, "score": hs["score"], "message": "snapshot saved"}


@api_router.get("/dashboard/health/history")
async def get_health_history(request: Request, months: int = 12):
    await get_current_user(request)
    docs = await db.health_snapshots.find({}).sort("month", -1).limit(months).to_list(months)
    docs.reverse()
    return [{k: v for k, v in d.items() if k != "_id"} for d in docs]


# ================== EXPANSION MODULE ==================
from expansion import compute_district_scores as _compute_districts, \
    simulate_breakeven as _simulate_be, DEFAULT_EXPANSION_CONFIG


async def _get_expansion_config():
    doc = await db.expansion_config.find_one({"_id": "singleton"})
    if not doc:
        await db.expansion_config.insert_one(dict(DEFAULT_EXPANSION_CONFIG))
        return dict(DEFAULT_EXPANSION_CONFIG)
    return doc


@api_router.get("/expansion/overview")
async def expansion_overview(request: Request, state: Optional[str] = None):
    user = await get_current_user(request)
    if user["role"] not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Expansion module is admin/manager only")
    cfg = await _get_expansion_config()
    projects = await db.projects.find({"deleted_at": {"$exists": False}}).to_list(5000)
    credits = await db.customer_credits.find({}).to_list(2000)
    returns = await db.brand_returns.find({}).to_list(500)
    branches = await db.branches.find({}).to_list(200)
    result = _compute_districts(projects, credits, returns, branches, cfg)
    if state:
        result["districts"] = [d for d in result["districts"] if d["state"] == state]
    return result


@api_router.get("/expansion/district/{district}")
async def expansion_district(district: str, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Forbidden")
    cfg = await _get_expansion_config()
    projects = await db.projects.find({"deleted_at": {"$exists": False}}).to_list(5000)
    credits = await db.customer_credits.find({}).to_list(2000)
    returns = await db.brand_returns.find({}).to_list(500)
    branches = await db.branches.find({}).to_list(200)
    all_ = _compute_districts(projects, credits, returns, branches, cfg)
    match = next((d for d in all_["districts"] if d["district"].lower() == district.lower()), None)
    if not match:
        raise HTTPException(status_code=404, detail="District not found")
    return match


class BreakEvenSimRequest(BaseModel):
    district: Optional[str] = None
    monthly_run_rate: float = 0
    monthly_branch_cost: float = 250000
    setup_capex: float = 1500000
    target_margin_pct: float = 20
    current_monthly_projects: float = 0
    current_avg_ticket: float = 300000


@api_router.post("/expansion/simulate")
async def expansion_simulate(payload: BreakEvenSimRequest, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Forbidden")
    return _simulate_be(
        monthly_run_rate=payload.monthly_run_rate,
        target_margin_pct=payload.target_margin_pct,
        monthly_branch_cost=payload.monthly_branch_cost,
        setup_capex=payload.setup_capex,
        current_monthly_projects=payload.current_monthly_projects,
        current_avg_ticket=payload.current_avg_ticket,
    )


@api_router.get("/expansion/config")
async def expansion_get_config(request: Request):
    await get_current_user(request)
    cfg = await _get_expansion_config()
    cfg.pop("_id", None)
    return cfg


@api_router.put("/expansion/config")
async def expansion_update_config(payload: Dict[str, Any], request: Request):
    await require_role("admin")(request)
    payload.pop("_id", None)
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.expansion_config.update_one({"_id": "singleton"}, {"$set": payload}, upsert=True)
    doc = await db.expansion_config.find_one({"_id": "singleton"})
    doc.pop("_id", None)
    return doc


# ── Branches CRUD ──
@api_router.get("/expansion/branches")
async def list_branches(request: Request):
    await get_current_user(request)
    docs = await db.branches.find({}).sort("name", 1).to_list(200)
    return [{**{k: v for k, v in d.items() if k != "_id"}, "id": str(d["_id"])} for d in docs]


@api_router.post("/expansion/branches")
async def create_branch(payload: Dict[str, Any], request: Request):
    await require_role("admin", "manager")(request)
    payload["created_at"] = datetime.now(timezone.utc).isoformat()
    if not payload.get("name"):
        raise HTTPException(status_code=400, detail="name is required")
    r = await db.branches.insert_one(payload)
    return {"id": str(r.inserted_id), "message": "Branch added"}


@api_router.put("/expansion/branches/{branch_id}")
async def update_branch(branch_id: str, payload: Dict[str, Any], request: Request):
    user = await require_role("admin", "manager")(request)
    payload.pop("_id", None); payload.pop("id", None)
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    old = await db.branches.find_one({"_id": ObjectId(branch_id)})
    r = await db.branches.update_one({"_id": ObjectId(branch_id)}, {"$set": payload})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    if old and payload.get("name") and old.get("name") != payload.get("name"):
        await create_audit_log(user["id"], user["name"], "branch_renamed", "branch", branch_id,
                                {"name": old.get("name")}, {"name": payload.get("name")})
    return {"message": "Branch updated"}


@api_router.delete("/expansion/branches/{branch_id}")
async def delete_branch(branch_id: str, request: Request):
    await require_role("admin", "manager")(request)
    r = await db.branches.delete_one({"_id": ObjectId(branch_id)})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    return {"message": "Branch deleted"}


async def _ceo_accounts_summary():
    """Latest snapshot per account_entries type for CEO Dashboard cards."""
    out = {}
    for t in ("cash_on_hand", "account_balance"):
        latest = await db.account_entries.find_one({"entry_type": t}, sort=[("entry_date", -1), ("created_at", -1)])
        if latest:
            out[t] = {
                "amount": latest.get("amount", 0),
                "entry_date": latest.get("entry_date"),
                "description": latest.get("description", ""),
                "entered_by": latest.get("entered_by", ""),
                "updated_at": latest.get("updated_at") or latest.get("created_at")
            }
        else:
            out[t] = {"amount": 0, "entry_date": None, "description": "", "entered_by": "", "updated_at": None}
    # Month-to-date expense + GST totals
    today_utc = datetime.now(timezone.utc)
    month_start = today_utc.strftime("%Y-%m-01")
    op_exp_total = 0.0
    gst_input_total = 0.0
    gst_paid_total = 0.0
    async for d in db.account_entries.find({"entry_type": "operational_expense", "entry_date": {"$gte": month_start}}):
        op_exp_total += float(d.get("amount", 0))
    async for d in db.account_entries.find({"entry_type": "gst_input", "entry_date": {"$gte": month_start}}):
        gst_input_total += float(d.get("amount", 0))
    async for d in db.account_entries.find({"entry_type": "gst_paid", "entry_date": {"$gte": month_start}}):
        gst_paid_total += float(d.get("amount", 0))
    out["operational_expense_mtd"] = round(op_exp_total, 2)
    out["gst_input_mtd"] = round(gst_input_total, 2)
    out["gst_paid_mtd"] = round(gst_paid_total, 2)
    out["gst_net_mtd"] = round(gst_paid_total - gst_input_total, 2)  # net liability (positive = owed to govt)
    # 30-day cash history
    docs = await db.account_entries.find({"entry_type": "cash_on_hand"}).sort("entry_date", -1).limit(30).to_list(30)
    out["cash_history"] = [{"date": d.get("entry_date"), "amount": d.get("amount", 0)} for d in reversed(docs)]
    return out

async def _ceo_readings_summary():
    docs = await db.readings.find({}).to_list(5000)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    active = completed = overdue = 0
    for d in docs:
        end_date = _compute_end_date(d.get("start_date", ""), d.get("days", 0))
        s = d.get("status", "active")
        if s == "active" and end_date and end_date < today:
            s = "overdue"
        if s == "active": active += 1
        elif s == "completed": completed += 1
        elif s == "overdue": overdue += 1
    return {"total": len(docs), "active": active, "completed": completed, "overdue": overdue}

# ================== REPORTS ENGINE (8 Consolidated Reports) ==================

async def _get_filtered_projects(query_base, date_from, date_to, system_type, status, customer, staff, project_id):
    query = {**query_base, "deleted_at": {"$exists": False}}
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})
        if isinstance(query["created_at"], dict):
            query["created_at"]["$lte"] = date_to + "T23:59:59"
        else:
            query["created_at"] = {"$gte": query["created_at"], "$lte": date_to + "T23:59:59"}
    if system_type and system_type != "all":
        query["solar_system.system_type"] = system_type
    if status and status != "all":
        query["status"] = status
    if customer:
        query["customer.name"] = {"$regex": customer, "$options": "i"}
    if staff:
        query["created_by_name"] = {"$regex": staff, "$options": "i"}
    if project_id:
        query["_id"] = ObjectId(project_id)
    return await db.projects.find(query).to_list(5000)

@api_router.get("/reports/{report_type}")
async def get_report(report_type: str, request: Request, date_from: str = None, date_to: str = None, system_type: str = None, status: str = None, project_id: str = None, tab: str = None, movement_type: str = None):
    user = await get_current_user(request)
    if user["role"] not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Reports are admin/manager only")
    projects = await _get_filtered_projects({}, date_from, date_to, system_type, status, None, None, project_id)
    inv_items = await db.inventory_items.find().to_list(1000)
    from collections import defaultdict, Counter

    # === 1. SALES & REVENUE ===
    if report_type == "sales_revenue":
        total_quotes = len(projects)
        approved = sum(1 for p in projects if p.get("status") in ["approved", "completed"])
        rejected = sum(1 for p in projects if p.get("status") == "rejected")
        revenue = sum(p.get("cost_estimation", {}).get("total_cost", 0) for p in projects if p.get("status") in ["approved", "completed"])
        # Lead sources
        sources = defaultdict(lambda: {"leads": 0, "converted": 0, "revenue": 0})
        for p in projects:
            src = p.get("custom_fields", {}).get("customer", {}).get("referral_source", "Direct")
            sources[src]["leads"] += 1
            if p.get("status") in ["approved", "completed"]:
                sources[src]["converted"] += 1
                sources[src]["revenue"] += p.get("cost_estimation", {}).get("total_cost", 0)
        lead_rows = [{"source": k, "leads": v["leads"], "converted": v["converted"], "conversion_rate": round((v["converted"]/v["leads"])*100,1) if v["leads"] else 0, "revenue": round(v["revenue"])} for k,v in sources.items()]
        tabs_data = {
            "overview": {"rows": [{"customer": p.get("customer",{}).get("name",""), "ref": p.get("reference_number",""), "status": p.get("status",""), "total": round(p.get("cost_estimation",{}).get("total_cost",0)), "date": p.get("created_at","")[:10]} for p in projects]},
            "lead_sources": {"rows": sorted(lead_rows, key=lambda x: x["revenue"], reverse=True)}
        }
        chart_data = [{"name": "Approved/Won", "value": approved}, {"name": "Rejected", "value": rejected}, {"name": "Pending", "value": total_quotes - approved - rejected}]
        return {"title": "Sales & Revenue Report", "summary": {"total_quotes": total_quotes, "approved": approved, "conversion_rate": round((approved/total_quotes)*100,1) if total_quotes else 0, "revenue": round(revenue)}, "rows": tabs_data.get(tab or "overview", tabs_data["overview"])["rows"], "tabs": list(tabs_data.keys()), "chart_data": [c for c in chart_data if c["value"]>0]}

    # === 2. PROFIT & LEAKAGE ===
    elif report_type == "profit_leakage":
        settings = await db.settings.find_one({"type": "thresholds"}) or {}
        min_margin = settings.get("min_margin_pct", 8)
        rows = []
        total_cost = total_selling = total_margin = total_leakage = 0
        for p in projects:
            ce = p.get("cost_estimation", {})
            selling = ce.get("total_cost", 0)
            margin = ce.get("margin_total", 0)
            gst = ce.get("gst_total", 0)
            base = selling - margin - gst
            margin_pct = round((margin/selling)*100,1) if selling else 0
            alert = ""
            leak = 0
            if margin_pct < min_margin and selling > 0:
                alert = f"Low margin ({margin_pct}% < {min_margin}%)"
                leak = round((min_margin - margin_pct) / 100 * selling)
            rows.append({"customer": p.get("customer",{}).get("name",""), "ref": p.get("reference_number",""), "base_cost": round(base), "selling": round(selling), "margin": round(margin), "margin_pct": margin_pct, "gst": round(gst), "alert": alert, "leakage": leak})
            total_cost += base; total_selling += selling; total_margin += margin; total_leakage += leak
        # Material variance tab
        usage_logs = await db.material_usage_logs.find().to_list(5000)
        item_data = defaultdict(lambda: {"est": 0, "act": 0, "waste": 0})
        for log in usage_logs:
            item_data[log.get("item_name","")]["est"] += log.get("estimated_qty",0)
            item_data[log.get("item_name","")]["act"] += log.get("actual_qty",0)
            item_data[log.get("item_name","")]["waste"] += log.get("wastage",0)
        variance_rows = [{"item": k, "estimated": round(v["est"],1), "actual": round(v["act"],1), "variance": round(v["act"]-v["est"],1), "wastage": round(v["waste"],1)} for k,v in item_data.items()]
        tabs_data = {"profit": {"rows": rows}, "material_variance": {"rows": variance_rows if variance_rows else [{"item": "No data", "estimated": 0, "actual": 0, "variance": 0, "wastage": 0}]}}
        chart_data = [{"name": "Base Cost", "value": round(total_cost)}, {"name": "Margin", "value": round(total_margin)}, {"name": "GST", "value": round(total_selling-total_cost-total_margin)}, {"name": "Leakage", "value": round(total_leakage)}]
        return {"title": "Profit & Leakage Report", "summary": {"total_selling": round(total_selling), "total_margin": round(total_margin), "avg_margin_pct": round((total_margin/total_selling)*100,1) if total_selling else 0, "total_leakage": round(total_leakage)}, "rows": tabs_data.get(tab or "profit", tabs_data["profit"])["rows"], "tabs": list(tabs_data.keys()), "chart_data": [c for c in chart_data if c["value"]>0]}

    # === 3. PROJECT EXECUTION ===
    elif report_type == "project_execution":
        rows = []
        for p in projects:
            el = p.get("electrical", {})
            ss = p.get("solar_system", {})
            created = p.get("created_at", "")[:10]
            updated = p.get("updated_at", "")[:10]
            is_complete = p.get("status") == "completed"
            rows.append({"customer": p.get("customer",{}).get("name",""), "ref": p.get("reference_number",""), "status": p.get("status",""), "system_type": ss.get("system_type",""), "kw": el.get("sanction_load_kw",0), "created": created, "updated": updated, "staff": p.get("created_by_name",""), "om_status": "Active" if is_complete else "N/A"})
        status_dist = Counter(r["status"] for r in rows)
        chart_data = [{"name": k, "value": v} for k,v in status_dist.items()]
        return {"title": "Project Execution Report", "summary": {"total": len(rows), "completed": sum(1 for r in rows if r["status"]=="completed"), "in_progress": sum(1 for r in rows if r["status"] in ["submitted","approved"]), "active_om": sum(1 for r in rows if r["om_status"]=="Active")}, "rows": rows, "chart_data": chart_data}

    # === 4. INVENTORY & MATERIAL ===
    elif report_type == "inventory_material":
        stock_rows = [{"name": i.get("name",""), "sku": i.get("sku_code",""), "category": i.get("category",""), "quantity": i.get("quantity",0), "unit_price": round(i.get("unit_price",0)), "total_value": round(i.get("unit_price",0)*i.get("quantity",0)), "reorder_level": i.get("reorder_level",5), "low_stock": i.get("quantity",0) <= i.get("reorder_level",5)} for i in inv_items]
        # Usage
        usage_logs = await db.material_usage_logs.find().to_list(5000)
        usage_agg = defaultdict(lambda: {"est": 0, "act": 0, "waste": 0})
        for log in usage_logs:
            usage_agg[log.get("item_name","")]["est"] += log.get("estimated_qty",0)
            usage_agg[log.get("item_name","")]["act"] += log.get("actual_qty",0)
            usage_agg[log.get("item_name","")]["waste"] += log.get("wastage",0)
        usage_rows = [{"item": k, "estimated": round(v["est"],1), "actual": round(v["act"],1), "variance": round(v["act"]-v["est"],1), "wastage": round(v["waste"],1), "status": "Excess" if v["act"]>v["est"] else "Shortage" if v["act"]<v["est"] else "OK"} for k,v in usage_agg.items()]
        # Alerts
        alert_rows = [{"name": i.get("name",""), "quantity": i.get("quantity",0), "reorder": i.get("reorder_level",5), "deficit": max(0, i.get("reorder_level",5)-i.get("quantity",0))} for i in inv_items if i.get("quantity",0) <= i.get("reorder_level",5)]

        # === Movement Analysis (fast / slow) ===
        # Determine time window from filters — fallback to last 30 days
        now_utc = datetime.now(timezone.utc)
        if date_from:
            try:
                window_start = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
            except (ValueError, TypeError):
                window_start = now_utc - timedelta(days=30)
        else:
            window_start = now_utc - timedelta(days=30)
        if date_to:
            try:
                window_end = datetime.fromisoformat(date_to + "T23:59:59").replace(tzinfo=timezone.utc)
            except (ValueError, TypeError):
                window_end = now_utc
        else:
            window_end = now_utc

        # Aggregate usage per item_name within window
        movement_agg = defaultdict(lambda: {"count": 0, "last_used": None, "qty": 0})
        for log in usage_logs:
            ts_str = log.get("created_at") or log.get("timestamp") or ""
            try:
                ts = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
                if ts.tzinfo is None:
                    ts = ts.replace(tzinfo=timezone.utc)
            except (ValueError, TypeError, AttributeError):
                continue
            if ts < window_start or ts > window_end:
                continue
            name = log.get("item_name", "")
            if not name:
                continue
            movement_agg[name]["count"] += 1
            movement_agg[name]["qty"] += log.get("actual_qty", 0)
            if movement_agg[name]["last_used"] is None or ts > movement_agg[name]["last_used"]:
                movement_agg[name]["last_used"] = ts

        FAST_THRESHOLD = 5  # ≥ 5 usages in window → Fast
        movement_rows = []
        for it in inv_items:
            nm = it.get("name", "")
            agg = movement_agg.get(nm, {"count": 0, "last_used": None, "qty": 0})
            mtype = "Fast" if agg["count"] >= FAST_THRESHOLD else "Slow"
            last_used_str = agg["last_used"].strftime("%Y-%m-%d") if agg["last_used"] else "-"
            procurement_date = it.get("procurement_date") or "-"
            active_status = "Active" if it.get("active", True) else "Inactive"
            movement_rows.append({
                "product": nm,
                "sku": it.get("sku_code", ""),
                "status": active_status,
                "procurement_date": procurement_date,
                "last_used_date": last_used_str,
                "usage_count": agg["count"],
                "qty_used": round(agg["qty"], 1),
                "movement_type": mtype
            })
        # Sort fast first then by usage_count desc (on full list BEFORE filter)
        movement_rows.sort(key=lambda r: (0 if r["movement_type"] == "Fast" else 1, -r["usage_count"]))

        # Compute summary counters on UNFILTERED list
        fast_count = sum(1 for r in movement_rows if r["movement_type"] == "Fast")
        slow_count = sum(1 for r in movement_rows if r["movement_type"] == "Slow")

        # Apply movement_type filter (only affects displayed rows, not summary)
        if movement_type and movement_type not in (None, "all", "All"):
            mt_norm = movement_type.strip().lower()
            movement_rows = [r for r in movement_rows if r["movement_type"].lower() == mt_norm]

        tabs_data = {
            "stock_levels": {"rows": stock_rows},
            "material_usage": {"rows": usage_rows if usage_rows else [{"item": "No logs", "estimated": 0, "actual": 0, "variance": 0, "wastage": 0, "status": "N/A"}]},
            "alerts": {"rows": alert_rows if alert_rows else [{"name": "All stock OK", "quantity": 0, "reorder": 0, "deficit": 0}]},
            "movement": {"rows": movement_rows if movement_rows else [{"product": "No items", "sku": "", "status": "-", "procurement_date": "-", "last_used_date": "-", "usage_count": 0, "qty_used": 0, "movement_type": "-"}]}
        }
        cat_val = defaultdict(float)
        for r in stock_rows: cat_val[r["category"]] += r["total_value"]
        chart_data = [{"name": k, "value": round(v)} for k,v in cat_val.items() if v>0]
        summary = {"total_items": len(stock_rows), "total_value": sum(r["total_value"] for r in stock_rows), "low_stock": sum(1 for r in stock_rows if r["low_stock"]), "materials_tracked": len(usage_agg)}
        if (tab or "") == "movement":
            summary = {"total_items": len(inv_items), "fast_moving": fast_count, "slow_moving": slow_count, "window_days": max(1, (window_end - window_start).days)}
        return {"title": "Inventory & Material Report", "summary": summary, "rows": tabs_data.get(tab or "stock_levels", tabs_data["stock_levels"])["rows"], "tabs": list(tabs_data.keys()), "chart_data": chart_data}

    # === 5. CUSTOMER CREDIT ===
    elif report_type == "customer_credit":
        all_payments = await db.payments.find().to_list(5000)
        project_payments = defaultdict(float)
        for pay in all_payments: project_payments[pay["project_id"]] += pay.get("amount",0)
        rows = []
        for p in projects:
            pid = str(p["_id"])
            total_cost = p.get("cost_estimation",{}).get("total_cost",0)
            paid = project_payments.get(pid, 0)
            balance = max(0, total_cost - paid)
            ps = "Paid" if paid >= total_cost and total_cost > 0 else "Partial" if paid > 0 else "Pending"
            rows.append({"customer": p.get("customer",{}).get("name",""), "ref": p.get("reference_number",""), "total_value": round(total_cost), "paid": round(paid), "balance": round(balance), "status": ps})
        chart_data = [{"name": "Paid", "value": sum(1 for r in rows if r["status"]=="Paid")}, {"name": "Partial", "value": sum(1 for r in rows if r["status"]=="Partial")}, {"name": "Pending", "value": sum(1 for r in rows if r["status"]=="Pending")}]
        return {"title": "Customer Credit Report", "summary": {"total_receivable": sum(r["total_value"] for r in rows), "collected": sum(r["paid"] for r in rows), "outstanding": sum(r["balance"] for r in rows)}, "rows": rows, "chart_data": [c for c in chart_data if c["value"]>0]}

    # === 6. TEAM PERFORMANCE ===
    elif report_type == "team_performance":
        team = defaultdict(lambda: {"name":"","assigned":0,"completed":0,"in_progress":0,"revenue":0})
        for p in projects:
            uid = p.get("created_by","")
            team[uid]["name"] = p.get("created_by_name","Unknown")
            team[uid]["assigned"] += 1
            if p.get("status") == "completed": team[uid]["completed"] += 1
            elif p.get("status") in ["submitted","approved"]: team[uid]["in_progress"] += 1
            if p.get("status") in ["approved","completed"]: team[uid]["revenue"] += p.get("cost_estimation",{}).get("total_cost",0)
        avg_load = sum(v["assigned"] for v in team.values())/len(team) if team else 0
        rows = [{"staff": v["name"], "assigned": v["assigned"], "in_progress": v["in_progress"], "completed": v["completed"], "revenue": round(v["revenue"]), "completion_rate": round((v["completed"]/v["assigned"])*100,1) if v["assigned"] else 0, "load_status": "Overloaded" if v["assigned"]>avg_load*1.5 else "Underutilized" if v["assigned"]<avg_load*0.5 else "Balanced"} for v in team.values()]
        chart_data = [{"name": r["staff"][:15], "value": r["assigned"]} for r in sorted(rows, key=lambda x: x["revenue"], reverse=True)[:8]]
        return {"title": "Team Performance Report", "summary": {"total_staff": len(rows), "avg_load": round(avg_load,1), "overloaded": sum(1 for r in rows if r["load_status"]=="Overloaded")}, "rows": sorted(rows, key=lambda x: x["revenue"], reverse=True), "chart_data": chart_data}

    # === 7. COMPLIANCE & TAX ===
    elif report_type == "compliance_tax":
        rows = []; total_gst = 0
        monthly_gst = defaultdict(float)
        for p in projects:
            ce = p.get("cost_estimation",{})
            gst = ce.get("gst_total",0); total_gst += gst
            monthly_gst[p.get("created_at","")[:7]] += gst
            rows.append({"customer": p.get("customer",{}).get("name",""), "ref": p.get("reference_number",""), "subtotal": round(ce.get("items_total",0)), "gst": round(gst), "total": round(ce.get("total_cost",0)), "date": p.get("created_at","")[:10]})
        chart_data = [{"name": k, "value": round(v)} for k,v in sorted(monthly_gst.items())[-6:]]
        return {"title": "Compliance & Tax Report", "summary": {"total_gst": round(total_gst), "invoices": len(rows)}, "rows": rows, "chart_data": chart_data}

    # === 8. CUSTOMER SATISFACTION ===
    elif report_type == "customer_satisfaction":
        rows = []
        for p in projects:
            fb = p.get("customer_feedback")
            rows.append({"customer": p.get("customer",{}).get("name",""), "ref": p.get("reference_number",""), "status": p.get("status",""), "feedback": fb or "No feedback", "has_feedback": bool(fb)})
        with_fb = sum(1 for r in rows if r["has_feedback"])
        chart_data = [{"name": "With Feedback", "value": with_fb}, {"name": "No Feedback", "value": len(rows)-with_fb}]
        return {"title": "Customer Satisfaction Report", "summary": {"total_customers": len(rows), "feedback_received": with_fb, "feedback_rate": round((with_fb/len(rows))*100,1) if rows else 0}, "rows": rows, "chart_data": chart_data}

    # === 9. INBOUND REPORT ===
    elif report_type == "inbound":
        pos = await db.purchase_orders.find().sort("created_at", -1).to_list(500)
        rows = []
        for po in pos:
            po_id = str(po["_id"])
            qc = po.get("qc", {}) or {}
            transport = po.get("transport", {}) or {}
            rows.append({"supplier": po.get("supplier_name",""), "items_count": len(po.get("items",[])), "total": round(po.get("total_amount",0)), "status": po.get("status",""), "qc_result": qc.get("overall","N/A"), "transporter": transport.get("transporter","N/A"), "vehicle": transport.get("vehicle","N/A"), "storage": po.get("storage_location","N/A"), "date": po.get("created_at","")[:10]})
        status_counts = {}
        for r in rows: status_counts[r["status"]] = status_counts.get(r["status"], 0) + 1
        chart_data = [{"name": k.replace("_"," ").title(), "value": v} for k,v in status_counts.items()]
        return {"title": "Inbound Report", "summary": {"total_pos": len(rows), "completed": sum(1 for r in rows if r["status"]=="completed"), "total_value": sum(r["total"] for r in rows)}, "rows": rows, "chart_data": chart_data}

    # === 10. OUTBOUND REPORT ===
    elif report_type == "outbound":
        dels = await db.deliveries.find().sort("created_at", -1).to_list(500)
        rows = [{"customer": d.get("customer_name",""), "items_count": len(d.get("items",[])), "transporter": d.get("transporter_name","N/A"), "vehicle": d.get("vehicle_number","N/A"), "distance_km": d.get("distance_km",0), "dispatch": d.get("dispatch_date",""), "delivery": d.get("delivery_date",""), "status": d.get("status","")} for d in dels]
        chart_data = [{"name": "Dispatched", "value": sum(1 for r in rows if r["status"]=="dispatched")}, {"name": "Delivered", "value": sum(1 for r in rows if r["status"]=="delivered")}]
        return {"title": "Outbound Report", "summary": {"total_deliveries": len(rows), "delivered": sum(1 for r in rows if r["status"]=="delivered"), "total_distance": sum(r["distance_km"] for r in rows)}, "rows": rows, "chart_data": [c for c in chart_data if c["value"]>0]}

    # === 11. AUDIT REPORT ===
    elif report_type == "audit":
        audits = await db.audits.find().sort("created_at", -1).to_list(500)
        rows = [{"title": a.get("title",""), "auditor": a.get("auditor_name",""), "status": a.get("status",""), "checklist_items": len(a.get("checklist",[])), "issues_count": len(a.get("issues",[])), "deadline": a.get("deadline","N/A"), "date": a.get("created_at","")[:10]} for a in audits]
        chart_data = [{"name": "Open", "value": sum(1 for r in rows if r["status"]=="open")}, {"name": "In Progress", "value": sum(1 for r in rows if r["status"]=="in_progress")}, {"name": "Resolved", "value": sum(1 for r in rows if r["status"]=="resolved")}]
        return {"title": "Audit Report", "summary": {"total_audits": len(rows), "open": sum(1 for r in rows if r["status"]=="open"), "total_issues": sum(r["issues_count"] for r in rows)}, "rows": rows, "chart_data": [c for c in chart_data if c["value"]>0]}

    # === 12. MARKETING REPORT ===
    elif report_type == "marketing":
        lead_updates = await db.daily_updates.find({"update_type": "leads"}).sort("created_at", -1).to_list(500)
        total_leads = sum(int(u.get("data",{}).get("total_leads",0) or 0) for u in lead_updates)
        qualified = sum(int(u.get("data",{}).get("qualified_leads",0) or 0) for u in lead_updates)
        site_visits = sum(int(u.get("data",{}).get("site_visits",0) or 0) for u in lead_updates)
        quotes_sent = sum(int(u.get("data",{}).get("quotes_sent",0) or 0) for u in lead_updates)
        followups = sum(int(u.get("data",{}).get("followups",0) or 0) for u in lead_updates)
        conversions = sum(int(u.get("data",{}).get("conversions",0) or 0) for u in lead_updates)
        conversion_rate = round((conversions / total_leads) * 100, 1) if total_leads > 0 else 0
        rows = [{"date": u.get("created_at","")[:10], "total_leads": u.get("data",{}).get("total_leads",0), "qualified": u.get("data",{}).get("qualified_leads",0), "site_visits": u.get("data",{}).get("site_visits",0), "quotes_sent": u.get("data",{}).get("quotes_sent",0), "conversions": u.get("data",{}).get("conversions",0), "by": u.get("created_by_name","")} for u in lead_updates]
        chart_data = [{"name": "Leads", "value": total_leads}, {"name": "Qualified", "value": qualified}, {"name": "Site Visits", "value": site_visits}, {"name": "Quotes", "value": quotes_sent}, {"name": "Conversions", "value": conversions}]
        if not rows:
            rows = [{"date": "-", "total_leads": 0, "qualified": 0, "site_visits": 0, "quotes_sent": 0, "conversions": 0, "by": "No data yet"}]
        return {"title": "Marketing Report", "summary": {"total_leads": total_leads, "qualified_leads": qualified, "quotes_sent": quotes_sent, "conversion_rate": conversion_rate}, "rows": rows, "chart_data": [c for c in chart_data if c["value"]>0]}

    # === 13. EXCESS MATERIAL REPORT (Iter 42 Change 4) ===
    elif report_type == "excess_material":
        docs = await db.material_reconciliation.find({}).to_list(2000)
        rows = []
        item_agg: Dict[str, Dict[str, Any]] = {}
        recoverable_value = 0.0
        damaged_total = 0.0
        for d in docs:
            proj = await db.projects.find_one({"_id": ObjectId(d["project_id"])}) if d.get("project_id") else None
            pname = (proj.get("customer", {}) or {}).get("name") if proj else d.get("project_id")
            variance_value_total = 0.0
            at_site_value_total = 0.0
            for l in d.get("lines", []):
                variance_value_total += l.get("variance_value", 0) or 0
                at_site_value_total += (l.get("qty_at_site", 0) or 0) * (l.get("unit_cost", 0) or 0)
                damaged_total += l.get("qty_damaged", 0) or 0
                key = l.get("name") or "Unknown"
                agg = item_agg.setdefault(key, {"name": key, "qty_issued": 0.0, "qty_consumed": 0.0, "variance": 0.0})
                agg["qty_issued"] += l.get("qty_issued", 0) or 0
                agg["qty_consumed"] += l.get("qty_consumed", 0) or 0
                agg["variance"] += l.get("variance", 0) or 0
            recoverable_value += at_site_value_total
            rows.append({"project": pname, "status": d.get("status"), "variance_value": round(variance_value_total, 2),
                         "value_at_site": round(at_site_value_total, 2), "reconciled_at": (d.get("reconciled_at") or "")[:10]})
        by_item = sorted(item_agg.values(), key=lambda x: abs(x["variance"]), reverse=True)[:10]
        chart_data = [{"name": it["name"][:14], "value": round(it["variance"], 1)} for it in by_item if it["variance"] > 0]
        if not rows:
            rows = [{"project": "-", "status": "-", "variance_value": 0, "value_at_site": 0, "reconciled_at": "No data yet"}]
        return {"title": "Excess Material Report", "summary": {
            "reconciliations": len(docs), "recoverable_value": round(recoverable_value, 2),
            "damaged_qty": round(damaged_total, 2),
        }, "rows": rows, "chart_data": chart_data}

    # === 14. AMC REPORT (Iter 43 Change 4) ===
    elif report_type == "amc":
        contracts = await db.amc_contracts.find({}).to_list(2000)
        active = [c for c in contracts if c.get("status") == "active"]
        arr = sum(c.get("annual_value", 0) or 0 for c in active)
        outstanding = sum(c.get("outstanding", 0) or 0 for c in contracts)
        renewed = sum(1 for c in contracts if c.get("status") == "renewed")
        churned = sum(1 for c in contracts if c.get("status") == "cancelled")
        renewal_rate = round((renewed / (renewed + churned)) * 100, 1) if (renewed + churned) else 0
        by_type: Dict[str, float] = {}
        for c in active:
            by_type[c.get("contract_type", "other")] = by_type.get(c.get("contract_type", "other"), 0) + (c.get("annual_value", 0) or 0)
        rows = [{"contract_number": c.get("contract_number"), "customer": c.get("customer_name"),
                 "contract_type": c.get("contract_type", "-"), "annual_value": round(c.get("annual_value", 0) or 0, 2),
                 "status": c.get("status"), "start_date": (c.get("start_date") or "")[:10],
                 "end_date": (c.get("end_date") or "")[:10], "outstanding": round(c.get("outstanding", 0) or 0, 2)} for c in contracts]
        if not rows:
            rows = [{"contract_number": "-", "customer": "-", "contract_type": "-", "annual_value": 0, "status": "No data yet", "start_date": "-", "end_date": "-", "outstanding": 0}]
        chart_data = [{"name": k.replace("_", " ").title(), "value": round(v, 1)} for k, v in by_type.items()]
        return {"title": "AMC Contracts Report", "summary": {
            "active_contracts": len(active), "arr": round(arr, 2), "mrr": round(arr / 12, 2),
            "renewal_rate": renewal_rate, "outstanding": round(outstanding, 2),
        }, "rows": rows, "chart_data": chart_data}

    # === 15. ASSETS REPORT (Iter 43 Change 4) ===
    elif report_type == "assets":
        assets_docs = await db.assets.find({"active": {"$ne": False}}).to_list(2000)
        by_status: Dict[str, int] = {}
        total_book_value = 0.0
        for a in assets_docs:
            by_status[a.get("status", "available")] = by_status.get(a.get("status", "available"), 0) + 1
            cost = a.get("purchase_cost", 0) or 0
            life = a.get("useful_life_years", 5) or 5
            try:
                purchased = datetime.fromisoformat(a["purchase_date"]) if a.get("purchase_date") else None
                if purchased and purchased.tzinfo is None: purchased = purchased.replace(tzinfo=timezone.utc)
            except Exception:
                purchased = None
            if purchased and life > 0:
                years_elapsed = (datetime.now(timezone.utc) - purchased).days / 365.25
                total_book_value += cost * max(0.0, 1 - (years_elapsed / life))
            else:
                total_book_value += cost
        rows = [{"asset_code": a.get("asset_code"), "name": a.get("name"), "category": a.get("category"),
                 "status": a.get("status"), "assigned_to": a.get("assigned_to_name") or "-",
                 "purchase_cost": round(a.get("purchase_cost", 0) or 0, 2)} for a in assets_docs]
        if not rows:
            rows = [{"asset_code": "-", "name": "-", "category": "-", "status": "No data yet", "assigned_to": "-", "purchase_cost": 0}]
        chart_data = [{"name": k.replace("_", " ").title(), "value": v} for k, v in by_status.items()]
        return {"title": "Assets Register Report", "summary": {
            "total_assets": len(assets_docs), "total_book_value": round(total_book_value, 2),
            "issued": by_status.get("issued", 0), "in_maintenance": by_status.get("in_maintenance", 0),
        }, "rows": rows, "chart_data": chart_data}

    # === 16. TOOLS REPORT (Iter 43 Change 4) — tool/equipment subset with utilisation ===
    elif report_type == "tools":
        tool_categories = {"power_tool", "hand_tool", "test_equipment"}
        tools_docs = await db.assets.find({"active": {"$ne": False}, "category": {"$in": list(tool_categories)}}).to_list(2000)
        tool_ids = [str(t["_id"]) for t in tools_docs]
        maint_docs = await db.asset_maintenance.find({"asset_id": {"$in": tool_ids}}).to_list(5000) if tool_ids else []
        maint_cost_by_tool: Dict[str, float] = {}
        for m in maint_docs:
            maint_cost_by_tool[m.get("asset_id")] = maint_cost_by_tool.get(m.get("asset_id"), 0) + (m.get("cost", 0) or 0)
        issued_now = sum(1 for t in tools_docs if t.get("status") == "issued")
        rows = [{"asset_code": t.get("asset_code"), "name": t.get("name"), "category": t.get("category"),
                 "status": t.get("status"), "assigned_to": t.get("assigned_to_name") or "-",
                 "maintenance_cost": round(maint_cost_by_tool.get(str(t["_id"]), 0), 2)} for t in tools_docs]
        if not rows:
            rows = [{"asset_code": "-", "name": "-", "category": "-", "status": "No data yet", "assigned_to": "-", "maintenance_cost": 0}]
        chart_data = [{"name": "Available", "value": sum(1 for t in tools_docs if t.get("status") == "available")},
                      {"name": "Issued", "value": issued_now},
                      {"name": "In Maintenance", "value": sum(1 for t in tools_docs if t.get("status") == "in_maintenance")}]
        return {"title": "Tools Utilisation Report", "summary": {
            "total_tools": len(tools_docs), "issued_now": issued_now,
            "utilisation_pct": round((issued_now / len(tools_docs)) * 100, 1) if tools_docs else 0,
            "total_maintenance_cost": round(sum(maint_cost_by_tool.values()), 2),
        }, "rows": rows, "chart_data": [c for c in chart_data if c["value"] > 0]}

    # === 17. EXPENSES REPORT (Iter 43 Change 4) ===
    elif report_type == "expenses":
        q: Dict[str, Any] = {"entry_type": {"$in": ["operational_expense", "marketing_expense", "gst_input", "gst_paid"]}}
        if date_from or date_to:
            q["entry_date"] = {}
            if date_from: q["entry_date"]["$gte"] = date_from
            if date_to: q["entry_date"]["$lte"] = date_to
        entries = await db.account_entries.find(q).sort("entry_date", -1).to_list(5000)
        by_type: Dict[str, float] = {}
        for e in entries:
            by_type[e.get("entry_type")] = by_type.get(e.get("entry_type"), 0) + (e.get("amount", 0) or 0)
        rows = [{"date": e.get("entry_date"), "type": (e.get("entry_type") or "").replace("_", " ").title(),
                 "amount": round(e.get("amount", 0) or 0, 2), "description": e.get("description", ""),
                 "channel": e.get("marketing_channel") or "-", "entered_by": e.get("entered_by", "")} for e in entries]
        if not rows:
            rows = [{"date": "-", "type": "-", "amount": 0, "description": "No data yet", "channel": "-", "entered_by": "-"}]
        op_exp = by_type.get("operational_expense", 0)
        mkt_exp = by_type.get("marketing_expense", 0)
        gst_input = by_type.get("gst_input", 0)
        gst_paid = by_type.get("gst_paid", 0)
        chart_data = [{"name": k.replace("_", " ").title(), "value": round(v, 1)} for k, v in by_type.items()]
        return {"title": "Expenses Report", "summary": {
            "operational_expense": round(op_exp, 2), "marketing_expense": round(mkt_exp, 2),
            "gst_input": round(gst_input, 2), "net_gst_liability": round(gst_paid - gst_input, 2),
        }, "rows": rows, "chart_data": chart_data}

    raise HTTPException(status_code=404, detail=f"Unknown report type: {report_type}")

# ================== ALERT ENGINE & RISK SCORING ==================

async def _get_thresholds():
    s = await db.settings.find_one({"type": "thresholds"})
    return {
        "min_margin_pct": (s or {}).get("min_margin_pct", 8),
        "max_material_variance_pct": (s or {}).get("max_material_variance_pct", 15),
        "payment_delay_days": (s or {}).get("payment_delay_days", 30),
        "max_project_duration_days": (s or {}).get("max_project_duration_days", 90),
        "underpriced_margin_pct": (s or {}).get("underpriced_margin_pct", 5)
    }

async def _generate_project_alerts(project, thresholds):
    alerts = []
    pid = str(project["_id"])
    ce = project.get("cost_estimation", {})
    selling = ce.get("total_cost", 0)
    margin = ce.get("margin_total", 0)
    margin_pct = round((margin / selling) * 100, 1) if selling else 0
    created_str = project.get("created_at", "")
    now = datetime.now(timezone.utc)

    # 1. Low Margin
    if selling > 0 and margin_pct < thresholds["min_margin_pct"]:
        alerts.append({"type": "low_margin", "severity": "high", "message": f"Margin {margin_pct}% below {thresholds['min_margin_pct']}% threshold", "impact": round((thresholds["min_margin_pct"] - margin_pct) / 100 * selling)})

    # 2. Underpriced Quote
    if selling > 0 and margin_pct < thresholds["underpriced_margin_pct"]:
        alerts.append({"type": "underpriced_quote", "severity": "high", "message": f"Quote margin {margin_pct}% critically low", "impact": round(selling * 0.05)})

    # 3. Payment Delay
    payments = await db.payments.find({"project_id": pid}).to_list(100)
    total_paid = sum(p.get("amount", 0) for p in payments)
    if selling > 0 and total_paid < selling and created_str:
        try:
            created_dt = datetime.fromisoformat(created_str.replace("Z", "+00:00"))
            days_since = (now - created_dt).days
            if days_since > thresholds["payment_delay_days"] and total_paid < selling:
                alerts.append({"type": "payment_delay", "severity": "high" if days_since > thresholds["payment_delay_days"] * 2 else "medium", "message": f"Payment overdue by {days_since - thresholds['payment_delay_days']} days. Outstanding: Rs {round(selling - total_paid):,}", "impact": round(selling - total_paid)})
        except (ValueError, TypeError):
            pass

    # 4. Project Delay
    if project.get("status") not in ["completed", "rejected", "draft"] and created_str:
        try:
            created_dt = datetime.fromisoformat(created_str.replace("Z", "+00:00"))
            days = (now - created_dt).days
            if days > thresholds["max_project_duration_days"]:
                alerts.append({"type": "project_delay", "severity": "medium", "message": f"Project open for {days} days (limit: {thresholds['max_project_duration_days']})", "impact": 0})
        except (ValueError, TypeError):
            pass

    # 5. Material Variance
    usage = await db.material_usage_logs.find({"project_id": pid}).to_list(100)
    for u in usage:
        est = u.get("estimated_qty", 0)
        act = u.get("actual_qty", 0)
        if est > 0 and act > 0:
            var_pct = ((act - est) / est) * 100
            if var_pct > thresholds["max_material_variance_pct"]:
                alerts.append({"type": "material_variance", "severity": "medium", "message": f"{u.get('item_name','')}: {round(var_pct,1)}% over estimated ({act} vs {est})", "impact": 0})

    return alerts

def _calc_risk_score(alerts):
    score = 0
    for a in alerts:
        if a["severity"] == "high": score += 25
        elif a["severity"] == "medium": score += 15
        else: score += 5
    return min(score, 100)

@api_router.get("/alerts/dashboard")
async def get_alerts_dashboard(request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Alerts dashboard requires admin/manager")
    thresholds = await _get_thresholds()
    projects = await db.projects.find({"deleted_at": {"$exists": False}, "status": {"$nin": ["draft"]}}).to_list(5000)
    all_alerts = []
    project_risks = []
    for p in projects:
        alerts = await _generate_project_alerts(p, thresholds)
        pid = str(p["_id"])
        risk = _calc_risk_score(alerts)
        for a in alerts:
            a["project_id"] = pid
            a["project_ref"] = p.get("reference_number", "")
            a["customer"] = p.get("customer", {}).get("name", "")
        all_alerts.extend(alerts)
        if risk > 0:
            project_risks.append({"id": pid, "ref": p.get("reference_number",""), "customer": p.get("customer",{}).get("name",""), "risk_score": risk, "risk_level": "High" if risk > 60 else "Medium" if risk > 30 else "Low", "alert_count": len(alerts), "status": p.get("status","")})
    total_leakage = sum(a.get("impact", 0) for a in all_alerts)
    by_type = {}
    for a in all_alerts:
        by_type.setdefault(a["type"], {"count": 0, "impact": 0})
        by_type[a["type"]]["count"] += 1
        by_type[a["type"]]["impact"] += a.get("impact", 0)
    chart_data = [{"name": k.replace("_", " ").title(), "value": v["impact"]} for k, v in by_type.items() if v["impact"] > 0]
    return {
        "total_leakage": round(total_leakage),
        "total_alerts": len(all_alerts),
        "risky_projects": len(project_risks),
        "top_risks": sorted(project_risks, key=lambda x: x["risk_score"], reverse=True)[:10],
        "alerts_by_type": by_type,
        "chart_data": chart_data,
        "thresholds": thresholds
    }

@api_router.get("/alerts/project/{project_id}")
async def get_project_alerts(project_id: str, request: Request):
    await get_current_user(request)
    project = await db.projects.find_one({"_id": ObjectId(project_id)})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    thresholds = await _get_thresholds()
    alerts = await _generate_project_alerts(project, thresholds)
    risk = _calc_risk_score(alerts)
    suggestions = []
    for a in alerts:
        if a["type"] == "low_margin": suggestions.append("Review pricing strategy or negotiate material costs")
        elif a["type"] == "payment_delay": suggestions.append("Immediate payment follow-up required")
        elif a["type"] == "material_variance": suggestions.append("Investigate installation team material handling")
        elif a["type"] == "project_delay": suggestions.append("Escalate to project manager for timeline review")
        elif a["type"] == "underpriced_quote": suggestions.append("Re-evaluate quotation margins before approval")
    return {"risk_score": risk, "risk_level": "High" if risk > 60 else "Medium" if risk > 30 else "Low", "alerts": alerts, "suggestions": list(set(suggestions))}

@api_router.get("/settings/thresholds")
async def get_thresholds(request: Request):
    await get_current_user(request)
    return await _get_thresholds()

@api_router.put("/settings/thresholds")
async def update_thresholds(body: ThresholdUpdate, request: Request):
    await require_permission(request, "can_manage_company")
    update = {}
    if body.min_margin_pct is not None: update["min_margin_pct"] = body.min_margin_pct
    if body.max_material_variance_pct is not None: update["max_material_variance_pct"] = body.max_material_variance_pct
    if body.payment_delay_days is not None: update["payment_delay_days"] = body.payment_delay_days
    if body.max_project_duration_days is not None: update["max_project_duration_days"] = body.max_project_duration_days
    if body.underpriced_margin_pct is not None: update["underpriced_margin_pct"] = body.underpriced_margin_pct
    await db.settings.update_one({"type": "thresholds"}, {"$set": {**update, "type": "thresholds"}}, upsert=True)
    return {"message": "Thresholds updated"}

# ================== AI RECOMMENDATIONS ==================

@api_router.post("/ai/recommendations")
async def get_ai_recommendations(data: AIRecommendationRequest, request: Request):
    await get_current_user(request)
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="AI service not configured")
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"solar-rec-{uuid.uuid4()}",
            system_message="""You are a solar energy consultant for Sensoper Controls and Renewables. 
            Provide concise, practical recommendations for solar system installations based on the user's energy consumption and site details.
            Focus on:
            1. Recommended system capacity (kW)
            2. Panel type and count
            3. Inverter recommendations
            4. Estimated savings
            5. ROI timeline
            Keep responses under 300 words and use bullet points."""
        )
        
        chat.with_model("openai", "gpt-5.2")
        
        prompt = f"""Based on the following details, provide solar system recommendations:
        
        - Monthly Electricity Consumption: {data.monthly_consumption_units} units
        - Sanctioned Load: {data.sanction_load_kw} kW
        - Roof Type: {data.roof_type}
        - Budget Range: {data.budget_range or 'Not specified'}
        
        Please recommend the optimal solar system configuration."""
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        return {"recommendation": response}
    
    except ImportError:
        return {"recommendation": f"""Based on your monthly consumption of {data.monthly_consumption_units} units:

**Recommended System:**
- Capacity: {max(1, data.monthly_consumption_units / 120):.1f} kW
- Panels: {int(max(1, data.monthly_consumption_units / 120) * 1000 / 540) + 1} x 540W panels
- Inverter: {max(1, data.monthly_consumption_units / 120):.1f} kW grid-tied

**Estimated Benefits:**
- Monthly Savings: Rs. {data.monthly_consumption_units * data.sanction_load_kw:.0f} approx
- ROI: 4-5 years
- System Life: 25+ years

*For detailed AI-powered analysis, please configure the AI service.*"""}
    except Exception as e:
        logger.error(f"AI recommendation error: {e}")
        return {"recommendation": f"Could not generate AI recommendation. Basic estimate: {max(1, data.monthly_consumption_units / 120):.1f} kW system recommended."}

# ================== APPROVALS ==================

APPROVAL_TYPES = ["deletion", "margin_change", "quotation_approval", "inventory_edit", "user_access_change"]

@api_router.get("/approvals")
async def list_approvals(request: Request, status: str = None, type: str = None, requested_by: str = None):
    user = await get_current_user(request)
    query = {}
    if status:
        query["status"] = status
    if type:
        query["type"] = type
    if requested_by:
        query["requested_by"] = requested_by
    # Staff can only see their own requests
    if user["role"] == "staff":
        query["requested_by"] = user["id"]
    
    approvals = await db.approvals.find(query).sort("timestamp", -1).to_list(200)
    return [
        {
            "id": str(a["_id"]),
            "type": a["type"],
            "requested_by": a["requested_by"],
            "requested_by_name": a.get("requested_by_name", "Unknown"),
            "role": a.get("role", "staff"),
            "description": a.get("description", ""),
            "entity_type": a.get("entity_type", ""),
            "entity_id": a.get("entity_id", ""),
            "data_payload": a.get("data_payload", {}),
            "status": a["status"],
            "approved_by": a.get("approved_by"),
            "approved_by_name": a.get("approved_by_name"),
            "rejection_reason": a.get("rejection_reason"),
            "timestamp": a["timestamp"],
            "resolved_at": a.get("resolved_at")
        }
        for a in approvals
    ]

@api_router.get("/approvals/pending-count")
async def get_pending_approvals_count(request: Request):
    user = await get_current_user(request)
    if user["role"] == "staff":
        return {"count": 0}
    count = await db.approvals.count_documents({"status": "pending"})
    return {"count": count}

@api_router.post("/approvals")
async def create_approval(request: Request):
    user = await get_current_user(request)
    body = await request.json()
    
    approval_type = body.get("type")
    if approval_type not in APPROVAL_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid type. Must be: {', '.join(APPROVAL_TYPES)}")
    
    doc = {
        "type": approval_type,
        "requested_by": user["id"],
        "requested_by_name": user["name"],
        "role": user["role"],
        "description": body.get("description", ""),
        "entity_type": body.get("entity_type", ""),
        "entity_id": body.get("entity_id", ""),
        "data_payload": body.get("data_payload", {}),
        "status": "pending",
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    result = await db.approvals.insert_one(doc)
    await create_audit_log(user["id"], user["name"], "create", "approval", str(result.inserted_id), None, {"type": approval_type})
    return {"id": str(result.inserted_id), "message": "Approval request created"}

@api_router.put("/approvals/{approval_id}/approve")
async def approve_request(approval_id: str, request: Request):
    user = await get_current_user(request)
    
    approval = await db.approvals.find_one({"_id": ObjectId(approval_id)})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="Approval already resolved")
    
    # Check permission based on type
    perm_map = {
        "deletion": "can_approve_deletion",
        "margin_change": "can_approve_margin",
        "quotation_approval": "can_approve_quotation",
        "inventory_edit": "can_approve_inventory",
        "user_access_change": "can_change_user_access"
    }
    perm = perm_map.get(approval["type"])
    if perm:
        has_perm = await check_permission(user, perm)
        if not has_perm:
            raise HTTPException(status_code=403, detail="You don't have permission to approve this type")
    
    # Execute the approval action
    executed = await execute_approval_action(approval)
    
    await db.approvals.update_one(
        {"_id": ObjectId(approval_id)},
        {"$set": {
            "status": "approved",
            "approved_by": user["id"],
            "approved_by_name": user["name"],
            "resolved_at": datetime.now(timezone.utc).isoformat(),
            "execution_result": executed
        }}
    )
    await create_audit_log(user["id"], user["name"], "approve", "approval", approval_id)
    return {"message": "Request approved and executed", "execution_result": executed}

@api_router.put("/approvals/{approval_id}/reject")
async def reject_request(approval_id: str, request: Request):
    user = await get_current_user(request)
    body = await request.json()
    reason = body.get("reason", "")
    
    approval = await db.approvals.find_one({"_id": ObjectId(approval_id)})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")
    if approval["status"] != "pending":
        raise HTTPException(status_code=400, detail="Approval already resolved")
    
    perm_map = {
        "deletion": "can_approve_deletion",
        "margin_change": "can_approve_margin",
        "quotation_approval": "can_approve_quotation",
        "inventory_edit": "can_approve_inventory",
        "user_access_change": "can_change_user_access"
    }
    perm = perm_map.get(approval["type"])
    if perm:
        has_perm = await check_permission(user, perm)
        if not has_perm:
            raise HTTPException(status_code=403, detail="You don't have permission to reject this type")
    
    await db.approvals.update_one(
        {"_id": ObjectId(approval_id)},
        {"$set": {
            "status": "rejected",
            "approved_by": user["id"],
            "approved_by_name": user["name"],
            "rejection_reason": reason,
            "resolved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    await create_audit_log(user["id"], user["name"], "reject", "approval", approval_id)
    return {"message": "Request rejected"}

async def execute_approval_action(approval: dict) -> str:
    """Execute the actual action when an approval is granted"""
    atype = approval["type"]
    payload = approval.get("data_payload", {})
    entity_id = approval.get("entity_id", "")
    
    try:
        if atype == "deletion":
            entity_type = approval.get("entity_type", "")
            if entity_type == "project" and entity_id:
                await db.projects.update_one(
                    {"_id": ObjectId(entity_id)},
                    {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat(), "status": "deleted"}}
                )
                return "Project deleted"
            elif entity_type == "inventory_item" and entity_id:
                await db.inventory_items.delete_one({"_id": ObjectId(entity_id)})
                return "Inventory item deleted"
        
        elif atype == "margin_change":
            if entity_id and "item_margins" in payload:
                project = await db.projects.find_one({"_id": ObjectId(entity_id)})
                if project:
                    items = project.get("selected_items", [])
                    for mu in payload["item_margins"]:
                        idx = mu.get("index")
                        if idx is not None and 0 <= idx < len(items):
                            items[idx]["margin_percentage"] = float(mu.get("margin_percentage", 0))
                    new_est = calculate_cost_estimation(items, project.get("manual_costs", []))
                    await db.projects.update_one(
                        {"_id": ObjectId(entity_id)},
                        {"$set": {"selected_items": items, "cost_estimation": new_est, "updated_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    return "Margins updated"
        
        elif atype == "quotation_approval":
            if entity_id:
                await db.projects.update_one(
                    {"_id": ObjectId(entity_id)},
                    {"$set": {"status": "approved", "approved_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                return "Quotation approved"
        
        elif atype == "inventory_edit":
            if entity_id and payload:
                update_fields = {}
                for k in ["name", "unit_price", "gst_percentage", "quantity", "category", "description"]:
                    if k in payload:
                        update_fields[k] = payload[k]
                if update_fields:
                    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
                    await db.inventory_items.update_one(
                        {"_id": ObjectId(entity_id)},
                        {"$set": update_fields}
                    )
                    return "Inventory item updated"
        
        elif atype == "user_access_change":
            user_id = payload.get("user_id")
            new_role = payload.get("new_role")
            if user_id and new_role:
                await db.users.update_one(
                    {"_id": ObjectId(user_id)},
                    {"$set": {"role": new_role}}
                )
                return f"User role changed to {new_role}"
        
        return "No action taken"
    except Exception as e:
        logger.error(f"Approval execution error: {e}")
        return f"Error: {str(e)}"

# ================== PERMISSIONS MANAGEMENT ==================

@api_router.get("/permissions")
async def get_all_permissions(request: Request):
    user = await require_role("admin")(request)
    roles = ["admin", "manager", "staff"]
    result = {}
    for role in roles:
        result[role] = await get_permissions(role)
    return result

@api_router.get("/permissions/{role_name}")
async def get_role_permissions(role_name: str, request: Request):
    user = await get_current_user(request)
    perms = await get_permissions(role_name)
    return {"role": role_name, "permissions": perms}

@api_router.put("/permissions/{role_name}")
async def update_role_permissions(role_name: str, request: Request):
    user = await require_role("admin")(request)
    body = await request.json()
    permissions = body.get("permissions", {})
    
    if role_name not in ["admin", "manager", "staff"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    await db.role_permissions.update_one(
        {"role_name": role_name},
        {"$set": {"role_name": role_name, "permissions": permissions, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": user["name"]}},
        upsert=True
    )
    await create_audit_log(user["id"], user["name"], "update", "permissions", role_name, None, permissions)
    return {"message": f"Permissions for {role_name} updated", "permissions": permissions}

@api_router.get("/company/logo-base64")
async def get_logo_base64():
    """Return company logo as base64 data URL to bypass CORS for PDF generation"""
    company = await db.company_profiles.find_one({"is_active": True})
    if not company:
        company = await db.company_profiles.find_one({})
    logo_url = company.get("logo_url") if company else None
    if not logo_url:
        return {"logo_base64": None}
    # If already base64
    if logo_url.startswith("data:"):
        return {"logo_base64": logo_url}
    try:
        import httpx
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(logo_url)
            if resp.status_code == 200:
                ct = resp.headers.get("content-type", "image/png")
                b64 = base64.b64encode(resp.content).decode("utf-8")
                return {"logo_base64": f"data:{ct};base64,{b64}"}
    except Exception as e:
        logger.error(f"Failed to fetch logo for base64: {e}")
    return {"logo_base64": None}

# ================== FORM TABS (Dynamic Form Engine) ==================

@api_router.get("/form-tabs")
async def get_form_tabs(request: Request):
    user = await get_current_user(request)
    query = {}
    if user["role"] != "admin":
        query["active"] = True
        query["roles_visible"] = user["role"]
    tabs = await db.form_tabs.find(query, {"_id": 0}).sort("order", 1).to_list(100)
    for t in tabs:
        doc = await db.form_tabs.find_one({"slug": t.get("slug")})
        if doc:
            t["id"] = str(doc["_id"])
        t.setdefault("system", False)
    return tabs

@api_router.post("/form-tabs")
async def create_form_tab(tab: FormTabCreate, request: Request):
    user = await require_permission(request, "can_manage_company")
    slug = tab.name.strip().lower().replace(" ", "_")
    slug = "".join(c for c in slug if c.isalnum() or c == "_")
    existing = await db.form_tabs.find_one({"slug": slug})
    if existing:
        raise HTTPException(status_code=400, detail="A tab with this name already exists")
    # Insert before site_docs: find max order among non-site_docs tabs
    non_docs = await db.form_tabs.find({"slug": {"$ne": "site_docs"}}).sort("order", -1).to_list(1)
    next_order = (non_docs[0]["order"] + 1) if non_docs else 1
    # Push site_docs order up
    await db.form_tabs.update_one({"slug": "site_docs"}, {"$set": {"order": next_order + 1}})
    doc = {
        "name": tab.name.strip(),
        "slug": slug,
        "fields": [f.model_dump() for f in tab.fields],
        "roles_visible": tab.roles_visible,
        "order": next_order,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.form_tabs.insert_one(doc)
    await create_audit_log(user["id"], user["name"], "create", "form_tab", str(result.inserted_id), None, {"name": tab.name})
    return {"id": str(result.inserted_id), "message": "Form tab created"}

@api_router.put("/form-tabs/reorder")
async def reorder_form_tabs(request: Request):
    await require_permission(request, "can_manage_company")
    body = await request.json()
    order_list = body.get("order", [])
    for idx, tab_id in enumerate(order_list):
        await db.form_tabs.update_one({"_id": ObjectId(tab_id)}, {"$set": {"order": idx + 1, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Tabs reordered"}

@api_router.put("/form-tabs/{tab_id}")
async def update_form_tab(tab_id: str, updates: FormTabUpdate, request: Request):
    user = await require_permission(request, "can_manage_company")
    tab = await db.form_tabs.find_one({"_id": ObjectId(tab_id)})
    if not tab:
        raise HTTPException(status_code=404, detail="Tab not found")
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if updates.name is not None:
        update_data["name"] = updates.name.strip()
        # System tabs keep their original slug (used to identify hardcoded content)
        if not tab.get("system"):
            new_slug = updates.name.strip().lower().replace(" ", "_")
            new_slug = "".join(c for c in new_slug if c.isalnum() or c == "_")
            existing = await db.form_tabs.find_one({"slug": new_slug, "_id": {"$ne": ObjectId(tab_id)}})
            if existing:
                raise HTTPException(status_code=400, detail="A tab with this name already exists")
            update_data["slug"] = new_slug
    if updates.fields is not None:
        update_data["fields"] = [f.model_dump() for f in updates.fields]
    if updates.roles_visible is not None:
        update_data["roles_visible"] = updates.roles_visible
    if updates.active is not None:
        update_data["active"] = updates.active
    await db.form_tabs.update_one({"_id": ObjectId(tab_id)}, {"$set": update_data})
    await create_audit_log(user["id"], user["name"], "update", "form_tab", tab_id)
    return {"message": "Tab updated"}

@api_router.delete("/form-tabs/{tab_id}")
async def delete_form_tab(tab_id: str, request: Request):
    user = await require_permission(request, "can_manage_company")
    tab = await db.form_tabs.find_one({"_id": ObjectId(tab_id)})
    if not tab:
        raise HTTPException(status_code=404, detail="Tab not found")
    await db.form_tabs.delete_one({"_id": ObjectId(tab_id)})
    await create_audit_log(user["id"], user["name"], "delete", "form_tab", tab_id, {"name": tab.get("name")})
    return {"message": "Tab deleted"}

# ================== DAILY UPDATES ==================

@api_router.post("/daily-updates")
async def create_daily_update(update: DailyUpdateCreate, request: Request):
    user = await get_current_user(request)
    doc = {
        "project_id": update.project_id,
        "update_type": update.update_type,
        "data": update.data,
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.daily_updates.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Update created"}

@api_router.get("/daily-updates")
async def list_daily_updates(request: Request, project_id: str = None, update_type: str = None, date_from: str = None, date_to: str = None):
    await get_current_user(request)
    query = {}
    if project_id:
        query["project_id"] = project_id
    if update_type:
        query["update_type"] = update_type
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    updates = await db.daily_updates.find(query).sort("created_at", -1).to_list(500)
    for u in updates:
        u["id"] = str(u.pop("_id"))
    return updates

@api_router.get("/daily-updates/project/{project_id}")
async def get_project_updates(project_id: str, request: Request):
    await get_current_user(request)
    updates = await db.daily_updates.find({"project_id": project_id}).sort("created_at", -1).to_list(200)
    for u in updates:
        u["id"] = str(u.pop("_id"))
    return updates

@api_router.put("/daily-updates/{update_id}")
async def update_daily_update(update_id: str, body: DailyUpdateUpdate, request: Request):
    user = await get_current_user(request)
    doc = await db.daily_updates.find_one({"_id": ObjectId(update_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Update not found")
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if body.data is not None:
        update_data["data"] = body.data
    if body.update_type is not None:
        update_data["update_type"] = body.update_type
    await db.daily_updates.update_one({"_id": ObjectId(update_id)}, {"$set": update_data})
    return {"message": "Update modified"}

@api_router.delete("/daily-updates/{update_id}")
async def delete_daily_update(update_id: str, request: Request):
    await get_current_user(request)
    result = await db.daily_updates.delete_one({"_id": ObjectId(update_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Update not found")
    return {"message": "Update deleted"}

# ================== PAYMENTS ==================

@api_router.post("/payments")
async def create_payment(payment: PaymentCreate, request: Request):
    user = await get_current_user(request)
    project = await db.projects.find_one({"_id": ObjectId(payment.project_id)})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    doc = {
        "project_id": payment.project_id,
        "amount": payment.amount,
        "payment_method": payment.payment_method,
        "notes": payment.notes,
        "recorded_by": user["id"],
        "recorded_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.payments.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Payment recorded"}

@api_router.get("/payments/project/{project_id}")
async def get_project_payments(project_id: str, request: Request):
    await get_current_user(request)
    payments = await db.payments.find({"project_id": project_id}).sort("created_at", -1).to_list(200)
    for p in payments:
        p["id"] = str(p.pop("_id"))
    return payments

# ================== MATERIAL USAGE LOGS ==================

@api_router.post("/material-usage")
async def create_material_usage(usage: MaterialUsageCreate, request: Request):
    user = await get_current_user(request)
    doc = {
        "project_id": usage.project_id,
        "item_name": usage.item_name,
        "estimated_qty": usage.estimated_qty,
        "actual_qty": usage.actual_qty,
        "wastage": usage.wastage,
        "variance": usage.actual_qty - usage.estimated_qty,
        "notes": usage.notes,
        "logged_by": user["id"],
        "logged_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.material_usage_logs.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Usage logged"}

@api_router.get("/material-usage/project/{project_id}")
async def get_project_material_usage(project_id: str, request: Request):
    await get_current_user(request)
    logs = await db.material_usage_logs.find({"project_id": project_id}).sort("created_at", -1).to_list(200)
    for l in logs:
        l["id"] = str(l.pop("_id"))
    return logs

# ================== DATA COMPLETENESS ==================

@api_router.get("/projects/{project_id}/completeness")
async def get_project_completeness(project_id: str, request: Request):
    await get_current_user(request)
    project = await db.projects.find_one({"_id": ObjectId(project_id)})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    score = 0
    checks = {}
    # Customer details (20%)
    cust = project.get("customer", {})
    cust_ok = bool(cust.get("name") and cust.get("phone") and cust.get("address"))
    checks["customer_details"] = cust_ok
    if cust_ok: score += 20
    # Site data (20%)
    loc = project.get("location", {})
    site_ok = bool(loc.get("site_location_words") or loc.get("address"))
    checks["site_data"] = site_ok
    if site_ok: score += 20
    # Electrical data (15%)
    elec = project.get("electrical", {})
    elec_ok = bool(elec.get("sanction_load_kw") and elec.get("monthly_consumption_units"))
    checks["electrical_data"] = elec_ok
    if elec_ok: score += 15
    # Costing (20%)
    items = project.get("selected_items", [])
    cost_ok = len(items) > 0
    checks["costing"] = cost_ok
    if cost_ok: score += 20
    # Drive link (10%)
    drive_ok = bool(project.get("drive_folder_link"))
    checks["site_docs"] = drive_ok
    if drive_ok: score += 10
    # Daily updates (15%)
    update_count = await db.daily_updates.count_documents({"project_id": project_id})
    updates_ok = update_count >= 1
    checks["daily_updates"] = updates_ok
    if updates_ok: score += 15
    return {"score": min(score, 100), "checks": checks, "update_count": update_count}

# ================== PROJECT REPORT (Per-Project Download) ==================

@api_router.get("/projects/{project_id}/report")
async def get_project_report(project_id: str, request: Request):
    await get_current_user(request)
    project = await db.projects.find_one({"_id": ObjectId(project_id)})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    payments = await db.payments.find({"project_id": project_id}).to_list(100)
    for p in payments:
        p["id"] = str(p.pop("_id"))
    material_logs = await db.material_usage_logs.find({"project_id": project_id}).to_list(100)
    for m in material_logs:
        m["id"] = str(m.pop("_id"))
    updates = await db.daily_updates.find({"project_id": project_id}).sort("created_at", -1).to_list(100)
    for u in updates:
        u["id"] = str(u.pop("_id"))
    total_paid = sum(p.get("amount", 0) for p in payments)
    total_cost = project.get("cost_estimation", {}).get("total_cost", 0)
    return {
        "project": {
            "id": project_id,
            "ref": project.get("reference_number", ""),
            "customer": project.get("customer", {}),
            "location": project.get("location", {}),
            "status": project.get("status", ""),
            "electrical": project.get("electrical", {}),
            "solar_system": project.get("solar_system", {}),
            "cost_estimation": project.get("cost_estimation", {}),
            "selected_items": project.get("selected_items", []),
            "manual_costs": project.get("manual_costs", []),
            "site_measurements": project.get("site_measurements", {}),
            "created_at": project.get("created_at", ""),
            "updated_at": project.get("updated_at", "")
        },
        "payments": payments,
        "total_paid": total_paid,
        "balance": max(0, total_cost - total_paid),
        "payment_status": "Paid" if total_paid >= total_cost and total_cost > 0 else "Partial" if total_paid > 0 else "Pending",
        "material_usage": material_logs,
        "daily_updates": updates
    }

# ================== CUSTOMER CREDITS ==================

@api_router.post("/credits")
async def create_credit(credit: CustomerCreditCreate, request: Request):
    user = await get_current_user(request)
    doc = {
        "customer_name": credit.customer_name, "customer_phone": credit.customer_phone,
        "invoice_ref": credit.invoice_ref, "total_amount": credit.total_amount,
        "amount_paid": 0, "balance": credit.total_amount,
        "due_date": credit.due_date, "status": "active", "notes": credit.notes,
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.customer_credits.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Credit created"}

@api_router.get("/credits")
async def list_credits(request: Request, status: str = None):
    await get_current_user(request)
    query = {}
    if status and status != "all": query["status"] = status
    credits = await db.customer_credits.find(query).sort("created_at", -1).to_list(500)
    now = datetime.now(timezone.utc)
    for c in credits:
        c["id"] = str(c.pop("_id"))
        if c.get("due_date") and c["status"] == "active":
            try:
                due = datetime.fromisoformat(c["due_date"])
                if due.tzinfo is None:
                    due = due.replace(tzinfo=timezone.utc)
                if now > due:
                    c["status"] = "overdue"
                    await db.customer_credits.update_one({"_id": ObjectId(c["id"])}, {"$set": {"status": "overdue"}})
            except (ValueError, TypeError):
                pass
    return credits

@api_router.post("/credits/{credit_id}/pay")
async def record_credit_payment(credit_id: str, payment: CreditPaymentCreate, request: Request):
    user = await get_current_user(request)
    credit = await db.customer_credits.find_one({"_id": ObjectId(credit_id)})
    if not credit: raise HTTPException(status_code=404, detail="Credit not found")
    pay_doc = {"credit_id": credit_id, "amount": payment.amount, "payment_method": payment.payment_method, "notes": payment.notes, "recorded_by": user["name"], "created_at": datetime.now(timezone.utc).isoformat()}
    await db.credit_payments.insert_one(pay_doc)
    new_paid = credit.get("amount_paid", 0) + payment.amount
    new_balance = max(0, credit["total_amount"] - new_paid)
    new_status = "closed" if new_balance <= 0 else credit.get("status", "active")
    await db.customer_credits.update_one({"_id": ObjectId(credit_id)}, {"$set": {"amount_paid": new_paid, "balance": new_balance, "status": new_status}})
    return {"message": "Payment recorded", "new_balance": new_balance}

@api_router.get("/credits/{credit_id}/payments")
async def get_credit_payments(credit_id: str, request: Request):
    await get_current_user(request)
    payments = await db.credit_payments.find({"credit_id": credit_id}).sort("created_at", -1).to_list(100)
    for p in payments: p["id"] = str(p.pop("_id"))
    return payments

@api_router.delete("/credits/{credit_id}")
async def delete_credit(credit_id: str, request: Request):
    await get_current_user(request)
    await db.customer_credits.delete_one({"_id": ObjectId(credit_id)})
    return {"message": "Credit deleted"}

# ================== PURCHASE ORDERS (INBOUND) ==================

async def _next_doc_sequence(key: str) -> int:
    doc = await db.counters.find_one_and_update(
        {"_id": key}, {"$inc": {"seq": 1}}, upsert=True, return_document=ReturnDocument.AFTER,
    )
    return doc["seq"]

@api_router.post("/purchase-orders")
async def create_po(po: PurchaseOrderCreate, request: Request):
    user = await get_current_user(request)
    total = sum(i.get("qty",0) * i.get("unit_price",0) for i in po.items)
    location_id = po.location_id or user.get("default_location_id")
    loc_code = None
    if location_id:
        loc = await db.locations.find_one({"_id": ObjectId(location_id)})
        loc_code = loc.get("code") if loc else None
    seq = await _next_doc_sequence(f"po_number_{loc_code or 'HO'}")
    doc = {
        "po_number": f"PO-{loc_code or 'HO'}-{seq:04d}",
        "supplier_name": po.supplier_name, "supplier_contact": po.supplier_contact,
        "items": po.items, "total_amount": round(total, 2),
        "expected_delivery": po.expected_delivery, "notes": po.notes,
        "location_id": location_id,
        "status": "pending", "qc": None, "transport": None, "storage_location": None,
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.purchase_orders.insert_one(doc)
    return {"id": str(result.inserted_id), "po_number": doc["po_number"], "message": "PO created"}

@api_router.get("/purchase-orders")
async def list_pos(request: Request, status: str = None, location_id: str = None):
    user = await get_current_user(request)
    query = {}
    if status and status != "all": query["status"] = status
    loc_filter = location_scope_filter(user, location_id)
    if loc_filter: query.update(loc_filter)
    pos = await db.purchase_orders.find(query).sort("created_at", -1).to_list(500)
    for p in pos: p["id"] = str(p.pop("_id"))
    return pos

@api_router.delete("/purchase-orders/{po_id}")
async def delete_po(po_id: str, request: Request):
    """Iter 43 Change 2 — a PO can be deleted freely while still 'pending' (nothing physical
    happened yet); once approved/arrived/received it must go through the existing reverse-inbound flow."""
    user = await get_current_user(request)
    po = await db.purchase_orders.find_one({"_id": ObjectId(po_id)})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    if po.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Only a still-pending PO can be deleted — use the reversal flow once it has progressed")
    await db.purchase_orders.delete_one({"_id": ObjectId(po_id)})
    await create_audit_log(user["id"], user["name"], "delete", "purchase_order", po_id, {"supplier_name": po.get("supplier_name")}, None)
    return {"message": "Purchase order deleted"}

@api_router.put("/purchase-orders/{po_id}/approve")
async def approve_po(po_id: str, request: Request):
    user = await require_permission(request, "can_manage_company")
    await db.purchase_orders.update_one({"_id": ObjectId(po_id)}, {"$set": {"status": "approved", "approved_by": user["name"], "approved_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "PO approved"}

@api_router.put("/purchase-orders/{po_id}/arrival")
async def record_arrival(po_id: str, request: Request):
    body = await request.json()
    await get_current_user(request)
    transport = {"transporter": body.get("transporter",""), "vehicle": body.get("vehicle",""), "driver_contact": body.get("driver_contact",""), "lr_number": body.get("lr_number","")}
    await db.purchase_orders.update_one({"_id": ObjectId(po_id)}, {"$set": {"status": "arrived", "transport": transport, "arrived_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Arrival recorded"}

@api_router.put("/purchase-orders/{po_id}/qc")
async def record_qc(po_id: str, request: Request):
    body = await request.json()
    await get_current_user(request)
    qc = {"qty_check": body.get("qty_check", "pass"), "damage_check": body.get("damage_check", "pass"), "spec_match": body.get("spec_match", "pass"), "overall": body.get("overall", "pass"), "notes": body.get("notes", "")}
    await db.purchase_orders.update_one({"_id": ObjectId(po_id)}, {"$set": {"status": "qc_done", "qc": qc, "qc_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "QC recorded"}

@api_router.put("/purchase-orders/{po_id}/inbound")
async def complete_inbound(po_id: str, request: Request):
    body = await request.json()
    current_user = await get_current_user(request)
    if not await check_module_permission(current_user, "module_purchase_inbound", "create"):
        raise HTTPException(status_code=403, detail="You don't have permission to receive inventory against a purchase order")

    po = await db.purchase_orders.find_one({"_id": ObjectId(po_id)})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    if po.get("status") == "completed":
        raise HTTPException(status_code=400, detail="This PO has already been received. Use Edit to correct a quantity, or Reverse to undo it.")

    location = body.get("storage_location", "")
    now_iso = datetime.now(timezone.utc).isoformat()
    received_items, failed_lines, movements = [], [], []
    for item in po.get("items", []):
        inv_item = None
        if item.get("inventory_item_id"):
            inv_item = await db.inventory_items.find_one({"_id": ObjectId(item["inventory_item_id"])})
        if not inv_item and item.get("sku_code"):
            inv_item = await db.inventory_items.find_one({"sku_code": item["sku_code"]})
        if not inv_item:
            failed_lines.append(item.get("name", "Unknown item"))
            continue
        qty = float(item.get("qty", 0))
        await db.inventory_items.update_one({"_id": inv_item["_id"]}, {"$inc": {"quantity": qty}})
        received_items.append({
            "inventory_item_id": str(inv_item["_id"]), "name": inv_item["name"],
            "sku_code": inv_item.get("sku_code"), "qty_received": qty, "received_at": now_iso,
        })
        movements.append({
            "inventory_item_id": str(inv_item["_id"]), "movement_type": "purchase_inbound", "quantity": qty,
            "reference_type": "purchase_order", "reference_id": po_id,
            "note": f"PO inbound: {inv_item['name']}", "created_by": current_user["id"], "created_at": now_iso,
        })
    if failed_lines:
        raise HTTPException(status_code=400,
            detail=f"Could not match to an inventory item: {', '.join(failed_lines)}. Link each PO line to an inventory item before completing the inbound.")
    if movements:
        await db.inventory_movements.insert_many(movements)
    await db.purchase_orders.update_one({"_id": ObjectId(po_id)}, {"$set": {
        "status": "completed", "storage_location": location, "completed_at": now_iso,
        "received_items": received_items,
        "completed_by": current_user["id"], "completed_by_name": current_user["name"],
    }})
    await create_audit_log(current_user["id"], current_user["name"], "inbound_completed", "purchase_order", po_id,
                            None, received_items)
    return {"message": "Inbound completed, inventory updated", "received_items": received_items}


@api_router.put("/purchase-orders/{po_id}/inbound/edit")
async def edit_inbound(po_id: str, payload: InboundEditRequest, request: Request):
    """Correct a completed inbound's received quantities. Applies only the DELTA to stock."""
    current_user = await get_current_user(request)
    if not await check_module_permission(current_user, "module_purchase_inbound", "edit"):
        raise HTTPException(status_code=403, detail="You don't have permission to edit a completed inbound")

    po = await db.purchase_orders.find_one({"_id": ObjectId(po_id)})
    if not po or po.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Only a completed inbound can be edited")

    lines = payload.lines
    old_by_id = {r["inventory_item_id"]: r for r in po.get("received_items", [])}
    now_iso = datetime.now(timezone.utc).isoformat()
    new_received, deltas, movements = [], [], []

    for line in lines:
        inv_id = line.inventory_item_id
        new_qty = float(line.qty_received)
        old_entry = old_by_id.get(inv_id, {})
        old_qty = float(old_entry.get("qty_received", 0))
        delta = new_qty - old_qty
        inv_item = await db.inventory_items.find_one({"_id": ObjectId(inv_id)})
        if not inv_item:
            raise HTTPException(status_code=400, detail=f"Inventory item {inv_id} not found")
        if delta != 0:
            projected = inv_item.get("quantity", 0) + delta
            if projected < 0:
                raise HTTPException(status_code=400,
                    detail=f"Cannot reduce '{inv_item['name']}' by {abs(delta)} — only {inv_item.get('quantity', 0)} left in stock (some has already been issued or sold elsewhere)")
            await db.inventory_items.update_one({"_id": ObjectId(inv_id)}, {"$inc": {"quantity": delta}})
            deltas.append({"inventory_item_id": inv_id, "name": inv_item["name"], "old_qty": old_qty, "new_qty": new_qty, "delta": delta})
            movements.append({
                "inventory_item_id": inv_id, "movement_type": "purchase_inbound_edit", "quantity": delta,
                "reference_type": "purchase_order", "reference_id": po_id,
                "note": f"Inbound correction: {inv_item['name']} ({old_qty} → {new_qty})",
                "created_by": current_user["id"], "created_at": now_iso,
            })
        new_received.append({**old_entry, "inventory_item_id": inv_id, "name": inv_item["name"],
                              "sku_code": inv_item.get("sku_code"), "qty_received": new_qty,
                              "notes": line.notes if line.notes is not None else old_entry.get("notes"),
                              "received_at": old_entry.get("received_at", now_iso)})

    if movements:
        await db.inventory_movements.insert_many(movements)
    edit_entry = {"edited_by": current_user["name"], "edited_at": now_iso, "deltas": deltas,
                  "storage_location_changed": bool(payload.storage_location and payload.storage_location != po.get("storage_location"))}
    update_fields = {
        "received_items": new_received, "edited_at": now_iso, "edited_by": current_user["name"], "edited": True,
    }
    if payload.storage_location:
        update_fields["storage_location"] = payload.storage_location
    await db.purchase_orders.update_one({"_id": ObjectId(po_id)}, {
        "$set": update_fields,
        "$push": {"edit_history": edit_entry},
    })
    await create_audit_log(current_user["id"], current_user["name"], "inbound_edited", "purchase_order", po_id,
                            old_by_id, new_received, details=json.dumps(deltas))
    return {"message": "Inbound updated", "deltas": deltas}


@api_router.delete("/purchase-orders/{po_id}/inbound")
async def reverse_inbound(po_id: str, request: Request):
    """Undo a completed inbound. Admins act directly; others without delete rights are queued for admin approval."""
    current_user = await get_current_user(request)
    po = await db.purchase_orders.find_one({"_id": ObjectId(po_id)})
    if not po or po.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Only a completed inbound can be reversed")

    now_iso = datetime.now(timezone.utc).isoformat()
    can_delete = await check_module_permission(current_user, "module_purchase_inbound", "delete")
    if not can_delete:
        existing = await db.inbound_action_requests.find_one({"po_id": po_id, "status": "pending"})
        if existing:
            return {"status": "pending_approval", "message": "A reversal request for this PO is already awaiting admin approval"}
        await db.inbound_action_requests.insert_one({
            "po_id": po_id, "supplier_name": po.get("supplier_name"), "action": "reverse",
            "requested_by": current_user["id"], "requested_by_name": current_user["name"],
            "status": "pending", "requested_at": now_iso, "location_id": po.get("location_id"),
            "received_items_snapshot": po.get("received_items", []),
        })
        await create_audit_log(current_user["id"], current_user["name"], "inbound_reversal_requested", "purchase_order", po_id)
        return {"status": "pending_approval", "message": "You don't have permission to reverse a completed inbound — request sent to an admin for approval"}

    blocked = []
    for r in po.get("received_items", []):
        inv_item = await db.inventory_items.find_one({"_id": ObjectId(r["inventory_item_id"])})
        if inv_item and inv_item.get("quantity", 0) < r.get("qty_received", 0):
            blocked.append(f"{r['name']} — only {inv_item.get('quantity', 0)} left, {r.get('qty_received', 0)} was received (some already issued or sold)")
    if blocked:
        raise HTTPException(status_code=400, detail="Cannot reverse — stock has already moved: " + "; ".join(blocked))

    movements = []
    for r in po.get("received_items", []):
        await db.inventory_items.update_one({"_id": ObjectId(r["inventory_item_id"])}, {"$inc": {"quantity": -r.get("qty_received", 0)}})
        movements.append({
            "inventory_item_id": r["inventory_item_id"], "movement_type": "purchase_inbound_reversal",
            "quantity": -r.get("qty_received", 0), "reference_type": "purchase_order", "reference_id": po_id,
            "note": f"Inbound reversed: {r['name']}", "created_by": current_user["id"], "created_at": now_iso,
        })
    if movements:
        await db.inventory_movements.insert_many(movements)
    await db.purchase_orders.update_one({"_id": ObjectId(po_id)}, {
        "$set": {"status": "qc_done", "received_items": [], "reversed_at": now_iso, "reversed_by": current_user["name"]},
        "$unset": {"completed_at": ""},
    })
    await create_audit_log(current_user["id"], current_user["name"], "inbound_reversed", "purchase_order", po_id,
                            po.get("received_items"), None)
    return {"status": "reversed", "message": "Inbound reversed, stock reverted"}


# ── Inbound reversal approval queue (managers without delete rights → admin, location-scoped) ──
@api_router.get("/inbound-action-requests")
async def list_inbound_action_requests(request: Request, status: Optional[str] = None):
    user = await require_role("admin", "manager")(request)
    query = {"status": status} if status else {}
    docs = await db.inbound_action_requests.find(query).sort("requested_at", -1).to_list(200)
    docs = [d for d in docs if _can_manage_request_location(user, d.get("location_id"))]
    return [{**{k: v for k, v in d.items() if k != "_id"}, "id": str(d["_id"])} for d in docs]


@api_router.post("/inbound-action-requests/{req_id}/approve")
async def approve_inbound_action_request(req_id: str, request: Request):
    user = await require_role("admin", "manager")(request)
    doc = await db.inbound_action_requests.find_one({"_id": ObjectId(req_id)})
    if not doc or doc.get("status") != "pending":
        raise HTTPException(status_code=400, detail="This request is no longer pending")
    if not _can_manage_request_location(user, doc.get("location_id")):
        raise HTTPException(status_code=403, detail="This request belongs to a location outside your assignment")
    po = await db.purchase_orders.find_one({"_id": ObjectId(doc["po_id"])})
    if not po or po.get("status") != "completed":
        raise HTTPException(status_code=400, detail="PO is no longer in a completed state")

    now_iso = datetime.now(timezone.utc).isoformat()
    blocked = []
    for r in po.get("received_items", []):
        inv_item = await db.inventory_items.find_one({"_id": ObjectId(r["inventory_item_id"])})
        if inv_item and inv_item.get("quantity", 0) < r.get("qty_received", 0):
            blocked.append(f"{r['name']} — only {inv_item.get('quantity', 0)} left")
    if blocked:
        raise HTTPException(status_code=400, detail="Cannot reverse — stock has already moved: " + "; ".join(blocked))

    movements = []
    for r in po.get("received_items", []):
        await db.inventory_items.update_one({"_id": ObjectId(r["inventory_item_id"])}, {"$inc": {"quantity": -r.get("qty_received", 0)}})
        movements.append({
            "inventory_item_id": r["inventory_item_id"], "movement_type": "purchase_inbound_reversal",
            "quantity": -r.get("qty_received", 0), "reference_type": "purchase_order", "reference_id": doc["po_id"],
            "note": f"Inbound reversed (approved): {r['name']}", "created_by": user["id"], "created_at": now_iso,
        })
    if movements:
        await db.inventory_movements.insert_many(movements)
    await db.purchase_orders.update_one({"_id": ObjectId(doc["po_id"])}, {
        "$set": {"status": "qc_done", "received_items": [], "reversed_at": now_iso, "reversed_by": user["name"]},
        "$unset": {"completed_at": ""},
    })
    await db.inbound_action_requests.update_one({"_id": ObjectId(req_id)}, {"$set": {
        "status": "approved", "resolved_by": user["id"], "resolved_by_name": user["name"], "resolved_at": now_iso,
    }})
    await create_audit_log(user["id"], user["name"], "inbound_reversal_approved", "purchase_order", doc["po_id"])
    return {"message": "Reversal approved and applied"}


@api_router.post("/inbound-action-requests/{req_id}/reject")
async def reject_inbound_action_request(req_id: str, request: Request):
    user = await require_role("admin", "manager")(request)
    doc = await db.inbound_action_requests.find_one({"_id": ObjectId(req_id)})
    if not doc or doc.get("status") != "pending":
        raise HTTPException(status_code=400, detail="This request is no longer pending")
    if not _can_manage_request_location(user, doc.get("location_id")):
        raise HTTPException(status_code=403, detail="This request belongs to a location outside your assignment")
    r = await db.inbound_action_requests.update_one(
        {"_id": ObjectId(req_id), "status": "pending"},
        {"$set": {"status": "rejected", "resolved_by": user["id"], "resolved_by_name": user["name"],
                   "resolved_at": datetime.now(timezone.utc).isoformat()}})
    if r.matched_count == 0:
        raise HTTPException(status_code=400, detail="This request is no longer pending")
    return {"message": "Request rejected"}


# ── Generic action-request queue (delivery cancel, asset archive, sale cancel — Iter 42/43) ──
async def _apply_action_request(doc: Dict[str, Any]):
    rt, rid = doc["resource_type"], doc["resource_id"]
    now_iso = datetime.now(timezone.utc).isoformat()
    if rt == "delivery":
        await db.deliveries.update_one({"_id": ObjectId(rid)}, {"$set": {"status": "cancelled", "cancelled_at": now_iso}})
    elif rt == "asset":
        await db.assets.update_one({"_id": ObjectId(rid)}, {"$set": {"active": False, "status": "scrapped"}})
    elif rt == "sale":
        from sales import apply_sale_cancellation
        await apply_sale_cancellation(db, rid)
    else:
        raise HTTPException(status_code=400, detail=f"Unknown resource type: {rt}")


def _can_manage_request_location(user: Dict[str, Any], location_id: Optional[str]) -> bool:
    """Iter 43 Change 3c — managers may only approve/reject requests scoped to their assigned
    location(s); admins and legacy (locationless) requests remain unrestricted."""
    if user.get("role") == "admin" or not location_id:
        return True
    return location_id in (user.get("location_ids") or [])


@api_router.get("/action-requests")
async def list_action_requests(request: Request, status: Optional[str] = None, resource_type: Optional[str] = None):
    user = await require_role("admin", "manager")(request)
    query: Dict[str, Any] = {}
    if status: query["status"] = status
    if resource_type: query["resource_type"] = resource_type
    docs = await db.action_requests.find(query).sort("requested_at", -1).to_list(200)
    docs = [d for d in docs if _can_manage_request_location(user, d.get("location_id"))]
    return [{**{k: v for k, v in d.items() if k != "_id"}, "id": str(d["_id"])} for d in docs]


@api_router.post("/action-requests/{req_id}/approve")
async def approve_action_request(req_id: str, request: Request):
    user = await require_role("admin", "manager")(request)
    doc = await db.action_requests.find_one({"_id": ObjectId(req_id)})
    if not doc or doc.get("status") != "pending":
        raise HTTPException(status_code=400, detail="This request is no longer pending")
    if not _can_manage_request_location(user, doc.get("location_id")):
        raise HTTPException(status_code=403, detail="This request belongs to a location outside your assignment")
    await _apply_action_request(doc)
    await db.action_requests.update_one({"_id": ObjectId(req_id)}, {"$set": {
        "status": "approved", "resolved_by": user["id"], "resolved_by_name": user["name"],
        "resolved_at": datetime.now(timezone.utc).isoformat(),
    }})
    await create_audit_log(user["id"], user["name"], f"{doc['resource_type']}_{doc['action']}_approved", doc["resource_type"], doc["resource_id"])
    return {"message": "Request approved and applied"}


@api_router.post("/action-requests/{req_id}/reject")
async def reject_action_request(req_id: str, request: Request):
    user = await require_role("admin", "manager")(request)
    doc = await db.action_requests.find_one({"_id": ObjectId(req_id)})
    if not doc or doc.get("status") != "pending":
        raise HTTPException(status_code=400, detail="This request is no longer pending")
    if not _can_manage_request_location(user, doc.get("location_id")):
        raise HTTPException(status_code=403, detail="This request belongs to a location outside your assignment")
    r = await db.action_requests.update_one(
        {"_id": ObjectId(req_id), "status": "pending"},
        {"$set": {"status": "rejected", "resolved_by": user["id"], "resolved_by_name": user["name"],
                   "resolved_at": datetime.now(timezone.utc).isoformat()}})
    if r.matched_count == 0:
        raise HTTPException(status_code=400, detail="This request is no longer pending")
    return {"message": "Request rejected"}

# ================== DELIVERY OUTBOUND ==================

@api_router.post("/deliveries")
async def create_delivery(delivery: DeliveryOutboundCreate, request: Request):
    user = await get_current_user(request)
    doc = {
        "project_id": delivery.project_id, "customer_name": delivery.customer_name,
        "customer_address": delivery.customer_address, "customer_contact": delivery.customer_contact,
        "items": delivery.items, "transporter_name": delivery.transporter_name,
        "vehicle_number": delivery.vehicle_number, "driver_contact": delivery.driver_contact,
        "dispatch_date": delivery.dispatch_date, "delivery_date": delivery.delivery_date,
        "distance_km": delivery.distance_km, "notes": delivery.notes,
        "location_id": user.get("default_location_id"),
        "status": "dispatched", "created_by": user["id"], "created_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.deliveries.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Delivery created"}

@api_router.get("/deliveries")
async def list_deliveries(request: Request, status: str = None, location_id: str = None):
    user = await get_current_user(request)
    query = {}
    if status and status != "all": query["status"] = status
    loc_filter = location_scope_filter(user, location_id)
    if loc_filter:
        query.update(loc_filter)
    dels = await db.deliveries.find(query).sort("created_at", -1).to_list(500)
    for d in dels: d["id"] = str(d.pop("_id"))
    return dels

@api_router.put("/deliveries/{delivery_id}/complete")
async def complete_delivery(delivery_id: str, request: Request):
    await get_current_user(request)
    await db.deliveries.update_one({"_id": ObjectId(delivery_id)}, {"$set": {"status": "delivered", "completed_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Delivery completed"}


@api_router.put("/deliveries/{delivery_id}")
async def edit_delivery(delivery_id: str, payload: DeliveryOutboundEdit, request: Request):
    """Edit a delivery's items/transporter/dates. Dispatched deliveries editable by anyone with edit
    rights; delivered ones require admin + a stated reason, since changing history is unusual."""
    current_user = await get_current_user(request)
    if not await check_module_permission(current_user, "module_delivery_outbound", "edit"):
        raise HTTPException(status_code=403, detail="You don't have permission to edit deliveries")
    d = await db.deliveries.find_one({"_id": ObjectId(delivery_id)})
    if not d:
        raise HTTPException(status_code=404, detail="Delivery not found")
    if d.get("status") == "delivered":
        if current_user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Only an admin can edit a delivery that has already been delivered")
        if not payload.admin_reason:
            raise HTTPException(status_code=400, detail="Please state a reason for editing a delivered outbound — this is unusual and gets logged")
    elif d.get("status") != "dispatched":
        raise HTTPException(status_code=400, detail=f"Cannot edit a delivery that is '{d.get('status')}'")

    if payload.items is not None and d.get("project_id"):
        recon = await db.material_reconciliation.find_one({"project_id": d["project_id"]})
        if recon and not payload.confirm_reconciliation_impact:
            return {"status": "needs_confirmation",
                    "message": "This project's material reconciliation has already started — changing quantities here will change its 'issued' figures. Confirm to proceed."}

    now_iso = datetime.now(timezone.utc).isoformat()
    before = {k: d.get(k) for k in ("items", "transporter_name", "vehicle_number", "driver_contact", "dispatch_date", "delivery_date", "notes")}
    update_fields: Dict[str, Any] = {"edited": True, "edited_at": now_iso, "edited_by": current_user["name"]}
    for field in ("items", "transporter_name", "vehicle_number", "driver_contact", "dispatch_date", "delivery_date", "notes"):
        val = getattr(payload, field)
        if val is not None:
            update_fields[field] = val
    if d.get("status") == "delivered":
        update_fields["post_delivery_edit_reason"] = payload.admin_reason
    after = {k: update_fields.get(k, before.get(k)) for k in before}
    edit_entry = {"edited_by": current_user["name"], "edited_at": now_iso, "before": before, "after": after,
                  "reason": payload.admin_reason if d.get("status") == "delivered" else None}
    await db.deliveries.update_one({"_id": ObjectId(delivery_id)}, {"$set": update_fields, "$push": {"edit_history": edit_entry}})
    await create_audit_log(current_user["id"], current_user["name"], "delivery_edited", "delivery", delivery_id, before, after)
    return {"status": "updated", "message": "Delivery updated"}


@api_router.delete("/deliveries/{delivery_id}")
async def cancel_delivery(delivery_id: str, request: Request):
    """Cancel a dispatched-but-not-yet-delivered outbound. No stock is decremented on dispatch
    today, so cancellation has nothing to restore — it just voids the record."""
    current_user = await get_current_user(request)
    d = await db.deliveries.find_one({"_id": ObjectId(delivery_id)})
    if not d:
        raise HTTPException(status_code=404, detail="Delivery not found")
    if d.get("status") != "dispatched":
        raise HTTPException(status_code=400, detail="Only a dispatched (not yet delivered) outbound can be cancelled")

    can_delete = await check_module_permission(current_user, "module_delivery_outbound", "delete")
    now_iso = datetime.now(timezone.utc).isoformat()
    if not can_delete:
        existing = await db.action_requests.find_one({"resource_type": "delivery", "resource_id": delivery_id, "status": "pending"})
        if existing:
            return {"status": "pending_approval", "message": "A cancellation request for this delivery is already awaiting admin approval"}
        await db.action_requests.insert_one({
            "resource_type": "delivery", "resource_id": delivery_id, "action": "cancel",
            "requested_by": current_user["id"], "requested_by_name": current_user["name"],
            "status": "pending", "requested_at": now_iso, "location_id": d.get("location_id"),
            "snapshot": {k: v for k, v in d.items() if k != "_id"},
        })
        await create_audit_log(current_user["id"], current_user["name"], "delivery_cancel_requested", "delivery", delivery_id)
        return {"status": "pending_approval", "message": "You don't have permission to cancel this delivery — request sent to an admin for approval"}

    await db.deliveries.update_one({"_id": ObjectId(delivery_id)}, {"$set": {"status": "cancelled", "cancelled_at": now_iso, "cancelled_by": current_user["name"]}})
    await create_audit_log(current_user["id"], current_user["name"], "delivery_cancelled", "delivery", delivery_id, d.get("items"), None)
    return {"status": "cancelled", "message": "Delivery cancelled"}

# ================== BRAND RETURNS ==================

@api_router.post("/returns")
async def create_return(ret: BrandReturnCreate, request: Request):
    user = await get_current_user(request)
    doc = {
        "project_id": ret.project_id, "supplier_name": ret.supplier_name,
        "item_name": ret.item_name, "quantity": ret.quantity,
        "reason": ret.reason, "notes": ret.notes,
        "status": "pending", "created_by": user["id"], "created_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.brand_returns.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Return created"}

@api_router.get("/returns")
async def list_returns(request: Request, status: str = None):
    await get_current_user(request)
    query = {}
    if status and status != "all": query["status"] = status
    rets = await db.brand_returns.find(query).sort("created_at", -1).to_list(500)
    for r in rets: r["id"] = str(r.pop("_id"))
    return rets

@api_router.put("/returns/{return_id}/complete")
async def complete_return(return_id: str, request: Request):
    await get_current_user(request)
    await db.brand_returns.update_one({"_id": ObjectId(return_id)}, {"$set": {"status": "completed", "completed_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Return completed"}

# ================== WEEKLY AUDITS ==================

@api_router.post("/audits")
async def create_audit(audit: AuditCreate, request: Request):
    user = await get_current_user(request)
    doc = {
        "title": audit.title, "project_id": audit.project_id,
        "auditor_name": audit.auditor_name, "deadline": audit.deadline,
        "checklist": audit.checklist, "notes": audit.notes,
        "issues": [], "status": "open",
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.audits.insert_one(doc)
    return {"id": str(result.inserted_id), "message": "Audit created"}

@api_router.get("/audits")
async def list_audits(request: Request, status: str = None):
    await get_current_user(request)
    query = {}
    if status and status != "all": query["status"] = status
    audits = await db.audits.find(query).sort("created_at", -1).to_list(500)
    for a in audits: a["id"] = str(a.pop("_id"))
    return audits

@api_router.put("/audits/{audit_id}")
async def update_audit(audit_id: str, request: Request):
    body = await request.json()
    await get_current_user(request)
    update = {}
    if "checklist" in body: update["checklist"] = body["checklist"]
    if "issues" in body: update["issues"] = body["issues"]
    if "status" in body: update["status"] = body["status"]
    if "notes" in body: update["notes"] = body["notes"]
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.audits.update_one({"_id": ObjectId(audit_id)}, {"$set": update})
    return {"message": "Audit updated"}

@api_router.put("/audits/{audit_id}/issue")
async def add_audit_issue(audit_id: str, request: Request):
    body = await request.json()
    await get_current_user(request)
    issue = {"description": body.get("description",""), "severity": body.get("severity","medium"), "fix_deadline": body.get("fix_deadline",""), "status": "open", "created_at": datetime.now(timezone.utc).isoformat()}
    await db.audits.update_one({"_id": ObjectId(audit_id)}, {"$push": {"issues": issue}})
    return {"message": "Issue added"}

# ================== HEALTH CHECK ==================

@api_router.get("/")
async def root():
    return {"message": "Sensoper Solar Estimator API", "status": "healthy"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# ================== STARTUP EVENTS ==================

@app.on_event("startup")
async def startup_event():
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.login_attempts.create_index("identifier")
    await db.projects.create_index("created_by")
    await db.projects.create_index("status")
    await db.audit_logs.create_index("timestamp")
    await db.audit_logs.create_index("entity_type")
    await db.inventory_items.create_index("sku_code", unique=True)
    await db.inventory_items.create_index("category")
    await db.inventory_categories.create_index("slug", unique=True)
    await db.terms_conditions.create_index([("language", 1), ("is_active", 1)])
    await db.deletion_requests.create_index("project_id")
    await db.deletion_requests.create_index("status")
    await db.approvals.create_index("status")
    await db.approvals.create_index("type")
    await db.form_tabs.create_index("slug", unique=True)
    await db.form_tabs.create_index("order")
    await db.daily_updates.create_index("project_id")
    await db.daily_updates.create_index("created_at")
    await db.payments.create_index("project_id")
    await db.material_usage_logs.create_index("project_id")
    await db.account_entries.create_index([("entry_type", 1), ("entry_date", -1)])
    await db.account_entries.create_index("entry_date")
    await db.readings.create_index("status")
    await db.readings.create_index("start_date")
    await db.customer_credits.create_index("status")
    await db.purchase_orders.create_index("status")
    await db.deliveries.create_index("status")
    await db.brand_returns.create_index("status")
    await db.audits.create_index("status")
    await db.approvals.create_index("requested_by")
    await db.approvals.create_index("timestamp")
    await db.role_permissions.create_index("role_name", unique=True)
    
    # Seed admin user
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@sensoper.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@123")
    
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        hashed = hash_password(admin_password)
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hashed,
            "name": "System Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin user created: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
        logger.info(f"Admin password updated: {admin_email}")
    
    # Update existing company profiles with new logo
    await db.company_profiles.update_many(
        {"logo_url": {"$regex": "job_solar-estimator|32se8qpu_snspr|x3rj9e2a_slg"}},
        {"$set": {"logo_url": "https://customer-assets.emergentagent.com/job_8c20414a-b147-464e-9c68-aaa2fa40fdbf/artifacts/q52gayft_snspr.png"}}
    )
    
    # Seed default company profile
    company = await db.company_profiles.find_one({})
    if not company:
        await db.company_profiles.insert_one({
            "company_name": "Sensoper Controls & Renewables",
            "tagline": "Solar Solutions Provider",
            "logo_url": "https://customer-assets.emergentagent.com/job_8c20414a-b147-464e-9c68-aaa2fa40fdbf/artifacts/q52gayft_snspr.png",
            "primary_color": "#4ADE40",
            "secondary_color": "#2D9BF0",
            "address": "123 Solar Street, Erode\nTamil Nadu, India - 638001",
            "phone": "+91 98765 43210",
            "email": "info@sensoper.com",
            "website": "www.sensoper.com",
            "gst_number": "33XXXXX1234X1ZX",
            "pan_number": "XXXXX1234X",
            "bank_details": {
                "account_name": "Sensoper Controls & Renewables",
                "account_number": "1234567890123456",
                "ifsc_code": "SBIN0001234",
                "bank_name": "State Bank of India",
                "branch": "Erode Main Branch"
            },
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info("Default company profile created")
    
    # Seed default inventory categories
    default_categories = [
        {"name": "Solar Panels", "slug": "solar_panels", "description": "Photovoltaic solar panels"},
        {"name": "Inverters", "slug": "inverters", "description": "Solar inverters and micro-inverters"},
        {"name": "Batteries", "slug": "batteries", "description": "Battery storage systems"},
        {"name": "Mounting Structures", "slug": "mounting_structures", "description": "Panel mounting and racking"},
        {"name": "Cables & Accessories", "slug": "cables_accessories", "description": "Wiring, connectors, and accessories"},
    ]
    for cat in default_categories:
        existing_cat = await db.inventory_categories.find_one({"slug": cat["slug"]})
        if not existing_cat:
            await db.inventory_categories.insert_one({**cat, "created_at": datetime.now(timezone.utc).isoformat()})
    logger.info("Inventory categories seeded")
    
    # Seed default permissions (and merge in any newly added module keys to existing roles)
    for role_name, perms in DEFAULT_PERMISSIONS.items():
        existing_perm = await db.role_permissions.find_one({"role_name": role_name})
        if not existing_perm:
            await db.role_permissions.insert_one({
                "role_name": role_name,
                "permissions": perms,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        else:
            existing_keys = set((existing_perm.get("permissions") or {}).keys())
            missing = {k: v for k, v in perms.items() if k not in existing_keys}
            if missing:
                merged = {**(existing_perm.get("permissions") or {}), **missing}
                await db.role_permissions.update_one(
                    {"role_name": role_name},
                    {"$set": {"permissions": merged, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
    logger.info("Default permissions seeded")
    
    # Seed system form tabs
    system_tabs = [
        {"name": "Customer", "slug": "customer", "system": True, "active": True, "icon": "User", "fields": [], "roles_visible": ["admin", "manager", "staff"]},
        {"name": "Location", "slug": "location", "system": True, "active": True, "icon": "MapPin", "fields": [], "roles_visible": ["admin", "manager", "staff"]},
        {"name": "Site & Electrical", "slug": "site_electrical", "system": True, "active": True, "icon": "Zap", "fields": [], "roles_visible": ["admin", "manager", "staff"]},
        {"name": "Materials", "slug": "materials", "system": True, "active": True, "icon": "Package", "fields": [], "roles_visible": ["admin", "manager", "staff"]},
        {"name": "Site Docs", "slug": "site_docs", "system": True, "active": True, "icon": "FolderOpen", "fields": [], "roles_visible": ["admin", "manager", "staff"]},
    ]
    for st in system_tabs:
        existing_tab = await db.form_tabs.find_one({"slug": st["slug"]})
        if not existing_tab:
            await db.form_tabs.insert_one({**st, "order": 0, "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()})
    # Normalize ordering: system tabs first (in defined order), custom tabs next, site_docs last
    system_order = ["customer", "location", "site_electrical", "materials"]
    all_tabs_list = await db.form_tabs.find().to_list(200)
    ordered = []
    for slug in system_order:
        tab = next((t for t in all_tabs_list if t.get("slug") == slug), None)
        if tab:
            ordered.append(tab)
    custom_tabs_sorted = sorted([t for t in all_tabs_list if not t.get("system")], key=lambda x: x.get("order", 999))
    ordered.extend(custom_tabs_sorted)
    site_docs = next((t for t in all_tabs_list if t.get("slug") == "site_docs"), None)
    if site_docs:
        ordered.append(site_docs)
    for idx, tab in enumerate(ordered):
        await db.form_tabs.update_one({"_id": tab["_id"]}, {"$set": {"order": idx + 1}})
    logger.info("System form tabs seeded")
    
    # Init object storage
    try:
        init_storage()
        logger.info("Object storage initialized")
    except Exception as e:
        logger.warning(f"Object storage init failed (non-critical): {e}")
    
    # Write test credentials
    try:
        os.makedirs("/app/memory", exist_ok=True)
        with open("/app/memory/test_credentials.md", "w") as f:
            f.write(f"""# Test Credentials

## Admin Account
- Email: {admin_email}
- Password: {admin_password}
- Role: admin

## Auth Endpoints
- POST /api/auth/login
- POST /api/auth/register
- POST /api/auth/logout
- GET /api/auth/me
- POST /api/auth/refresh
""")
    except Exception as e:
        logger.warning(f"Could not write test credentials: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# Include the router in the main app
# ================== ACCOUNTS (Cash on Hand, Meter Readings, Account Balance) ==================

class AccountEntryCreate(BaseModel):
    entry_type: str  # cash_on_hand | account_balance | operational_expense | marketing_expense | gst_input | gst_paid
    entry_date: str  # ISO date YYYY-MM-DD
    amount: float
    description: Optional[str] = ""
    # Marketing-expense-only attribution fields (Iter 39 Change 3)
    marketing_channel: Optional[str] = None       # google_ads|meta|referral|organic|hoardings|local_events|whatsapp|print|tv_radio|other
    campaign_name: Optional[str] = None
    target_district: Optional[str] = None

class AccountEntryUpdate(BaseModel):
    entry_type: Optional[str] = None
    entry_date: Optional[str] = None
    amount: Optional[float] = None
    description: Optional[str] = None
    marketing_channel: Optional[str] = None
    campaign_name: Optional[str] = None
    target_district: Optional[str] = None

ACCOUNT_TYPES = {"cash_on_hand", "account_balance", "operational_expense", "marketing_expense", "gst_input", "gst_paid"}

@api_router.get("/accounts")
async def list_accounts(request: Request, entry_type: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    user = await get_current_user(request)
    if user["role"] not in ["admin", "manager", "staff"]:
        raise HTTPException(status_code=403, detail="Forbidden")
    q = {}
    if entry_type:
        q["entry_type"] = entry_type
    if date_from or date_to:
        q["entry_date"] = {}
        if date_from: q["entry_date"]["$gte"] = date_from
        if date_to: q["entry_date"]["$lte"] = date_to
    cursor = db.account_entries.find(q).sort("entry_date", -1).limit(2000)
    docs = await cursor.to_list(2000)
    return [{
        "id": str(d["_id"]),
        "entry_type": d.get("entry_type"),
        "entry_date": d.get("entry_date"),
        "amount": d.get("amount", 0),
        "description": d.get("description", ""),
        "marketing_channel": d.get("marketing_channel"),
        "campaign_name": d.get("campaign_name"),
        "target_district": d.get("target_district"),
        "entered_by_id": d.get("entered_by_id"),
        "entered_by": d.get("entered_by", ""),
        "created_at": d.get("created_at")
    } for d in docs]

@api_router.post("/accounts")
async def create_account_entry(entry: AccountEntryCreate, request: Request):
    user = await get_current_user(request)
    if entry.entry_type not in ACCOUNT_TYPES:
        raise HTTPException(status_code=400, detail=f"entry_type must be one of {sorted(ACCOUNT_TYPES)}")
    doc = {
        "entry_type": entry.entry_type,
        "entry_date": entry.entry_date,
        "amount": float(entry.amount),
        "description": entry.description or "",
        "entered_by_id": user["id"],
        "entered_by": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    if entry.entry_type == "marketing_expense":
        doc["marketing_channel"] = entry.marketing_channel or "other"
        doc["campaign_name"] = entry.campaign_name or ""
        doc["target_district"] = entry.target_district or ""
    result = await db.account_entries.insert_one(doc)
    await create_audit_log(user["id"], user.get("name",""), "create", "account_entry", str(result.inserted_id), None, doc)
    return {"id": str(result.inserted_id), "message": "Account entry created"}

@api_router.put("/accounts/{entry_id}")
async def update_account_entry(entry_id: str, updates: AccountEntryUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.account_entries.find_one({"_id": ObjectId(entry_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Entry not found")
    if user["role"] == "staff" and existing.get("entered_by_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Staff can only edit their own entries")
    update_data = {k: v for k, v in updates.dict(exclude_unset=True).items() if v is not None}
    if "entry_type" in update_data and update_data["entry_type"] not in ACCOUNT_TYPES:
        raise HTTPException(status_code=400, detail="Invalid entry_type")
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.account_entries.update_one({"_id": ObjectId(entry_id)}, {"$set": update_data})
    await create_audit_log(user["id"], user.get("name",""), "update", "account_entry", entry_id, None, update_data)
    return {"message": "Account entry updated"}

@api_router.delete("/accounts/{entry_id}")
async def delete_account_entry(entry_id: str, request: Request):
    user = await get_current_user(request)
    existing = await db.account_entries.find_one({"_id": ObjectId(entry_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Entry not found")
    if user["role"] not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admin/manager can delete account entries")
    await db.account_entries.delete_one({"_id": ObjectId(entry_id)})
    await create_audit_log(user["id"], user.get("name",""), "delete", "account_entry", entry_id, existing, None)
    return {"message": "Deleted"}

@api_router.get("/accounts/summary")
async def accounts_summary(request: Request):
    """Latest snapshot per account type + month-to-date totals for expense types — used by Accounts page + CEO Dashboard."""
    await get_current_user(request)
    summary = {}
    for t in ACCOUNT_TYPES:
        latest = await db.account_entries.find_one({"entry_type": t}, sort=[("entry_date", -1), ("created_at", -1)])
        if latest:
            summary[t] = {
                "amount": latest.get("amount", 0),
                "entry_date": latest.get("entry_date"),
                "description": latest.get("description", ""),
                "entered_by": latest.get("entered_by", ""),
                "updated_at": latest.get("updated_at") or latest.get("created_at")
            }
        else:
            summary[t] = {"amount": 0, "entry_date": None, "description": "", "entered_by": "", "updated_at": None}
    # Month-to-date totals for expense / gst types
    today = datetime.now(timezone.utc)
    month_start = today.strftime("%Y-%m-01")
    for t in ("operational_expense", "gst_input", "gst_paid"):
        cursor = db.account_entries.find({"entry_type": t, "entry_date": {"$gte": month_start}})
        total = 0.0
        async for d in cursor:
            total += float(d.get("amount", 0))
        summary[f"{t}_mtd"] = round(total, 2)
    summary["gst_net_mtd"] = round(summary.get("gst_paid_mtd", 0) - summary.get("gst_input_mtd", 0), 2)
    return summary


# ================== READINGS (Site reading-phase tracker) ==================

class ReadingCreate(BaseModel):
    site_name: str
    site_ref: Optional[str] = ""
    site_address: Optional[str] = ""
    device_id: Optional[str] = ""
    device_type: Optional[str] = ""
    device_serial: Optional[str] = ""
    customer_name: Optional[str] = ""
    customer_phone: Optional[str] = ""
    customer_account: Optional[str] = ""
    start_date: str  # YYYY-MM-DD
    days: int = 30
    status: str = "active"  # active | completed | overdue
    notes: Optional[str] = ""

class ReadingUpdate(BaseModel):
    site_name: Optional[str] = None
    site_ref: Optional[str] = None
    site_address: Optional[str] = None
    device_id: Optional[str] = None
    device_type: Optional[str] = None
    device_serial: Optional[str] = None
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_account: Optional[str] = None
    start_date: Optional[str] = None
    days: Optional[int] = None
    status: Optional[str] = None
    notes: Optional[str] = None

READING_STATUSES = {"active", "completed", "overdue"}

def _compute_end_date(start_date_str: str, days: int) -> str:
    try:
        d = datetime.strptime(start_date_str, "%Y-%m-%d")
    except (ValueError, TypeError):
        return ""
    return (d + timedelta(days=int(days or 0))).strftime("%Y-%m-%d")

def _serialise_reading(d: dict) -> dict:
    end_date = _compute_end_date(d.get("start_date", ""), d.get("days", 0))
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    raw_status = d.get("status", "active")
    # Auto-classify overdue (without persisting) when end date passed but still active
    derived_status = raw_status
    if raw_status == "active" and end_date and end_date < today:
        derived_status = "overdue"
    return {
        "id": str(d["_id"]),
        "site_name": d.get("site_name", ""),
        "site_ref": d.get("site_ref", ""),
        "site_address": d.get("site_address", ""),
        "device_id": d.get("device_id", ""),
        "device_type": d.get("device_type", ""),
        "device_serial": d.get("device_serial", ""),
        "customer_name": d.get("customer_name", ""),
        "customer_phone": d.get("customer_phone", ""),
        "customer_account": d.get("customer_account", ""),
        "start_date": d.get("start_date", ""),
        "days": d.get("days", 0),
        "end_date": end_date,
        "status": derived_status,
        "notes": d.get("notes", ""),
        "created_by": d.get("created_by", ""),
        "created_at": d.get("created_at"),
        "updated_at": d.get("updated_at")
    }

@api_router.get("/readings")
async def list_readings(request: Request, status: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    user = await get_current_user(request)
    if user["role"] not in ["admin", "manager", "staff"]:
        raise HTTPException(status_code=403, detail="Forbidden")
    q = {}
    if date_from or date_to:
        q["start_date"] = {}
        if date_from: q["start_date"]["$gte"] = date_from
        if date_to: q["start_date"]["$lte"] = date_to
    docs = await db.readings.find(q).sort("start_date", -1).to_list(2000)
    rows = [_serialise_reading(d) for d in docs]
    if status and status != "all":
        rows = [r for r in rows if r["status"] == status]
    return rows

@api_router.get("/readings/summary")
async def readings_summary(request: Request):
    await get_current_user(request)
    docs = await db.readings.find({}).to_list(5000)
    rows = [_serialise_reading(d) for d in docs]
    return {
        "total": len(rows),
        "active": sum(1 for r in rows if r["status"] == "active"),
        "completed": sum(1 for r in rows if r["status"] == "completed"),
        "overdue": sum(1 for r in rows if r["status"] == "overdue")
    }

@api_router.post("/readings")
async def create_reading(reading: ReadingCreate, request: Request):
    user = await get_current_user(request)
    if reading.status not in READING_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    doc = reading.dict()
    doc["created_by"] = user.get("name", "")
    doc["created_by_id"] = user["id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["updated_at"] = doc["created_at"]
    result = await db.readings.insert_one(doc)
    await create_audit_log(user["id"], user.get("name",""), "create", "reading", str(result.inserted_id), None, doc)
    fresh = await db.readings.find_one({"_id": result.inserted_id})
    return _serialise_reading(fresh)

@api_router.put("/readings/{reading_id}")
async def update_reading(reading_id: str, updates: ReadingUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.readings.find_one({"_id": ObjectId(reading_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Reading not found")
    update_data = {k: v for k, v in updates.dict(exclude_unset=True).items() if v is not None}
    if "status" in update_data and update_data["status"] not in READING_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.readings.update_one({"_id": ObjectId(reading_id)}, {"$set": update_data})
    await create_audit_log(user["id"], user.get("name",""), "update", "reading", reading_id, None, update_data)
    fresh = await db.readings.find_one({"_id": ObjectId(reading_id)})
    return _serialise_reading(fresh)

@api_router.delete("/readings/{reading_id}")
async def delete_reading(reading_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admin/manager can delete readings")
    existing = await db.readings.find_one({"_id": ObjectId(reading_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Reading not found")
    await db.readings.delete_one({"_id": ObjectId(reading_id)})
    await create_audit_log(user["id"], user.get("name",""), "delete", "reading", reading_id, existing, None)
    return {"message": "Deleted"}



# ============================================================
# ============ TNEB / SOLAR REPORT MODULE ====================
# ============================================================
# Pluggable bill-fetch provider. Set BILL_FETCH_PROVIDER + BILL_FETCH_API_KEY
# in backend/.env. Currently supports: setu, decentro, signzy (placeholders).
# When unset, returns success=False so frontend falls into manual entry mode.

import math
from io import BytesIO
from pypdf import PdfReader, PdfWriter
from fastapi.responses import StreamingResponse

class TnebFetchRequest(BaseModel):
    service_number: str
    phone: str

class TnebConsumerData(BaseModel):
    consumer_name: Optional[str] = None
    address: Optional[str] = None
    sanctioned_load_kw: Optional[float] = None
    avg_monthly_consumption: Optional[float] = None
    avg_monthly_bill: Optional[float] = None
    tariff_category: Optional[str] = None  # Domestic / Commercial / Industrial
    connection_type: Optional[str] = None  # Single Phase / Three Phase
    historical_12m: Optional[List[dict]] = None  # [{"month":"2025-01","units":350,"amount":2100}]

@api_router.post("/tneb/fetch")
async def tneb_fetch(payload: TnebFetchRequest, current_user: dict = Depends(get_current_user)):
    """Try a configured 3rd-party bill-fetch provider. Falls back to manual entry."""
    # Basic format validation
    svc = (payload.service_number or "").strip()
    phone = (payload.phone or "").strip()
    if not svc or len(svc) < 6:
        raise HTTPException(status_code=400, detail="Invalid TNEB service number format")
    if not phone.isdigit() or len(phone) != 10:
        raise HTTPException(status_code=400, detail="Phone must be a 10-digit Indian mobile number")

    provider = os.environ.get("BILL_FETCH_PROVIDER", "").lower().strip()
    api_key = os.environ.get("BILL_FETCH_API_KEY", "").strip()

    if not provider or not api_key:
        return {
            "success": False,
            "fallback": "manual",
            "message": "TNEB live fetch not configured. Please enter consumer details manually.",
            "data": None,
        }

    # Provider-specific call (stubs — wire real endpoints once user provides key)
    try:
        if provider == "setu":
            # Setu BBPS-style bill fetch (example contract — exact contract per Setu docs)
            r = http_requests.post(
                "https://uat.setu.co/api/bill-fetch/v1/fetch",
                headers={"x-api-key": api_key, "Content-Type": "application/json"},
                json={"biller_id": "TNEB", "service_number": svc, "phone": phone},
                timeout=20,
            )
            r.raise_for_status()
            payload_raw = r.json()
        elif provider == "decentro":
            r = http_requests.post(
                "https://in.decentro.tech/v2/billers/fetch",
                headers={"client_id": api_key.split(":")[0], "client_secret": api_key.split(":")[-1]},
                json={"biller": "TNEB", "consumer_id": svc, "mobile": phone},
                timeout=20,
            )
            r.raise_for_status()
            payload_raw = r.json()
        elif provider == "signzy":
            r = http_requests.post(
                "https://preproduction.signzy.tech/api/v2/patrons/utility/electricity",
                headers={"Authorization": api_key},
                json={"board": "TNEB", "consumerNumber": svc, "phone": phone},
                timeout=20,
            )
            r.raise_for_status()
            payload_raw = r.json()
        else:
            return {"success": False, "fallback": "manual",
                    "message": f"Unknown provider '{provider}'. Set BILL_FETCH_PROVIDER to setu|decentro|signzy.",
                    "data": None}

        # Normalize provider response → TnebConsumerData. Real shapes will need
        # tuning once a contract is locked; we map best-effort common keys.
        d = payload_raw.get("data") or payload_raw.get("result") or payload_raw
        normalized = {
            "consumer_name": d.get("customerName") or d.get("consumer_name") or d.get("name"),
            "address": d.get("address") or d.get("billingAddress"),
            "sanctioned_load_kw": float(d.get("sanctionedLoad") or d.get("sanctioned_load") or 0) or None,
            "avg_monthly_consumption": float(d.get("avgUnits") or d.get("units") or 0) or None,
            "avg_monthly_bill": float(d.get("amount") or d.get("billAmount") or 0) or None,
            "tariff_category": d.get("tariff") or d.get("tariffCategory"),
            "connection_type": d.get("phase") or d.get("connectionType"),
            "historical_12m": d.get("history") or None,
        }
        return {"success": True, "fallback": None,
                "message": f"Fetched from {provider}.",
                "data": normalized, "provider": provider}
    except http_requests.RequestException as e:
        logger.warning(f"TNEB provider {provider} error: {e}")
        return {"success": False, "fallback": "manual",
                "message": f"Provider error: {str(e)[:120]}. Please enter details manually.",
                "data": None}


@api_router.get("/solar/irradiation")
async def solar_irradiation(lat: float, lng: float, current_user: dict = Depends(get_current_user)):
    """Fetch annual avg daily solar irradiation (kWh/m²/day) from NASA POWER. Free, no key."""
    if not (-90 <= lat <= 90 and -180 <= lng <= 180):
        raise HTTPException(status_code=400, detail="Invalid coordinates")
    try:
        r = http_requests.get(
            "https://power.larc.nasa.gov/api/temporal/climatology/point",
            params={
                "parameters": "ALLSKY_SFC_SW_DWN",
                "community": "RE",
                "longitude": lng,
                "latitude": lat,
                "format": "JSON",
            },
            timeout=20,
        )
        r.raise_for_status()
        data = r.json()
        monthly = data.get("properties", {}).get("parameter", {}).get("ALLSKY_SFC_SW_DWN", {}) or {}
        # Annual is keyed as "ANN" in NASA POWER climatology
        annual = monthly.get("ANN")
        if annual is None or annual < 0:
            # Fallback to India avg
            annual = 5.0
        # Strip ANN from monthly for cleaner response
        months = {k: v for k, v in monthly.items() if k != "ANN" and isinstance(v, (int, float)) and v >= 0}
        return {
            "annual_avg_kwh_m2_day": round(float(annual), 3),
            "monthly_kwh_m2_day": months,
            "source": "NASA POWER",
            "lat": lat,
            "lng": lng,
        }
    except http_requests.RequestException as e:
        logger.warning(f"NASA POWER fetch failed: {e}. Using India avg fallback.")
        return {
            "annual_avg_kwh_m2_day": 5.0,
            "monthly_kwh_m2_day": {},
            "source": "fallback (India avg)",
            "lat": lat,
            "lng": lng,
            "warning": "NASA POWER unreachable; used 5.0 kWh/m²/day fallback.",
        }


class SolarSizingRequest(BaseModel):
    monthly_consumption_units: float
    sanctioned_load_kw: Optional[float] = None
    tariff_category: str = "Domestic"  # Domestic / Commercial / Industrial
    connection_type: str = "Single Phase"
    avg_monthly_bill: Optional[float] = None
    irradiation_kwh_m2_day: float = 5.0  # India avg fallback
    system_type: Literal["on-grid", "off-grid", "hybrid", "solar-pump"] = "on-grid"
    panel_wattage_w: int = 550
    cost_per_kwp: float = 55000  # ₹/kWp installed (2026 India avg residential on-grid)
    battery_autonomy_days: float = 1.0
    battery_voltage: int = 48
    state: Optional[str] = "Tamil Nadu"

@api_router.post("/solar/sizing")
async def solar_sizing(req: SolarSizingRequest, current_user: dict = Depends(get_current_user)):
    """Pure calculator: produces full solar system sizing + 25-year financial projection."""
    PR = 0.75  # Performance ratio (typical India)
    DEGRADATION_PCT = 0.7  # per year
    TARIFF_ESCALATION_PCT = 2.5  # per year
    GRID_EF_KG_PER_KWH = 0.82  # India grid CO2 emission factor

    monthly_units = max(req.monthly_consumption_units, 1)
    irradiation = max(req.irradiation_kwh_m2_day, 0.1)

    # Tariff (₹/unit) — infer if bill provided else use category default
    if req.avg_monthly_bill and req.avg_monthly_bill > 0:
        tariff_per_unit = req.avg_monthly_bill / monthly_units
    else:
        tariff_per_unit = {"Domestic": 6.5, "Commercial": 9.0, "Industrial": 8.0}.get(req.tariff_category, 7.0)

    # System sizing
    daily_units_needed = monthly_units / 30.0
    kwp_needed = daily_units_needed / (irradiation * PR)
    # Round up to nearest 0.5 kWp; cap at sanctioned load if provided
    kwp_recommended = math.ceil(kwp_needed * 2) / 2
    if req.sanctioned_load_kw and kwp_recommended > req.sanctioned_load_kw:
        kwp_recommended = req.sanctioned_load_kw

    num_panels = math.ceil((kwp_recommended * 1000) / req.panel_wattage_w)
    inverter_capacity_kw = round(kwp_recommended * 1.1, 2)

    # Battery (off-grid/hybrid only)
    battery_ah = 0
    if req.system_type in ("off-grid", "hybrid"):
        daily_wh = daily_units_needed * 1000
        usable_wh = daily_wh * req.battery_autonomy_days / 0.85  # 85% round-trip efficiency
        battery_ah = math.ceil(usable_wh / req.battery_voltage)

    # Monthly generation & savings
    monthly_generation = kwp_recommended * irradiation * 30 * PR
    monthly_savings = monthly_generation * tariff_per_unit
    annual_savings = monthly_savings * 12

    # Cost & subsidy (PM Surya Ghar — residential domestic on-grid only, India 2026)
    total_cost = round(kwp_recommended * req.cost_per_kwp)
    subsidy = 0
    if req.tariff_category == "Domestic" and req.system_type == "on-grid":
        if kwp_recommended <= 1:
            subsidy = 30000
        elif kwp_recommended <= 2:
            subsidy = 60000
        else:
            subsidy = 78000  # cap for ≥3 kW
    net_cost = max(total_cost - subsidy, 0)

    # Payback (simple)
    payback_years = round(net_cost / annual_savings, 2) if annual_savings > 0 else None

    # 25-year savings projection (with tariff escalation + panel degradation)
    yearly_breakdown = []
    cumulative = 0
    for year in range(1, 26):
        degraded_gen = monthly_generation * 12 * ((1 - DEGRADATION_PCT / 100) ** (year - 1))
        escalated_tariff = tariff_per_unit * ((1 + TARIFF_ESCALATION_PCT / 100) ** (year - 1))
        year_savings = degraded_gen * escalated_tariff
        cumulative += year_savings
        yearly_breakdown.append({
            "year": year,
            "generation_units": round(degraded_gen, 1),
            "tariff": round(escalated_tariff, 2),
            "savings": round(year_savings, 2),
            "cumulative": round(cumulative, 2),
        })
    total_25yr_savings = round(cumulative)
    roi_pct = round(((total_25yr_savings - net_cost) / net_cost) * 100, 1) if net_cost > 0 else None

    # Technical KPIs
    annual_generation = monthly_generation * 12
    cuf_pct = round((annual_generation / (kwp_recommended * 8760)) * 100, 2) if kwp_recommended > 0 else 0
    co2_offset_kg = round(annual_generation * GRID_EF_KG_PER_KWH)

    return {
        "sizing": {
            "kwp_recommended": round(kwp_recommended, 2),
            "num_panels": num_panels,
            "panel_wattage_w": req.panel_wattage_w,
            "inverter_capacity_kw": inverter_capacity_kw,
            "battery_ah": battery_ah,
            "battery_voltage": req.battery_voltage if battery_ah else 0,
        },
        "financials": {
            "tariff_per_unit": round(tariff_per_unit, 2),
            "total_cost": total_cost,
            "subsidy": subsidy,
            "net_cost": net_cost,
            "monthly_generation_units": round(monthly_generation, 1),
            "monthly_savings": round(monthly_savings, 2),
            "annual_savings": round(annual_savings, 2),
            "payback_years": payback_years,
            "roi_pct": roi_pct,
            "total_25yr_savings": total_25yr_savings,
            "yearly_breakdown": yearly_breakdown,
        },
        "technical": {
            "performance_ratio": PR,
            "cuf_pct": cuf_pct,
            "annual_generation_units": round(annual_generation),
            "co2_offset_kg_per_year": co2_offset_kg,
            "irradiation_kwh_m2_day": round(irradiation, 3),
            "degradation_pct_per_year": DEGRADATION_PCT,
        },
        "assumptions": {
            "tariff_escalation_pct": TARIFF_ESCALATION_PCT,
            "grid_emission_factor_kg_per_kwh": GRID_EF_KG_PER_KWH,
            "cost_per_kwp_inr": req.cost_per_kwp,
            "subsidy_scheme": "PM Surya Ghar (residential domestic on-grid only)",
        },
    }


@api_router.post("/solar/merge-pdf")
async def solar_merge_pdf(
    generated_pdf: UploadFile = File(...),
    uploaded_pdf: UploadFile = File(...),
    position: Literal["prepend", "append"] = Form("prepend"),
    current_user: dict = Depends(get_current_user),
):
    """Merge frontend-generated report PDF with user-uploaded PDF. Returns merged PDF stream."""
    try:
        gen_bytes = await generated_pdf.read()
        up_bytes = await uploaded_pdf.read()
        if not gen_bytes or not up_bytes:
            raise HTTPException(status_code=400, detail="Both PDF files are required and must be non-empty")

        # Size sanity check (50MB max each)
        if len(gen_bytes) > 50 * 1024 * 1024 or len(up_bytes) > 50 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="PDF file too large (max 50MB each)")

        writer = PdfWriter()
        gen_reader = PdfReader(BytesIO(gen_bytes))
        up_reader = PdfReader(BytesIO(up_bytes))
        first, second = (gen_reader, up_reader) if position == "prepend" else (up_reader, gen_reader)
        for pg in first.pages:
            writer.add_page(pg)
        for pg in second.pages:
            writer.add_page(pg)

        out = BytesIO()
        writer.write(out)
        out.seek(0)

        filename = (uploaded_pdf.filename or "report").rsplit(".", 1)[0] + "_merged.pdf"
        return StreamingResponse(
            out,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"PDF merge failed: {e}")
        raise HTTPException(status_code=500, detail=f"PDF merge failed: {str(e)[:200]}")


# Include the router in the main app
app.include_router(api_router)

# CORS Configuration
frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[frontend_url, "http://localhost:3000", "https://solar-ops-management.preview.emergentagent.com"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# Include the router in the main app
app.include_router(api_router)

# CORS Configuration
frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[frontend_url, "http://localhost:3000", "https://solar-ops-management.preview.emergentagent.com"],
    allow_methods=["*"],
    allow_headers=["*"],
)
